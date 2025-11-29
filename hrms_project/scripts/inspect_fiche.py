import os
from openpyxl import load_workbook
import glob

exports_dir = os.path.join(os.getcwd(), 'exports')
pattern = os.path.join(exports_dir, 'fiche_*.xlsx')
files = sorted(glob.glob(pattern), key=os.path.getmtime, reverse=True)
if not files:
    print('No fiche_*.xlsx found in exports/')
    raise SystemExit(1)
file = files[0]
print('Inspecting', file)
wb = load_workbook(file, data_only=True)
ws = wb['FICHE'] if 'FICHE' in wb.sheetnames else wb.active
cells = ['B16','B17','B18','B19','B20','J16','H26','H27','H29','J52','M40']
for c in cells:
    val = None
    try:
        val = ws[c].value
    except Exception:
        try:
            val = ws.cell(row=int(c[1:]), column=ord(c[0])-64).value
        except Exception:
            val = '<unreadable>'
    print(f"{c}: {val}")
# If there's a _META sheet, print A2
if '_META' in wb.sheetnames:
    meta = wb['_META']
    try:
        a2 = meta['A2'].value if 'A2' in meta else meta.cell(row=2, column=1).value
        print('\n_META A2:', a2)
    except Exception:
        pass
