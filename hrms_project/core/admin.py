from django.contrib import admin
from django.utils.html import format_html
from django.urls import reverse
from django.shortcuts import get_object_or_404
from .models import Category, Employee, Leave, Payroll
from .models import (
    Contract, ContractHistory, Document, LeaveHistory, Absence, Presence, Alerte, Historique,
    LeaveRequest, LeaveBalance, PositionHistory,
)
from django.utils import timezone
from datetime import timedelta, date
from django.urls import path
from django.http import HttpResponse, HttpResponseRedirect
from django.template.loader import render_to_string
from django.conf import settings
import io
import tempfile
import shutil
import subprocess
from .calculators import PayrollCalculator
from .permissions import is_rh_user
try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None
try:
    from weasyprint import HTML
except Exception:
    HTML = None
import os


@admin.register(Category)
class CategoryAdmin(admin.ModelAdmin):
    list_display = ('id', 'name')


@admin.register(Employee)
class EmployeeAdmin(admin.ModelAdmin):
    list_display = ('matricule', 'last_name', 'first_name', 'category', 'function', 'department', 'service', 'hire_date', 'archived')
    search_fields = ('matricule', 'last_name', 'first_name')
    list_filter = ('category', 'department', 'service', 'archived')
    change_form_template = 'admin/core/employee/change_form.html'

    class DocumentInline(admin.TabularInline):
        model = Document
        extra = 0
        fields = ('type', 'file_name', 'file', 'valid_from', 'valid_to', 'is_active')
        readonly_fields = ('file_name',)
        can_delete = False

    class ContractHistoryInline(admin.TabularInline):
        model = ContractHistory
        fk_name = 'employee'
        extra = 0
        fields = ('date_action', 'action', 'details')
        readonly_fields = fields
        can_delete = False
        ordering = ('-date_action',)
        verbose_name = 'Historique contrat'
        verbose_name_plural = 'Historique des contrats'

    class PositionHistoryInline(admin.TabularInline):
        model = PositionHistory
        fk_name = 'employee'
        extra = 0
        fields = ('effective_date', 'old_position', 'new_position', 'old_salary', 'new_salary', 'reason')
        readonly_fields = fields
        can_delete = False
        ordering = ('-effective_date', '-created_at')

    class _NothingInline(admin.TabularInline):
        model = Contract
        extra = 0
        max_num = 0
        can_delete = False
        verbose_name_plural = ''

    actions = ['action_archive', 'action_unarchive']

    def action_archive(self, request, queryset):
        updated = queryset.update(archived=True)
        self.message_user(request, f"{updated} employee(s) archived")

    def action_unarchive(self, request, queryset):
        updated = queryset.update(archived=False)
        self.message_user(request, f"{updated} employee(s) restored from archive")

    action_archive.short_description = 'Archiver sélection'
    action_unarchive.short_description = 'Désarchiver sélection'

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('<int:employee_id>/export_fiche/', self.admin_site.admin_view(self.export_fiche_view), name='export_employee_fiche'),
            path('<int:employee_id>/contract/<str:action>/', self.admin_site.admin_view(self.contract_action_view), name='core_employee_contract_action'),
        ]
        return custom_urls + urls

    # attach the Document inline dynamically
    def get_inline_instances(self, request, obj=None):
        base = super().get_inline_instances(request, obj)
        inlines = []
        for inline in base:
            if isinstance(inline, self._NothingInline):
                continue
            inlines.append(inline)
        if obj is not None:
            for inline_class in (self.DocumentInline, self.ContractHistoryInline, self.PositionHistoryInline):
                try:
                    inlines.append(inline_class(self.model, self.admin_site))
                except Exception:
                    continue
        return inlines

    def export_fiche_view(self, request, employee_id):
        """Generate an XLSX fiche for the given employee (and attempt a PDF).
        Saves files under exports/ and returns the XLSX as attachment. If PDF conversion
        is available (WeasyPrint), also return it as attachment (zipped) or provide both.
        """
        # load employee
        try:
            emp = Employee.objects.get(pk=employee_id)
        except Employee.DoesNotExist:
            self.message_user(request, 'Employee not found', level='error')
            return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/admin/'))

        # Determine period from POST (form) or query params or default to current month
        year = int(request.POST.get('year', request.GET.get('year', date.today().year)))
        month = int(request.POST.get('month', request.GET.get('month', date.today().month)))
        # start and end dates
        start = request.POST.get('start', request.GET.get('start'))
        end = request.POST.get('end', request.GET.get('end'))
        if not start:
            start = date(year, month, 1)
        if not end:
            # naive end of month
            if month == 12:
                end = date(year + 1, 1, 1) - timedelta(days=1)
            else:
                end = date(year, month + 1, 1) - timedelta(days=1)

        # Path to template
        template_path = None
        candidate = os.path.join(os.getcwd(), 'exports', 'FICHE_DE_PAIE.xlsx')
        if os.path.exists(candidate):
            template_path = candidate
        else:
            # try nearby paths relative to app
            candidate2 = os.path.join(os.path.dirname(os.path.dirname(__file__)), '..', 'exports', 'FICHE_DE_PAIE.xlsx')
            if os.path.exists(candidate2):
                template_path = candidate2

        if not template_path or not load_workbook:
            # cannot generate XLSX; inform user
            self.message_user(request, 'Template XLSX or openpyxl not available on server; export failed.', level='error')
            return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/admin/'))

        # load workbook and replace variables
        wb = load_workbook(template_path)
        ws = wb['FICHE'] if 'FICHE' in wb.sheetnames else wb.active

        # map employee fields to cells used in the template; allow overriding via form params (hours sup, indemnites...)
        # base extras from form/query
        extra = {
            'hours_sup_30': float(request.POST.get('hours_sup_30', request.GET.get('hours_sup_30', 0) or 0)),
            'hours_sup_50': float(request.POST.get('hours_sup_50', request.GET.get('hours_sup_50', 0) or 0)),
            'night_hours': float(request.POST.get('night_hours', request.GET.get('night_hours', 0) or 0)),
            'indemnites': float(request.POST.get('indemnites', request.GET.get('indemnites', 0) or 0)),
        }

        # Load external override parameters if present (preferred: exports/do.json, fallback: exports/do.txt)
        external_params = {}
        try:
            json_path = os.path.join(os.getcwd(), 'exports', 'do.json')
            txt_path = os.path.join(os.getcwd(), 'exports', 'do.txt')
            if os.path.exists(json_path):
                import json
                with open(json_path, 'r', encoding='utf-8') as jf:
                    external_params = json.load(jf) or {}
                self.message_user(request, f'Loaded external params from exports/do.json')
            elif os.path.exists(txt_path):
                # simple key=value parser, ignore comments starting with #
                with open(txt_path, 'r', encoding='utf-8') as tf:
                    for line in tf:
                        line = line.strip()
                        if not line or line.startswith('#'):
                            continue
                        if '=' in line:
                            k, v = line.split('=', 1)
                            external_params[k.strip()] = v.strip()
                self.message_user(request, f'Loaded external params from exports/do.txt')
        except Exception:
            external_params = {}

        def _get_ext_number(key, fallback):
            """Helper: get numeric value from external_params if present, else fallback."""
            val = external_params.get(key)
            if val is None:
                return fallback
            try:
                return float(val)
            except Exception:
                return fallback

        # If external params include overrides, apply them to extras and salary
        if external_params:
            # salary keys: salary, salaire, base_salary
            salary_override = None
            for k in ('salary', 'salaire', 'base_salary'):
                if k in external_params:
                    try:
                        salary_override = float(external_params[k])
                        break
                    except Exception:
                        salary_override = None
            if salary_override is not None:
                # will be applied below when mapping J16 is built
                extra['_salary_override'] = salary_override
            # overtime/indemnities
            extra['hours_sup_30'] = _get_ext_number('hours_sup_30', extra['hours_sup_30'])
            extra['hours_sup_50'] = _get_ext_number('hours_sup_50', extra['hours_sup_50'])
            extra['night_hours'] = _get_ext_number('night_hours', extra['night_hours'])
            extra['indemnites'] = _get_ext_number('indemnites', extra['indemnites'])
            # optional CNAPS cap
            cap_cnaps = external_params.get('cap_cnaps')
            if cap_cnaps is not None:
                try:
                    extra['_cap_cnaps'] = float(cap_cnaps)
                except Exception:
                    pass

        mapping = {
            'B16': f"{getattr(emp, 'first_name', '')} {getattr(emp, 'last_name', '')}".strip(),
            'B17': getattr(emp, 'matricule', ''),
            'B18': getattr(emp, 'function', '') or getattr(emp, 'fonction', ''),
            'B19': getattr(emp, 'cnaps_number', '') if hasattr(emp, 'cnaps_number') else getattr(emp, 'cnaps', getattr(emp, 'cnaps_num', '')),
            'B20': emp.hire_date.isoformat() if getattr(emp, 'hire_date', None) else '',
            'F10': start.isoformat() if isinstance(start, date) else str(start),
            'N10': end.isoformat() if isinstance(end, date) else str(end),
            # salary: prefer external override, then employee fields
            'J16': extra.get('_salary_override', getattr(emp, 'salary', getattr(emp, 'salaire', getattr(emp, 'base_salary', 0))) or 0),
            # extra mapped cells that exist in template
            'H26': extra['hours_sup_30'],
            'H27': extra['hours_sup_50'],
            'H28': extra['night_hours'],
            'J52': extra['indemnites'],
        }
        # Apply mapping
        for cell, val in mapping.items():
            try:
                ws[cell].value = val
            except Exception:
                pass

        # Save workbook to BytesIO
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        # Prepare XLSX bytes
        xlsx_bytes = out.getvalue()

        # Prefer LibreOffice (soffice) headless conversion for faithful XLSX->PDF rendering.
        pdf_generated = False
        pdf_bytes = None

        def try_libreoffice_convert(xlsx_bytes):
            """Try converting XLSX bytes to PDF using soffice headless. Returns pdf bytes or None."""
            try:
                with tempfile.TemporaryDirectory() as td:
                    xlsx_path = os.path.join(td, 'temp_fiche.xlsx')
                    pdf_path = os.path.join(td, 'temp_fiche.pdf')
                    # write xlsx
                    with open(xlsx_path, 'wb') as f:
                        f.write(xlsx_bytes)
                    # run soffice conversion
                    cmd = ['soffice', '--headless', '--convert-to', 'pdf', '--outdir', td, xlsx_path]
                    proc = subprocess.run(cmd, capture_output=True, timeout=30)
                    # check stdout/stderr for debugging if needed
                    if proc.returncode != 0:
                        return None
                    # find generated pdf (may have same base name but .pdf)
                    if os.path.exists(pdf_path):
                        with open(pdf_path, 'rb') as pf:
                            return pf.read()
                    # fallback: find any .pdf in td
                    for fname in os.listdir(td):
                        if fname.lower().endswith('.pdf'):
                            with open(os.path.join(td, fname), 'rb') as pf:
                                return pf.read()
                return None
            except Exception:
                return None

        # try LibreOffice first
        try:
            pdf_bytes = try_libreoffice_convert(xlsx_bytes)
            if pdf_bytes:
                pdf_generated = True
                # save for auditing
                export_dir = os.path.join(os.getcwd(), 'exports')
                os.makedirs(export_dir, exist_ok=True)
                pdf_path = os.path.join(export_dir, f"fiche_{emp.pk}_{year}_{month}.pdf")
                with open(pdf_path, 'wb') as f:
                    f.write(pdf_bytes)
        except Exception:
            pdf_generated = False

        # If LibreOffice failed and WeasyPrint is available, fall back to HTML->PDF
        if not pdf_generated and HTML is not None:
            try:
                try:
                    salaire = float(mapping.get('J16', 0) or 0)
                except Exception:
                    salaire = 0.0
                taux_journalier = round(salaire / 30) if salaire else 0
                taux_horaire = round(salaire / 173.33) if salaire else 0
                brut = salaire
                cap = ws['M40'].value if ws['M40'].value else None
                try:
                    cnaps_emp = round(min(brut * 0.01, cap)) if cap else round(brut * 0.01)
                except Exception:
                    cnaps_emp = round(brut * 0.01)
                sanitaire = round(brut * 0.01)

                def compute_irsa(brut_v):
                    tax = 0
                    tax += max(0, min(brut_v - 350000, 50000)) * 0.05
                    tax += max(0, min(brut_v - 400000, 100000)) * 0.10
                    tax += max(0, min(brut_v - 500000, 100000)) * 0.15
                    tax += max(0, min(brut_v - 600000, 3400000)) * 0.20
                    tax += max(0, brut_v - 4000000) * 0.25
                    return round(max(0, tax))

                irsa = compute_irsa(brut)
                total_retentions = cnaps_emp + sanitaire + irsa
                net = brut - total_retentions

                html = render_to_string('core/fiche_pdf.html', {
                    'employee': emp,
                    'period_start': start,
                    'period_end': end,
                    'salaire': salaire,
                    'taux_journalier': taux_journalier,
                    'taux_horaire': taux_horaire,
                    'brut': brut,
                    'cnaps_emp': cnaps_emp,
                    'sanitaire': sanitaire,
                    'irsa': irsa,
                    'total_retentions': total_retentions,
                    'net': net,
                })
                pdf_file = io.BytesIO()
                HTML(string=html).write_pdf(pdf_file)
                pdf_file.seek(0)
                pdf_bytes = pdf_file.read()
                export_dir = os.path.join(os.getcwd(), 'exports')
                os.makedirs(export_dir, exist_ok=True)
                pdf_path = os.path.join(export_dir, f"fiche_{emp.pk}_{year}_{month}.pdf")
                with open(pdf_path, 'wb') as f:
                    f.write(pdf_bytes)
                pdf_generated = True
            except Exception:
                pdf_generated = False

        # If PDF generated, return a zip with both; otherwise return XLSX
        if pdf_generated and pdf_bytes:
            import zipfile
            zip_io = io.BytesIO()
            with zipfile.ZipFile(zip_io, mode='w', compression=zipfile.ZIP_DEFLATED) as zf:
                zf.writestr(f"fiche_{emp.pk}_{year}_{month}.xlsx", xlsx_bytes)
                zf.writestr(f"fiche_{emp.pk}_{year}_{month}.pdf", pdf_bytes)
            zip_io.seek(0)
            zip_name = f"fiche_{emp.pk}_{year}_{month}.zip"
            response = HttpResponse(zip_io.read(), content_type='application/zip')
            response['Content-Disposition'] = f'attachment; filename="{zip_name}"'
            self.message_user(request, f'Fiche XLSX+PDF générée et zippée (exports/)')
            return response
        else:
            filename = f"fiche_{emp.pk}_{year}_{month}.xlsx"
            response = HttpResponse(xlsx_bytes, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            if HTML is None:
                self.message_user(request, 'Fiche XLSX générée. PDF conversion not available on server (WeasyPrint missing)')
            else:
                self.message_user(request, 'Fiche XLSX générée. PDF conversion failed or not produced.')
            return response

    def contract_action_view(self, request, employee_id, action):
        employee = get_object_or_404(Employee, pk=employee_id)
        redirect_url = reverse('admin:core_employee_change', args=[employee.pk])
        if request.method != 'POST':
            return HttpResponseRedirect(redirect_url)
        contract = employee.get_active_contract()
        if not contract:
            self.message_user(request, "Aucun contrat actif pour appliquer cette action.", level='error')
            return HttpResponseRedirect(redirect_url)
        if action == 'renew_trial':
            if contract.type != 'ESSAI':
                self.message_user(request, "La période d'essai n'est pas active.", level='warning')
            else:
                if contract.renew_trial():
                    self.message_user(request, "Période d'essai renouvelée de 6 mois.")
                else:
                    self.message_user(request, "Impossible de renouveler la période d'essai (limite atteinte).", level='error')
        elif action == 'convert_cdi':
            if contract.type == 'CDI':
                self.message_user(request, "Le contrat est déjà en CDI.", level='warning')
            else:
                contract.convert_to_cdi()
                self.message_user(request, "Contrat converti en CDI.")
        elif action == 'terminate_contract':
            contract.terminate(date_termination=date.today(), reason='Terminé via action rapide admin')
            self.message_user(request, "Contrat marqué comme terminé.")
        else:
            self.message_user(request, "Action non reconnue.", level='error')
        return HttpResponseRedirect(redirect_url)

    def changeform_view(self, request, object_id=None, form_url='', extra_context=None):
        extra_context = extra_context or {}
        obj = None
        if object_id:
            obj = self.get_object(request, object_id)
        if obj:
            extra_context.update({
                'active_contract': obj.get_active_contract(),
                'contract_badge': obj.get_contract_badge(),
                'contract_banner': obj.get_contract_status_banner(),
                'employee_deadlines': obj.get_upcoming_deadlines(),
                'salary_summary': obj.get_salary_summary(),
            })
        return super().changeform_view(request, object_id, form_url, extra_context=extra_context)


@admin.register(Leave)
class LeaveAdmin(admin.ModelAdmin):
    list_display = ('employee', 'leave_type', 'start_date', 'end_date', 'status')
    list_filter = ('leave_type', 'status')


@admin.register(LeaveRequest)
class LeaveRequestAdmin(admin.ModelAdmin):
    list_display = ('employee', 'leave_type', 'start_date', 'end_date', 'status', 'requested_at')
    list_filter = ('status', 'leave_type')
    actions = ['action_request', 'action_dept_approve', 'action_hr_approve', 'action_refuse']
    
    # Champs automatiques en lecture seule
    readonly_fields = ('requested_at',)

    def action_request(self, request, queryset):
        updated = queryset.update(status='REQUESTED')
        self.message_user(request, f"{updated} requests set to REQUESTED")

    def action_dept_approve(self, request, queryset):
        updated = queryset.update(status='DEPT_APPROVED')
        self.message_user(request, f"{updated} requests set to DEPT_APPROVED")

    def action_hr_approve(self, request, queryset):
        if not is_rh_user(request.user):
            self.message_user(request, "Seuls les utilisateurs RH peuvent valider les demandes.", level='error')
            return
        # when HR approves, check entitlement before creating Leave records
        created = 0
        skipped = 0
        for req in queryset.filter(status__in=['REQUESTED', 'DEPT_APPROVED']):
            success, msg = req.approve_by_hr(hr_user=request.user)
            if success:
                created += 1
            else:
                skipped += 1
                # attach message for skipped request
                self.message_user(request, f"Skipped {req}: {msg}")
        self.message_user(request, f"{created} Leave records created, {skipped} skipped due to entitlement checks")

    def action_refuse(self, request, queryset):
        if not is_rh_user(request.user):
            self.message_user(request, "Seuls les utilisateurs RH peuvent refuser ou annuler des demandes.", level='error')
            return
        updated = queryset.update(status='REFUSED')
        self.message_user(request, f"{updated} requests set to REFUSED")

    action_request.short_description = 'Mark selected as REQUESTED'
    action_dept_approve.short_description = 'Mark selected as DEPT_APPROVED'
    action_hr_approve.short_description = 'HR approve and create Leave records'
    action_refuse.short_description = 'Mark selected as REFUSED'


@admin.register(Payroll)
class PayrollAdmin(admin.ModelAdmin):
    list_display = ('employee', 'month', 'year', 'net_salary', 'gross_salary', 'export_payroll')
    list_filter = ('year', 'month')
    
    # Les champs calculés ne doivent PAS être saisis manuellement
    # make salary_base and computed fields readonly to prevent CRUD edits
    readonly_fields = ('salary_base', 'gross_salary', 'net_salary', 'deductions', 'created_at')

    def get_fields(self, request, obj=None):
        """Masquer les champs calculés lors de la création (ADD), les afficher en lecture seule lors de la modification."""
        if obj:  # Modification d'un objet existant
            # For existing payrolls we only expose month/year/created to prevent
            # manual CRUD of salary fields. The diagnostic panel shows employee
            # and computed breakdown.
            return (
                'month',
                'year',
                'created_at',
            )
        else:  # Création d'un nouvel objet - only select employee, month, year
            return (
                'employee',
                'month',
                'year',
            )

    def export_payroll(self, obj):
        """Return a link to export the payroll fiche as PDF for this payroll (opens in a new tab).
        Points to the `export_payroll_pdf` view which generates a PDF (when environment supports it)
        and falls back to the existing XLSX fiche export when PDF libs are not available.
        """
        if not obj:
            return '-'
        try:
            url = reverse('export_payroll_pdf', args=[obj.pk])
            # Request both XLSX+PDF (ZIP) so admin exports always include the exact Excel model
            url = f"{url}?format=both"
            return format_html('<a class="button" href="{}" target="_blank">Exporter (XLSX+PDF)</a>', url)
        except Exception:
            # fallback to employee fiche xlsx
            try:
                url = reverse('export_payroll_pdf', args=[obj.employee.pk])
                url = f"{url}?year={obj.year}&month={obj.month}"
                return format_html('<a class="button" href="{}" target="_blank">Exporter</a>', url)
            except Exception:
                return '-'

    export_payroll.short_description = 'Exporter fiche'

    def changelist_view(self, request, extra_context=None):
        """Inject a link to the reports listing so admins can quickly download recent reports."""
        extra_context = extra_context or {}
        try:
            extra_context['reports_url'] = reverse('admin:hr_reports')
        except Exception:
            extra_context['reports_url'] = None
        return super().changelist_view(request, extra_context=extra_context)

    def changeform_view(self, request, object_id=None, form_url='', extra_context=None):
        """Show additional payroll diagnostics on the change form: current contract,
        presence history for the payroll month, related histories and a computed
        payroll breakdown to explain the PDF values.
        """
        extra_context = extra_context or {}
        try:
            # Determine the employee and target month/year
            obj = None
            employee = None
            year = None
            month = None
            if object_id:
                try:
                    obj = self.get_object(request, object_id)
                except Exception:
                    obj = None
            if obj:
                employee = getattr(obj, 'employee', None)
                year = getattr(obj, 'year', None)
                month = getattr(obj, 'month', None)
            else:
                # try to get from GET params when creating a new payroll
                emp_id = request.GET.get('employee') or request.GET.get('employee_id')
                if emp_id:
                    from .models import Employee
                    try:
                        employee = Employee.objects.filter(pk=emp_id).first()
                    except Exception:
                        employee = None
                year = int(request.GET.get('year')) if request.GET.get('year') else None
                month = int(request.GET.get('month')) if request.GET.get('month') else None

            if employee:
                # current active contract
                from .models import Contract, Presence, Absence, ContractHistory, PositionHistory, Historique
                contract = Contract.objects.filter(employee=employee, active=True).order_by('-date_start').first()
                # check whether contract ends in the requested month/year
                contract_ending = False
                contract_end_date = None
                if contract and contract.date_end:
                    try:
                        if contract.date_end.year == int(year) and contract.date_end.month == int(month):
                            contract_ending = True
                            contract_end_date = contract.date_end
                    except Exception:
                        contract_ending = False

                # presences / absences for the given month/year (fallback to current month/year)
                from datetime import date
                today = date.today()
                if not year:
                    year = today.year
                if not month:
                    month = today.month

                presences = list(Presence.objects.filter(employee=employee, date__year=year, date__month=month).order_by('date'))
                absences = list(Absence.objects.filter(employee=employee, date__year=year, date__month=month).order_by('date'))
                # Filter histories to the selected month/year so diagnostics show
                # only events that occurred in that period.
                contract_hist = list(
                    ContractHistory.objects.filter(employee=employee, date_action__year=year, date_action__month=month)
                    .order_by('-date_action')
                )
                from django.db import models as dj_models
                position_hist = list(
                    PositionHistory.objects.filter(employee=employee)
                    .filter(dj_models.Q(effective_date__year=year) | dj_models.Q(created_at__year=year))
                    .filter(dj_models.Q(effective_date__month=month) | dj_models.Q(created_at__month=month))
                    .order_by('-created_at')
                )
                histories = list(
                    Historique.objects.filter(employee=employee, date_action__year=year, date_action__month=month)
                    .order_by('-date_action')
                )
                # Include leaves and leave requests in the diagnostics for the period
                diag_leaves = list(Leave.objects.filter(employee=employee, start_date__year=year, start_date__month=month))
                from django.db import models as dj_models
                diag_leave_requests = list(LeaveRequest.objects.filter(employee=employee).filter(
                    dj_models.Q(start_date__year=year, start_date__month=month) | dj_models.Q(end_date__year=year, end_date__month=month)
                ))

                # compute payroll breakdown using shared logic
                try:
                    from .payroll import compute_payroll_for_employee
                    breakdown = compute_payroll_for_employee(employee, int(year), int(month), dry_run=True)
                except Exception:
                    breakdown = {}

                # If there are no presence records for the month, assume the employee
                # worked normally for the period: no overtime/night/sunday/holiday premiums
                # and no absence deductions. Construct a fallback breakdown so the
                # diagnostics and PDF comparison reflect a full-month standard pay.
                diag_assume_full_work = False
                if not presences:
                    diag_assume_full_work = True
                    # Determine base salary from contract or employee record
                    try:
                        salary = float(contract.salary if contract and getattr(contract, 'salary', None) is not None else (getattr(employee, 'salary', None) or getattr(employee, 'base_salary', 0) or 0))
                    except Exception:
                        salary = 0.0

                    # choose default monthly hours based on sector setting
                    sector = getattr(contract, 'sector', None) or getattr(employee, 'sector', None) or getattr(settings, 'HR_SECTEUR', 'non_agricole')
                    if sector == 'agricole':
                        default_hours = getattr(settings, 'HR_HEURES_AGRICOLE', 200.0)
                    else:
                        default_hours = getattr(settings, 'HR_HEURES_NON_AGRICOLE', 173.33)

                    try:
                        hourly_rate = round(salary / default_hours, 2) if default_hours else 0
                    except Exception:
                        hourly_rate = 0

                    # CNAPS and sanitary contributions (employee share)
                    try:
                        cnaps_rate = float(getattr(settings, 'HR_TAUX_CNAPS_SALARIE', 0.01))
                    except Exception:
                        cnaps_rate = 0.01
                    try:
                        sanitary_rate = float(getattr(settings, 'HR_TAUX_SANITAIRE_SALARIE', 0.01))
                    except Exception:
                        sanitary_rate = 0.01

                    # apply CNAPS cap if configured
                    cap = getattr(settings, 'HR_PLAFOND_CNAPS', None)
                    try:
                        cnaps_emp = round(min(salary * cnaps_rate, cap)) if cap else round(salary * cnaps_rate)
                    except Exception:
                        cnaps_emp = round(salary * cnaps_rate) if salary else 0

                    sanitaire = round(salary * sanitary_rate) if salary else 0

                    def compute_irsa(brut_v):
                        tax = 0
                        tax += max(0, min(brut_v - 350000, 50000)) * 0.05
                        tax += max(0, min(brut_v - 400000, 100000)) * 0.10
                        tax += max(0, min(brut_v - 500000, 100000)) * 0.15
                        tax += max(0, min(brut_v - 600000, 3400000)) * 0.20
                        tax += max(0, brut_v - 4000000) * 0.25
                        return round(max(0, tax))

                    irsa = compute_irsa(salary)
                    total_retentions = cnaps_emp + sanitaire + irsa
                    net = round(salary - total_retentions)

                    breakdown = {
                        'salary_base': salary,
                        'hours_worked': default_hours,
                        'overtime_hours': 0,
                        'details': {
                            'overtime_pay': 0,
                            'night_premium': 0,
                            'sunday_premium': 0,
                            'holiday_premium': 0,
                        },
                        'hourly_rate': hourly_rate,
                        'gross': salary,
                        'cnaps_employee': cnaps_emp,
                        'ostie': sanitaire,
                        'absence_deduction': 0,
                        'deductions': total_retentions,
                        'net': net,
                    }

                extra_context.update({
                    'diag_assume_full_work': diag_assume_full_work,
                })

                extra_context.update({
                    'diag_employee': employee,
                    'diag_contract': contract,
                    'diag_contract_ending': contract_ending,
                    'diag_contract_end_date': contract_end_date,
                    'diag_presences': presences,
                    'diag_absences': absences,
                    'diag_contract_histories': contract_hist,
                    'diag_position_histories': position_hist,
                    'diag_histories': histories,
                    'diag_leaves': diag_leaves,
                    'diag_leave_requests': diag_leave_requests,
                    'diag_payroll_breakdown': breakdown,
                    'diag_year': year,
                    'diag_month': month,
                })
        except Exception:
            # fail silently to avoid breaking admin UI
            pass

        # Hide save buttons when viewing computed payrolls? Keep defaults.
        return super().changeform_view(request, object_id, form_url, extra_context=extra_context)


@admin.register(Contract)
class ContractAdmin(admin.ModelAdmin):
    list_display = ('employee', 'type', 'sector', 'date_start', 'date_end', 'active', 'export_contract')
    list_filter = ('type', 'sector', 'active')
    actions = ['renew_trial', 'convert_to_cdd', 'convert_to_cdi']
    
    # Champs calculés/automatiques en lecture seule
    readonly_fields = ('trial_renewals', 'created_at', 'updated_at')
    
    # Organisation du formulaire par sections
    fieldsets = (
        ('Informations principales', {
            'fields': ('employee', 'type', 'sector', 'active')
        }),
        ('Dates', {
            'fields': ('date_start', 'date_end')
        }),
        ('Conditions', {
            'fields': ('salary', 'full_time')
        }),
        ('Gestion période d\'essai', {
            'fields': ('trial_renewals', 'max_trial_renewals', 'auto_convert_to_cdi'),
            'description': 'trial_renewals est calculé automatiquement lors des renouvellements.'
        }),
        ('Détails', {
            'fields': ('notes', 'created_at', 'updated_at'),
            'classes': ('collapse',)  # Section repliable
        }),
    )

    def renew_trial(self, request, queryset):
        updated = 0
        for c in queryset:
            if c.type != 'ESSAI':
                continue
            if c.trial_renewals >= c.max_trial_renewals:
                self.message_user(request, f"Cannot renew {c}: max renewals reached")
                continue
            # extend by ~6 months (180 days)
            if c.date_end:
                c.date_end = c.date_end + timedelta(days=180)
            else:
                c.date_end = c.date_start + timedelta(days=180)
            c.trial_renewals = c.trial_renewals + 1
            c.save()
            ContractHistory.objects.create(employee=c.employee, contract=c, action='TRIAL_RENEWED', details=f'Renewed trial to {c.date_end}')
            updated += 1
        self.message_user(request, f"{updated} trial(s) renewed")

    def convert_to_cdd(self, request, queryset):
        updated = 0
        for c in queryset:
            c.type = 'CDD'
            # set a default CDD duration of 12 months from today if none
            if not c.date_end or c.date_end <= date.today():
                c.date_end = date.today() + timedelta(days=365)
            c.save()
            ContractHistory.objects.create(employee=c.employee, contract=c, action='CONVERTED_TO_CDD', details=f'Converted to CDD until {c.date_end}')
            updated += 1
        self.message_user(request, f"{updated} contract(s) converted to CDD")

    def convert_to_cdi(self, request, queryset):
        updated = 0
        for c in queryset:
            c.type = 'CDI'
            c.date_end = None
            c.save()
            ContractHistory.objects.create(employee=c.employee, contract=c, action='CONVERTED_TO_CDI', details='Converted to CDI')
            updated += 1
        self.message_user(request, f"{updated} contract(s) converted to CDI")

    renew_trial.short_description = 'Renew selected trial contracts (6 months)'
    convert_to_cdd.short_description = 'Convert selected contracts to CDD (default 12 months)'
    convert_to_cdi.short_description = 'Convert selected contracts to CDI'

    def export_contract(self, obj):
        if not obj:
            return '-'
        try:
            url = reverse('export_contract_pdf', args=[obj.pk])
            return format_html('<a class="button" href="{}" target="_blank">Exporter PDF</a>', url)
        except Exception:
            return '-'

    export_contract.short_description = 'Exporter'


@admin.register(ContractHistory)
class ContractHistoryAdmin(admin.ModelAdmin):
    list_display = ('employee', 'contract', 'action', 'date_action')
    # make history read-only in admin
    readonly_fields = [f.name for f in ContractHistory._meta.fields]
    actions = None

    def has_add_permission(self, request):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    def get_readonly_fields(self, request, obj=None):
        return self.readonly_fields

    def changeform_view(self, request, object_id=None, form_url='', extra_context=None):
        extra_context = extra_context or {}
        # hide save buttons
        extra_context.update({'show_save': False, 'show_save_and_continue': False, 'show_save_and_add_another': False})
        return super().changeform_view(request, object_id, form_url, extra_context=extra_context)


@admin.register(Document)
class DocumentAdmin(admin.ModelAdmin):
    list_display = ('employee', 'type', 'file_name', 'date_upload', 'valid_from', 'valid_to', 'is_active')
    list_filter = ('type', 'is_active')
    search_fields = ('file_name', 'employee__first_name', 'employee__last_name', 'employee__matricule')
    date_hierarchy = 'date_upload'


@admin.register(LeaveHistory)
class LeaveHistoryAdmin(admin.ModelAdmin):
    list_display = ('employee', 'leave', 'action', 'date_action')
    # make history read-only in admin
    readonly_fields = [f.name for f in LeaveHistory._meta.fields]
    actions = None

    def has_add_permission(self, request):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    def get_readonly_fields(self, request, obj=None):
        return self.readonly_fields

    def changeform_view(self, request, object_id=None, form_url='', extra_context=None):
        extra_context = extra_context or {}
        extra_context.update({'show_save': False, 'show_save_and_continue': False, 'show_save_and_add_another': False})
        return super().changeform_view(request, object_id, form_url, extra_context=extra_context)


@admin.register(Absence)
class AbsenceAdmin(admin.ModelAdmin):
    list_display = ('employee', 'date', 'justified')


@admin.register(Presence)
class PresenceAdmin(admin.ModelAdmin):
    list_display = ('employee', 'date', 'time_in', 'time_out', 'minutes_late', 'pause_minutes', 'pause_excess_minutes')
    change_list_template = 'admin/core/presence/change_list.html'

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        try:
            extra_context['generate_presence_url'] = reverse('admin:generate_presence_report')
        except Exception:
            extra_context['generate_presence_url'] = None

        today = date.today()
        default_start = today.replace(day=1)
        defaults = {
            'start': request.GET.get('start', default_start.isoformat()),
            'end': request.GET.get('end', today.isoformat()),
            'department': request.GET.get('department', ''),
            'format': request.GET.get('format', 'xlsx'),
        }
        extra_context['presence_form_defaults'] = defaults
        extra_context['presence_formats'] = ['xlsx', 'csv']
        try:
            dept_qs = (
                Employee.objects.exclude(department__isnull=True)
                .exclude(department__exact='')
                .order_by('department')
                .values_list('department', flat=True)
                .distinct()
            )
            extra_context['presence_departments'] = list(dept_qs)
        except Exception:
            extra_context['presence_departments'] = []

        return super().changelist_view(request, extra_context=extra_context)


@admin.register(Alerte)
class AlerteAdmin(admin.ModelAdmin):
    """Admin for Alerte: make alerts non-creatable/non-deletable from admin,
    make most fields read-only and expose a quick action to mark alerts as resolved.
    Alerts are intended to be generated by the system (management command).
    """
    list_display = ('employee', 'type', 'message', 'statut', 'date_creation')
    # Make all fields readonly by default except 'statut' so admins can only change status
    try:
        _all_fields = [f.name for f in Alerte._meta.fields]
    except Exception:
        _all_fields = []
    readonly_fields = [f for f in _all_fields if f != 'statut']
    actions = ['mark_resolved']

    def has_add_permission(self, request):
        # Alerts should be created by the automated generator or via code, not manually through admin
        return False

    def has_delete_permission(self, request, obj=None):
        # Prevent deletion from admin to preserve audit trail
        return False

    def get_readonly_fields(self, request, obj=None):
        # Keep fields read-only on the change form except allow editing statut
        return self.readonly_fields

    def get_model_perms(self, request):
        """Explicitly hide the 'add' permission in the admin app index so the
        'Add alerte' link/button does not appear. We already prevent adds via
        has_add_permission, but removing the perm here ensures the admin index
        and app list won't show the add link for this model.
        """
        perms = super().get_model_perms(request)
        perms['add'] = False
        return perms

    def mark_resolved(self, request, queryset):
        updated = queryset.update(statut='CLOSED')
        self.message_user(request, f"{updated} alerte(s) marquée(s) comme résolues")

    mark_resolved.short_description = 'Marquer la sélection comme RÉSOLUE (CLOSED)'


@admin.register(Historique)
class HistoriqueAdmin(admin.ModelAdmin):
    list_display = ('employee', 'action', 'date_action')
    # make history read-only in admin
    readonly_fields = [f.name for f in Historique._meta.fields]
    actions = None

    def has_add_permission(self, request):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    def get_readonly_fields(self, request, obj=None):
        return self.readonly_fields

    def changeform_view(self, request, object_id=None, form_url='', extra_context=None):
        extra_context = extra_context or {}
        extra_context.update({'show_save': False, 'show_save_and_continue': False, 'show_save_and_add_another': False})
        return super().changeform_view(request, object_id, form_url, extra_context=extra_context)


@admin.register(LeaveBalance)
class LeaveBalanceAdmin(admin.ModelAdmin):
    list_display = ('employee', 'year', 'entitlement_days', 'used_days')
    # used_days est calculé automatiquement quand les congés sont approuvés
    readonly_fields = ('used_days',)


from .models import Report


@admin.register(Report)
class ReportAdmin(admin.ModelAdmin):
    list_display = ('name', 'created_at', 'created_by', 'has_pdf', 'download_links')
    list_filter = ('created_at', 'created_by')
    search_fields = ('name',)
    date_hierarchy = 'created_at'
    ordering = ('-created_at',)
    list_per_page = 25

    def has_pdf(self, obj):
        return bool(obj.pdf_path)
    has_pdf.boolean = True
    has_pdf.short_description = 'PDF'

    def download_links(self, obj):
        links = []
        if obj.xlsx_path:
            url = reverse('admin:hr_report_download', args=[os.path.basename(obj.xlsx_path)])
            links.append(f"<a href='{url}'>XLSX</a>")
        if obj.pdf_path:
            url = reverse('admin:hr_report_download', args=[os.path.basename(obj.pdf_path)])
            links.append(f"<a href='{url}'>PDF</a>")
        return format_html(' | '.join(links))

    download_links.allow_tags = True
    download_links.short_description = 'Télécharger'

    def has_add_permission(self, request):
        # reports are generated by the management command, not manually
        return False

    def changelist_view(self, request, extra_context=None):
        """Inject some lightweight statistics and a link to the HR dashboard
        into the reports changelist page so admins can quickly jump to metrics.
        """
        extra_context = extra_context or {}
        try:
            # total persisted reports
            total_reports = Report.objects.count()
            recent = Report.objects.order_by('-created_at')[:10]
            recent_list = [{'name': r.name, 'created_at': r.created_at, 'created_by': getattr(r.created_by, 'username', None)} for r in recent]
            extra_context.update({
                'reports_total': total_reports,
                'recent_reports': recent_list,
            })
        except Exception:
            # fail silently to avoid breaking admin if Report model changed
            pass
        try:
            extra_context['hr_dashboard_url'] = reverse('admin:hr_dashboard')
        except Exception:
            extra_context['hr_dashboard_url'] = None
        return super().changelist_view(request, extra_context=extra_context)


# --- Admin custom views for HR reports (list + download) ---
from django.urls import re_path
from django.http import FileResponse, Http404
import json
from django.shortcuts import render
from django.http import JsonResponse
from django.core.management import call_command
import glob


def _reports_view(request):
    """Redirect to the admin changelist for persisted Report objects.

    We persist reports in the `Report` model; use its changelist for pagination/filters/preview.
    """
    try:
        return HttpResponseRedirect(reverse('admin:core_report_changelist'))
    except Exception:
        # fall back to exports folder listing if something goes wrong
        export_dir = os.path.join(os.getcwd(), 'exports')
        files = []
        if os.path.isdir(export_dir):
            for fname in os.listdir(export_dir):
                if not (fname.startswith('hr_report_') and (fname.endswith('.xlsx') or fname.endswith('.pdf'))):
                    continue
                fpath = os.path.join(export_dir, fname)
                st = os.stat(fpath)
                files.append({
                    'name': fname,
                    'mtime': st.st_mtime,
                    'size': st.st_size,
                })
        grouped = {}
        for f in sorted(files, key=lambda x: x['mtime'], reverse=True):
            base = os.path.splitext(f['name'])[0]
            grouped.setdefault(base, {})
            if f['name'].endswith('.xlsx'):
                grouped[base]['xlsx'] = f
            elif f['name'].endswith('.pdf'):
                grouped[base]['pdf'] = f
        context = {'reports': grouped, 'title': 'Rapports RH'}
        html = render_to_string('admin/hr_reports_list.html', context=context, request=request)
        return HttpResponse(html)


def _hr_dashboard_view(request):
    """Admin HR dashboard: compute core KPIs and render charts.

    This view computes a small set of HR statistics (headcount, gender split,
    contract types, average age, absence rate, turnover) and renders a simple
    dashboard using Chart.js.
    """
    from datetime import date, timedelta
    today = date.today()
    one_year_ago = today - timedelta(days=365)

    employees = Employee.objects.all()
    active_emps = employees.filter(is_active=True, archived=False)
    total = employees.count()
    active = active_emps.count()

    # contract counts
    cdd = 0
    cdi = 0
    for e in active_emps:
        latest = e.contracts.order_by('-date_start').first()
        if latest:
            if latest.type == 'CDD':
                cdd += 1
            elif latest.type == 'CDI':
                cdi += 1

    # gender distribution
    genders = {label: employees.filter(gender=code).count() for code, label in Employee.GENDER_CHOICES}

    # avg age
    ages = []
    for e in employees:
        if e.birth_date:
            ages.append((today - e.birth_date).days // 365)
    avg_age = int(sum(ages) / len(ages)) if ages else None

    # top departments
    by_dept = {}
    for e in employees:
        key = e.department or 'UNASSIGNED'
        by_dept[key] = by_dept.get(key, 0) + 1
    top_depts = sorted(by_dept.items(), key=lambda kv: kv[1], reverse=True)[:10]

    # absence days last 12 months
    total_absence_days = Absence.objects.filter(date__gte=one_year_ago).count()
    absence_rate = round(total_absence_days / max(1, active), 2)

    # turnover: contracts terminated over last 12 months
    terminations = Contract.objects.filter(date_end__gte=one_year_ago, date_end__lte=today).count()
    avg_headcount = max(1, (total + active) / 2)
    turnover = round((terminations / avg_headcount) * 100, 2)

    context = {
        'title': 'Tableau de bord RH',
        'total_employees': total,
        'active_employees': active,
        'cdi_count': cdi,
        'cdd_count': cdd,
        'genders': json.dumps(genders),
        'avg_age': avg_age,
        'top_depts': json.dumps(top_depts),
        'absence_days': total_absence_days,
        'absence_rate': absence_rate,
        'turnover': turnover,
    }

    # render using a template that extends admin base
    return render(request, 'admin/hr_dashboard.html', context)


def _hr_stats_json(request):
    """Return detailed HR statistics as JSON (for dashboards or API clients)."""
    from datetime import date
    today = date.today()

    employees = Employee.objects.all()
    active_emps = employees.filter(is_active=True, archived=False)

    total = employees.count()
    active = active_emps.count()

    # Gender counts
    gender_counts = {label: employees.filter(gender=code).count() for code, label in Employee.GENDER_CHOICES}

    # Age distribution and stats
    ages = []
    for e in employees:
        if e.birth_date:
            ages.append((today - e.birth_date).days // 365)
    ages_sorted = sorted(ages)
    age_stats = {}
    if ages:
        import statistics
        age_stats['min'] = int(min(ages_sorted))
        age_stats['max'] = int(max(ages_sorted))
        age_stats['avg'] = float(sum(ages_sorted)) / len(ages_sorted)
        try:
            age_stats['median'] = statistics.median(ages_sorted)
        except Exception:
            age_stats['median'] = age_stats['avg']
        # buckets: <25,25-34,35-44,45-54,55+
        buckets = {'<25': 0, '25-34': 0, '35-44': 0, '45-54': 0, '55+': 0}
        for a in ages_sorted:
            if a < 25:
                buckets['<25'] += 1
            elif a < 35:
                buckets['25-34'] += 1
            elif a < 45:
                buckets['35-44'] += 1
            elif a < 55:
                buckets['45-54'] += 1
            else:
                buckets['55+'] += 1
        age_stats['buckets'] = buckets
    else:
        age_stats = {'min': None, 'max': None, 'avg': None, 'median': None, 'buckets': {}}

    # Service / department distribution
    by_service = {}
    for e in employees:
        key = e.service or e.department or 'UNASSIGNED'
        by_service[key] = by_service.get(key, 0) + 1
    # top 20 services
    top_services = sorted(by_service.items(), key=lambda kv: kv[1], reverse=True)[:20]

    # Contract type counts among active employees
    contract_types = {}
    for e in active_emps:
        latest = e.contracts.order_by('-date_start').first()
        t = latest.type if latest else 'NONE'
        contract_types[t] = contract_types.get(t, 0) + 1

    # Full-time vs part-time
    fulltime = active_emps.filter(contracts__full_time=True).distinct().count()
    parttime = active - fulltime

    # Tenure (anciennete) buckets in years
    tenure_buckets = {'<1': 0, '1-3': 0, '4-6': 0, '7-10': 0, '10+': 0}
    for e in active_emps:
        if e.hire_date:
            yrs = (today - e.hire_date).days // 365
        else:
            # fallback via contracts
            yrs = 0
            if e.contracts.exists():
                first = e.contracts.order_by('date_start').first()
                if first and first.date_start:
                    yrs = (today - first.date_start).days // 365
        if yrs < 1:
            tenure_buckets['<1'] += 1
        elif yrs <= 3:
            tenure_buckets['1-3'] += 1
        elif yrs <= 6:
            tenure_buckets['4-6'] += 1
        elif yrs <= 10:
            tenure_buckets['7-10'] += 1
        else:
            tenure_buckets['10+'] += 1

    data = {
        'total_employees': total,
        'active_employees': active,
        'gender_counts': gender_counts,
        'age_stats': age_stats,
        'top_services': top_services,
        'contract_types': contract_types,
        'full_time': fulltime,
        'part_time': parttime,
        'tenure_buckets': tenure_buckets,
    }

    return JsonResponse(data)


def _report_download_view(request, filename):
    export_dir = os.path.join(os.getcwd(), 'exports')
    # sanitize filename to avoid directory traversal
    if '..' in filename or filename.startswith('/') or filename.startswith('\\'):
        raise Http404('Invalid filename')
    fpath = os.path.join(export_dir, filename)
    if not os.path.exists(fpath):
        raise Http404('File not found')
    return FileResponse(open(fpath, 'rb'), as_attachment=True, filename=filename)


def _generate_presence_report_view(request):
    """Generate a presence report and return the generated file as attachment.

    Accepts optional GET params: start, end (YYYY-MM-DD), department, format (xlsx|csv).
    """
    from datetime import date
    start = request.GET.get('start')
    end = request.GET.get('end')
    dept = request.GET.get('department')
    fmt = request.GET.get('format', 'xlsx')

    # defaults
    today = date.today()
    if not start:
        start = today.replace(day=1).isoformat()
    if not end:
        end = today.isoformat()

    outdir = os.path.join(os.getcwd(), 'exports')

    try:
        # call the management command to generate the report and create a Report row
        call_command('generate_presence_report', start=start, end=end, department=dept or None, format=fmt, output_dir=outdir, create_report=True)
    except Exception as e:
        return HttpResponse(f'Error generating report: {e}', status=500)

    # find generated file
    pattern = os.path.join(outdir, f'presence_report_{start}_{end}*')
    matches = sorted(glob.glob(pattern), key=os.path.getmtime, reverse=True)
    if not matches:
        return HttpResponse(f'Report generated but file not found (searched: {pattern})', status=404)
    fpath = matches[0]
    filename = os.path.basename(fpath)
    return FileResponse(open(fpath, 'rb'), as_attachment=True, filename=filename)


# Register URLs into admin site
def _wrap_admin_urls(original_get_urls):
    def get_urls():
        urls = original_get_urls()
        my_urls = [
            path('reports/', admin.site.admin_view(_reports_view), name='hr_reports'),
            path('reports/download/<str:filename>/', admin.site.admin_view(_report_download_view), name='hr_report_download'),
            # Expose the HR dashboard to all visitors (no admin login required)
            path('hr-dashboard/', _hr_dashboard_view, name='hr_dashboard'),
            path('reports/presence/generate/', admin.site.admin_view(_generate_presence_report_view), name='generate_presence_report'),
            path('hr-stats.json', _hr_stats_json, name='hr_stats_json'),
        ]
        return my_urls + urls
    return get_urls


# Monkey-patch admin site's get_urls to expose our report pages
admin.site.get_urls = _wrap_admin_urls(admin.site.get_urls)
