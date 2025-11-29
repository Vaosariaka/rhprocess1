from django.db.models.signals import pre_save, post_save
from django.dispatch import receiver
from django.utils import timezone
from django.conf import settings
from .models import Contract, ContractHistory, Leave, LeaveHistory, Alerte, Presence


@receiver(pre_save, sender=Contract)
def contract_pre_save(sender, instance: Contract, **kwargs):
    # detect changes vs existing contract
    if not instance.pk:
        return
    try:
        old = Contract.objects.get(pk=instance.pk)
    except Contract.DoesNotExist:
        return
    if old.type != instance.type or old.active != instance.active or old.date_end != instance.date_end:
        # record history
        ContractHistory.objects.create(
            employee=instance.employee,
            contract=instance,
            action=f"UPDATED: {old.type} -> {instance.type}",
            details=f"Changed active {old.active}->{instance.active}, date_end {old.date_end}->{instance.date_end}")


@receiver(post_save, sender=Contract)
def contract_post_save(sender, instance: Contract, created, **kwargs):
    # if contract of type ESSAI ended, create an alert or auto-transition if configured
    if instance.type == 'ESSAI' and instance.date_end:
        today = timezone.localdate()
        if instance.date_end < today:
            default_after = getattr(settings, 'HR_PAYROLL', {}).get('DEFAULT_AFTER_ESSAI')
            if default_after in ('CDD', 'CDI'):
                # create new contract automatically starting the day after date_end
                Contract.objects.create(
                    employee=instance.employee,
                    type=default_after,
                    sector=instance.sector,
                    date_start=instance.date_end + timezone.timedelta(days=1),
                    salary=instance.salary,
                    full_time=instance.full_time,
                    active=True,
                )
                ContractHistory.objects.create(
                    employee=instance.employee,
                    contract=instance,
                    action=f"AUTO_PROMOTE_TO_{default_after}",
                    details=f"Trial ended on {instance.date_end}, auto-created {default_after}")
            else:
                Alerte.objects.create(
                    employee=instance.employee,
                    type='ESSAI_EXPIRED',
                    message=f'Trial contract expired on {instance.date_end} for {instance.employee}. Please update contract.',
                )


@receiver(pre_save, sender=Leave)
def leave_pre_save(sender, instance: Leave, **kwargs):
    if not instance.pk:
        return
    try:
        old = Leave.objects.get(pk=instance.pk)
    except Leave.DoesNotExist:
        return
    if old.status != instance.status:
        LeaveHistory.objects.create(
            employee=instance.employee,
            leave=instance,
            action=f"STATUS_CHANGE {old.status} -> {instance.status}",
            details=f"Changed by save"
        )


@receiver(post_save, sender=Presence)
def presence_post_save(sender, instance: Presence, created, **kwargs):
    # if late above threshold, create an alert
    late_threshold = getattr(settings, 'HR_PAYROLL', {}).get('LATE_THRESHOLD', 15)
    try:
        if instance.minutes_late and instance.minutes_late > late_threshold:
            Alerte.objects.create(
                employee=instance.employee,
                type='LATE',
                message=f'Late {instance.minutes_late} minutes on {instance.date} for {instance.employee}',
            )
        if instance.pause_excess_minutes and instance.pause_excess_minutes > 0:
            Alerte.objects.get_or_create(
                employee=instance.employee,
                type='PAUSE_EXCESSIVE',
                message=f'Pause excessive ({instance.pause_minutes} min) le {instance.date} pour {instance.employee}',
                defaults={'statut': 'OPEN'},
            )
    except Exception:
        pass
import logging
from django.db.models.signals import post_save, post_delete
from django.dispatch import receiver
from django.conf import settings
from django.db import transaction
from pathlib import Path
import openpyxl
from openpyxl import Workbook
from io import BytesIO
import shutil
from datetime import datetime

from .models import Employee, Payroll, Leave

logger = logging.getLogger(__name__)


def _get_workspace_parent():
    # settings.BASE_DIR points to hrms_project/; files are in parent (workspace)
    return Path(settings.BASE_DIR).parent


def write_employees_to_fiche():
    base = _get_workspace_parent()
    path = base / 'FICHE DE PAIE .xlsx'
    # create backups directory
    backups_dir = base / 'backups'
    backups_dir.mkdir(parents=True, exist_ok=True)
    try:
        # backup existing file before modifying
        if path.exists():
            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_path = backups_dir / f"{path.name}.{ts}.bak"
            try:
                shutil.copy2(path, backup_path)
                logger.info(f'Backup created: {backup_path}')
            except Exception:
                logger.exception('Failed to create backup for %s', path)
        if path.exists():
            wb = openpyxl.load_workbook(path)
        else:
            wb = Workbook()
        # choose sheet name 'FICHE' or active
        if 'FICHE' in wb.sheetnames:
            ws = wb['FICHE']
        else:
            ws = wb.active
            ws.title = 'FICHE'
        # clear sheet
        ws.delete_rows(1, ws.max_row)
        headers = ['Matricule', 'CNAPS', 'Last Name', 'First Name', 'Category', 'Function', 'Hire Date', 'Salary Base']
        ws.append(headers)
        for e in Employee.objects.select_related('category').all().order_by('matricule'):
            ws.append([
                e.matricule,
                e.cnaps_number or '',
                e.last_name,
                e.first_name,
                str(e.category) if e.category else '',
                e.function,
                e.hire_date.isoformat() if e.hire_date else '',
                float(e.salary_base or 0),
            ])
        wb.save(path)
        logger.info(f'Wrote employees to {path}')
    except Exception as exc:
        logger.exception('Failed to write employees fiche: %s', exc)


def write_payrolls_and_leaves_to_etat():
    base = _get_workspace_parent()
    path = base / 'ETAT DE PAIE MAJ.xlsx'
    backups_dir = base / 'backups'
    backups_dir.mkdir(parents=True, exist_ok=True)
    try:
        # backup existing file before modifying
        if path.exists():
            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_path = backups_dir / f"{path.name}.{ts}.bak"
            try:
                shutil.copy2(path, backup_path)
                logger.info(f'Backup created: {backup_path}')
            except Exception:
                logger.exception('Failed to create backup for %s', path)
        if path.exists():
            wb = openpyxl.load_workbook(path)
        else:
            wb = Workbook()
        # Payrolls sheet
        if 'Payrolls' in wb.sheetnames:
            ws_p = wb['Payrolls']
        else:
            ws_p = wb.create_sheet('Payrolls')
        ws_p.delete_rows(1, ws_p.max_row)
        headers_p = ['Matricule', 'Last Name', 'First Name', 'Month', 'Year', 'Gross', 'Net', 'Created At']
        ws_p.append(headers_p)
        for p in Payroll.objects.select_related('employee').all().order_by('year', 'month'):
            ws_p.append([
                p.employee.matricule,
                p.employee.last_name,
                p.employee.first_name,
                p.month,
                p.year,
                float(p.gross_salary or 0),
                float(p.net_salary or 0),
                p.created_at.isoformat(),
            ])

        # Leaves sheet
        if 'Leaves' in wb.sheetnames:
            ws_l = wb['Leaves']
        else:
            ws_l = wb.create_sheet('Leaves')
        ws_l.delete_rows(1, ws_l.max_row)
        headers_l = ['Matricule', 'Last Name', 'First Name', 'Type', 'Start', 'End', 'Days', 'Status', 'Note']
        ws_l.append(headers_l)
        for l in Leave.objects.select_related('employee').all().order_by('-start_date'):
            ws_l.append([
                l.employee.matricule,
                l.employee.last_name,
                l.employee.first_name,
                l.leave_type,
                l.start_date.isoformat() if l.start_date else '',
                l.end_date.isoformat() if l.end_date else '',
                l.days,
                l.status,
                l.note or '',
            ])

        # Save workbook
        # If workbook had a default sheet created by Workbook() and it's empty, remove it
        if 'Sheet' in wb.sheetnames and len(wb.sheetnames) > 1:
            std = wb['Sheet']
            if std.max_row == 1 and std.max_column == 1 and std['A1'].value is None:
                wb.remove(std)

        wb.save(path)
        logger.info(f'Wrote payrolls and leaves to {path}')
    except Exception as exc:
        logger.exception('Failed to write etat de paie: %s', exc)


@receiver(post_save, sender=Employee)
def employee_saved(sender, instance, created, **kwargs):
    # schedule write after transaction commit
    try:
        transaction.on_commit(write_employees_to_fiche)
    except Exception:
        write_employees_to_fiche()


@receiver(post_delete, sender=Employee)
def employee_deleted(sender, instance, **kwargs):
    try:
        transaction.on_commit(write_employees_to_fiche)
    except Exception:
        write_employees_to_fiche()


@receiver(post_save, sender=Payroll)
def payroll_saved(sender, instance, created, **kwargs):
    try:
        transaction.on_commit(write_payrolls_and_leaves_to_etat)
    except Exception:
        write_payrolls_and_leaves_to_etat()


@receiver(post_delete, sender=Payroll)
def payroll_deleted(sender, instance, **kwargs):
    try:
        transaction.on_commit(write_payrolls_and_leaves_to_etat)
    except Exception:
        write_payrolls_and_leaves_to_etat()


@receiver(post_save, sender=Leave)
def leave_saved(sender, instance, created, **kwargs):
    try:
        transaction.on_commit(write_payrolls_and_leaves_to_etat)
    except Exception:
        write_payrolls_and_leaves_to_etat()


@receiver(post_delete, sender=Leave)
def leave_deleted(sender, instance, **kwargs):
    try:
        transaction.on_commit(write_payrolls_and_leaves_to_etat)
    except Exception:
        write_payrolls_and_leaves_to_etat()


# --- Audit logging and leave workflow enhancements ---
from django.db.models.signals import post_delete
from .models import AuditLog, Document, LeaveRequest, Payroll, Historique
from .middleware import get_current_user
from django.forms.models import model_to_dict
from datetime import date, datetime
from decimal import Decimal
from uuid import UUID


def _sanitize_for_json(obj):
    """Recursively convert common non-JSON-serializable types to serializable forms.

    Converts date/datetime to isoformat, Decimal to str, UUID to str, and
    walks dicts/lists/tuples.
    """
    if obj is None:
        return None
    if isinstance(obj, (str, int, float, bool)):
        return obj
    if isinstance(obj, (date, datetime)):
        return obj.isoformat()
    if isinstance(obj, Decimal):
        # keep precision by storing as string
        return str(obj)
    if isinstance(obj, UUID):
        return str(obj)
    if isinstance(obj, dict):
        return {str(k): _sanitize_for_json(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple, set)):
        return [_sanitize_for_json(v) for v in obj]
    # fallback to string representation
    try:
        return str(obj)
    except Exception:
        return None


WATCHED = (Employee, Contract, LeaveRequest, Document, Payroll)


def _create_audit(instance, action):
    try:
        user = get_current_user()
        raw_changes = model_to_dict(instance) if instance is not None else None
        try:
            changes = _sanitize_for_json(raw_changes)
        except Exception:
            changes = None

        def _do_create():
            try:
                AuditLog.objects.create(
                    user=user,
                    action=action,
                    model_name=instance.__class__.__name__ if instance is not None else None,
                    object_pk=str(getattr(instance, 'pk', None)),
                    changes=changes,
                )
            except Exception:
                logger.exception('Failed to create AuditLog')

        # schedule audit creation after transaction commit to avoid breaking
        # the current atomic block in case something goes wrong with the DB
        try:
            transaction.on_commit(_do_create)
        except Exception:
            # if on_commit not available or we're outside a transaction, create now
            _do_create()
    except Exception:
        logger.exception('Unexpected error in _create_audit')


# Populate Historique for Payroll model so admin SQL like `select * from historique where action like '%payroll%'` will return rows
@receiver(post_save, sender=Payroll)
def payroll_history_on_save(sender, instance, created, **kwargs):
    try:
        act = 'PAYROLL_CREATED' if created else 'PAYROLL_UPDATED'
        details = ''
        try:
            details = str(model_to_dict(instance))
        except Exception:
            details = f'Payroll {instance.pk} saved.'
        Historique.objects.create(employee=instance.employee, action=act, details=details)
    except Exception:
        pass


@receiver(post_delete, sender=Payroll)
def payroll_history_on_delete(sender, instance, **kwargs):
    try:
        details = f'Payroll {getattr(instance, "pk", None)} deleted.'
        Historique.objects.create(employee=instance.employee, action='PAYROLL_DELETED', details=details)
    except Exception:
        pass


@receiver(post_save)
def generic_post_save(sender, instance, created, **kwargs):
    try:
        if sender in WATCHED:
            _create_audit(instance, 'CREATE' if created else 'UPDATE')
    except Exception:
        pass


@receiver(post_delete)
def generic_post_delete(sender, instance, **kwargs):
    try:
        if sender in WATCHED:
            _create_audit(instance, 'DELETE')
    except Exception:
        pass


# LeaveRequest workflow: when status transitions to REQUESTED, create alert
@receiver(post_save, sender=LeaveRequest)
def leave_request_notify(sender, instance, created, **kwargs):
    try:
        # create alert when transitioning into REQUESTED
        if instance.status == 'REQUESTED':
            Alerte.objects.create(employee=instance.employee, type='LEAVE_REQUEST', message=f'Leave request: {instance}', statut='OPEN')
    except Exception:
        pass
