from django.core.management.base import BaseCommand
from django.utils import timezone
from core.models import Employee, Contract, Absence, Presence, compute_anciennete
from datetime import date, timedelta
import os
from openpyxl import Workbook
from core.models import Report
import tempfile
import subprocess


class Command(BaseCommand):
    help = 'Generate HR statistics report (Excel) and optionally convert to PDF via LibreOffice'

    def add_arguments(self, parser):
        parser.add_argument('--outdir', type=str, default=None, help='Output directory (defaults to exports/)')
        parser.add_argument('--pdf', action='store_true', help='Also convert generated XLSX to PDF using soffice')

    def handle(self, *args, **options):
        outdir = options.get('outdir') or os.path.join(os.getcwd(), 'exports')
        os.makedirs(outdir, exist_ok=True)
        today = date.today()

        employees = Employee.objects.all()
        active_emps = employees.filter(is_active=True, archived=False)
        total = employees.count()
        active = active_emps.count()

        # counts by contract type (use latest active contract if any)
        cdd = 0
        cdi = 0
        for e in active_emps:
            latest = e.contracts.order_by('-date_start').first()
            if latest:
                if latest.type == 'CDD':
                    cdd += 1
                elif latest.type == 'CDI':
                    cdi += 1

        # gender distribution
        genders = {}
        for g, _ in Employee.GENDER_CHOICES:
            genders[g] = employees.filter(gender=g).count()

        # age stats (compute age in years where birth_date available)
        ages = []
        for e in employees:
            if e.birth_date:
                ages.append((today - e.birth_date).days // 365)
        avg_age = int(sum(ages) / len(ages)) if ages else None

        # category distribution
        by_category = {}
        for e in employees:
            key = e.category.name if e.category else 'UNASSIGNED'
            by_category[key] = by_category.get(key, 0) + 1

        # anciennete distribution (years buckets)
        buckets = {'<1': 0, '1-3': 0, '3-5': 0, '5-10': 0, '>10': 0}
        anciennete_years = []
        for e in employees:
            anc = compute_anciennete(e, as_of_date=today)
            yrs = anc.get('years', 0)
            anciennete_years.append(yrs)
            if yrs < 1:
                buckets['<1'] += 1
            elif yrs < 3:
                buckets['1-3'] += 1
            elif yrs < 5:
                buckets['3-5'] += 1
            elif yrs < 10:
                buckets['5-10'] += 1
            else:
                buckets['>10'] += 1
        avg_anciennete = int(sum(anciennete_years) / len(anciennete_years)) if anciennete_years else None

        # employees by department
        by_dept = {}
        for e in employees:
            key = e.department or 'UNASSIGNED'
            by_dept[key] = by_dept.get(key, 0) + 1

        # mass salariale (monthly) — sum salary_base for active employees
        total_payroll = 0
        for e in active_emps:
            try:
                total_payroll += float(e.salary_base or 0)
            except Exception:
                pass

        # Absence rate (last 12 months): total absence days / active headcount
        one_year_ago = today - timedelta(days=365)
        total_absence_days = 0
        for a in Absence.objects.filter(date__gte=one_year_ago):
            total_absence_days += 1
        absence_rate_per_employee = round(total_absence_days / max(1, active), 2)

        # Punctuality: average minutes_late per presence over last 12 months
        presences = Presence.objects.filter(date__gte=one_year_ago)
        total_minutes_late = 0
        pres_count = presences.count()
        for p in presences:
            total_minutes_late += (p.minutes_late or 0)
        avg_minutes_late = int(total_minutes_late / pres_count) if pres_count else None

        # turnover: terminated contracts in last 12 months / average headcount (simple)
        one_year_ago = today - timedelta(days=365)
        terminations = Contract.objects.filter(date_end__gte=one_year_ago, date_end__lte=today).count()
        avg_headcount = max(1, (total + active) / 2)
        turnover = round((terminations / avg_headcount) * 100, 2)

        # Build Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Summary'
        ws.append(['Report generated', today.isoformat()])
        ws.append([])
        ws.append(['Total employees', total])
        ws.append(['Active employees', active])
        ws.append(['CDI', cdi])
        ws.append(['CDD', cdd])
        ws.append(['Average age', avg_age or 'N/A'])
        ws.append(['Turnover (%) last 12 months', turnover])
        ws.append(['Monthly payroll total', total_payroll])
        ws.append([])
        ws.append(['Gender distribution'])
        for k, v in genders.items():
            ws.append([k, v])
        ws.append([])
        ws.append(['Category distribution'])
        for k, v in by_category.items():
            ws.append([k, v])
        ws.append([])
        ws.append(['Anciennete (years) average', avg_anciennete or 'N/A'])
        ws.append(['Anciennete distribution'])
        for k, v in buckets.items():
            ws.append([k, v])
        ws.append([])
        ws.append(['Absence days (last 12 months)', total_absence_days])
        ws.append(['Absence days per active employee (last 12 months)', absence_rate_per_employee])
        ws.append(['Average minutes late (last 12 months)', avg_minutes_late or 'N/A'])
        ws2 = wb.create_sheet('By Department')
        ws2.append(['Department', 'Count'])
        for k, v in sorted(by_dept.items(), key=lambda kv: kv[1], reverse=True):
            ws2.append([k, v])
        ws3 = wb.create_sheet('Anciennete')
        ws3.append(['Bucket', 'Count'])
        for k, v in buckets.items():
            ws3.append([k, v])
        ws4 = wb.create_sheet('Category')
        ws4.append(['Category', 'Count'])
        for k, v in by_category.items():
            ws4.append([k, v])

        # add simple charts using openpyxl
        try:
            from openpyxl.chart import PieChart, Reference, BarChart
            pie = PieChart()
            # Build a small contract type pie on main sheet
            contract_start = ws.max_row + 2
            ws.append([])
            ws.append(['Contract type', 'Count'])
            ws.append(['CDI', cdi])
            ws.append(['CDD', cdd])
            pie = PieChart()
            labels = Reference(ws, min_col=1, min_row=contract_start+1, max_row=contract_start+2)
            data = Reference(ws, min_col=2, min_row=contract_start+1, max_row=contract_start+2)
            pie.add_data(data, titles_from_data=False)
            pie.set_categories(labels)
            pie.title = 'Contract types (active)'
            ws.add_chart(pie, f'E2')
            # department bar chart
            dept_rows = list(sorted(by_dept.items(), key=lambda kv: kv[1], reverse=True))[:10]
            dept_start = ws.max_row + 2
            ws.append([])
            ws.append(['Top Departments', 'Count'])
            for k, v in dept_rows:
                ws.append([k, v])
            bc = BarChart()
            data = Reference(ws, min_col=2, min_row=dept_start+1, max_row=dept_start+len(dept_rows))
            cats = Reference(ws, min_col=1, min_row=dept_start+1, max_row=dept_start+len(dept_rows))
            bc.add_data(data, titles_from_data=False)
            bc.set_categories(cats)
            bc.title = 'Top Departments by headcount'
            ws.add_chart(bc, 'E20')
        except Exception:
            # if openpyxl charting not available, ignore charts
            pass

        # save workbook
        fname = f'hr_report_{today.isoformat()}.xlsx'
        out_xlsx = os.path.join(outdir, fname)
        wb.save(out_xlsx)
        self.stdout.write(self.style.SUCCESS(f'Wrote HR report to {out_xlsx}'))

        # Persist report record in DB (no user for management command)
        try:
            pdf_path = None
            if options.get('pdf'):
                pdf_path = os.path.join(outdir, fname.rsplit('.', 1)[0] + '.pdf')
                if not os.path.exists(pdf_path):
                    pdf_path = None
            Report.objects.create(name=fname, xlsx_path=out_xlsx, pdf_path=pdf_path)
            self.stdout.write(self.style.SUCCESS('Persisted report record in DB'))
        except Exception as e:
            self.stderr.write(self.style.ERROR(f'Failed to persist Report record: {e}'))

        # optional PDF conversion via soffice
        if options.get('pdf'):
            try:
                proc = subprocess.run(['soffice', '--headless', '--convert-to', 'pdf', '--outdir', outdir, out_xlsx], capture_output=True, timeout=60)
                if proc.returncode == 0:
                    pdfname = fname.rsplit('.', 1)[0] + '.pdf'
                    self.stdout.write(self.style.SUCCESS(f'Converted report to PDF: {os.path.join(outdir, pdfname)}'))
                else:
                    self.stderr.write(self.style.ERROR(f'LibreOffice conversion failed: {proc.stderr.decode()[:200]}'))
            except Exception as e:
                self.stderr.write(self.style.ERROR(f'PDF conversion error: {e}'))
