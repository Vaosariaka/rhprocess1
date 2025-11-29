from django.core.management.base import BaseCommand, CommandError
from django.conf import settings
from core.models import Employee, Payroll, Category
from pathlib import Path
import openpyxl
from datetime import datetime
from decimal import Decimal, InvalidOperation


def _clean_decimal(value):
    if value is None:
        return None
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    s = str(value).strip()
    if s == '':
        return None
    # Replace comma decimal separator with dot, remove spaces
    s = s.replace('\u00A0', '').replace(' ', '').replace(',', '.')
    # remove any non-number characters except dot and -
    allowed = '0123456789.-'
    s = ''.join(ch for ch in s if ch in allowed)
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


class Command(BaseCommand):
    help = 'Import payroll data from Excel files into Employee and Payroll models.'

    def add_arguments(self, parser):
        parser.add_argument('--payroll-file', dest='payroll_file', default='ETAT DE PAIE MAJ.xlsx', help='Fichier Excel contenant l\'état de paie')
        parser.add_argument('--sheet', dest='sheet', default=None, help='Nom de la feuille à lire (optionnel)')
        parser.add_argument('--month', dest='month', type=int, default=None, help='Mois à utiliser si la colonne DATE est absente (1-12)')
        parser.add_argument('--year', dest='year', type=int, default=None, help='Année à utiliser si la colonne DATE est absente')
        parser.add_argument('--dry-run', dest='dry_run', action='store_true', help='Ne pas sauvegarder en base, afficher seulement ce qui serait importé')

    def handle(self, *args, **options):
        base = Path(settings.BASE_DIR)
        payroll_path = base / options['payroll_file']
        if not payroll_path.exists():
            raise CommandError(f"Fichier introuvable: {payroll_path}")

        self.stdout.write(self.style.NOTICE(f"Ouverture du fichier: {payroll_path}"))
        wb = openpyxl.load_workbook(payroll_path, data_only=True)

        sheet = None
        if options['sheet']:
            if options['sheet'] in wb.sheetnames:
                sheet = wb[options['sheet']]
            else:
                raise CommandError(f"Feuille '{options['sheet']}' introuvable dans le fichier")
        else:
            # heuristique: chercher feuille contenant 'ETAT' dans le nom
            for name in wb.sheetnames:
                if 'ETAT' in name.upper() or 'PAIE' in name.upper():
                    sheet = wb[name]
                    break
            if sheet is None:
                sheet = wb[wb.sheetnames[0]]

        # find header row by searching for known header keywords (robust)
        header_row = None
        headers = {}
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
            if not any(row):
                continue
            row_lc = [str(c).strip().upper() if c is not None else '' for c in row]
            # look for matricule header heuristically
            if any(('MATR' in c or 'N\u00b0' in c or 'NUM' in c or 'ID' == c) for c in row_lc):
                header_row = row_idx
                for idx, cell in enumerate(row_lc):
                    headers[cell] = idx
                break

        if header_row is None:
            self.stdout.write(self.style.WARNING("Impossible de détecter la ligne d'en-tête ; j'essaierai la première ligne utile."))
            first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            header_row = 1
            headers = { (str(cell).strip().upper() if cell else f'COL{i}'): i for i, cell in enumerate(first_row) }

        # Debug: show detected headers
        self.stdout.write(self.style.NOTICE(f'Detected header row: {header_row}'))
        self.stdout.write(self.style.NOTICE(f'Headers mapping (sample): {list(headers.items())[:10]}'))

        # normalize header keys mapping to known fields
        key_map = {}
        for h, idx in headers.items():
            # normalize header label
            label = h or ''
            if any(token in label for token in ('MATR', 'N\u00b0', 'NUM', 'ID')):
                key_map['matricule'] = idx
            if any(token in label for token in ('NOM', 'PREN', 'NAME')):
                # column that contains name might be 'NOM PRENOM' or 'NOM ET PRENOM'
                key_map.setdefault('name', idx)
            if 'SALAIRE' in label and 'BASE' in label:
                key_map['salary_base'] = idx
            if 'BRUT' in label:
                key_map['gross_salary'] = idx
            if 'NET' in label and 'NET' not in key_map:
                key_map['net_salary'] = idx
            if 'DATE' in label or 'MOIS' in label:
                key_map.setdefault('date', idx)
            if 'CNAPS' in label:
                key_map.setdefault('cnaps', idx)

        if 'matricule' not in key_map:
            raise CommandError('Colonne matricule non trouvée dans le fichier')

        created_employees = 0
        created_payrolls = 0
        skipped = 0
        rows_iter = sheet.iter_rows(min_row=sheet.min_row+1, values_only=True)

        for row in rows_iter:
            try:
                matr = row[key_map['matricule']]
            except Exception:
                continue
            if matr is None:
                skipped += 1
                continue
            matr = str(matr).strip()
            # name
            name = None
            if 'name' in key_map:
                name = row[key_map['name']]
            # salary base
            salary_base = None
            if 'salary_base' in key_map:
                salary_base = _clean_decimal(row[key_map['salary_base']])
            gross = None
            net = None
            if 'gross_salary' in key_map:
                gross = _clean_decimal(row[key_map['gross_salary']])
            if 'net_salary' in key_map:
                net = _clean_decimal(row[key_map['net_salary']])

            # date -> month/year
            month = options.get('month')
            year = options.get('year')
            if 'date' in key_map and (month is None or year is None):
                dval = row[key_map['date']]
                if isinstance(dval, datetime):
                    month = dval.month
                    year = dval.year
                else:
                    # try parse
                    try:
                        parsed = datetime.strptime(str(dval), '%Y-%m-%d')
                        month = parsed.month
                        year = parsed.year
                    except Exception:
                        pass

            if month is None or year is None:
                # fallback to current
                now = datetime.now()
                month = now.month if month is None else month
                year = now.year if year is None else year

            # create or get employee
            emp, created = Employee.objects.get_or_create(matricule=matr, defaults={
                'first_name': (str(name).split()[-1] if name else '')[:128],
                'last_name': (str(name).split()[0] if name else '')[:128],
                'salary_base': salary_base or 0,
            })
            if created:
                created_employees += 1

            # update salary_base if provided and different
            if salary_base is not None and emp.salary_base != (salary_base or 0):
                emp.salary_base = salary_base
                if not options['dry_run']:
                    emp.save()

            # create payroll
            if gross is None and net is None:
                # nothing to create
                skipped += 1
                continue

            # avoid duplicates enforced by unique_together
            exists = Payroll.objects.filter(employee=emp, month=month, year=year).first()
            if exists:
                self.stdout.write(self.style.WARNING(f'Payroll exists for {emp} {month}/{year}, skipping'))
                skipped += 1
                continue

            if not options['dry_run']:
                pr = Payroll.objects.create(employee=emp, month=month, year=year, gross_salary=(gross or 0), net_salary=(net or 0))
                created_payrolls += 1
            else:
                created_payrolls += 1

        self.stdout.write(self.style.SUCCESS(f'Import terminé: employés créés={created_employees}, fiches paie ajoutées={created_payrolls}, lignes ignorées={skipped}'))
