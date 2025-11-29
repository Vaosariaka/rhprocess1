from django.core.management.base import BaseCommand
from django.utils import timezone
from django.conf import settings
from openpyxl import load_workbook
import os
from datetime import date
from core.models import Employee

_CANDIDATES = [
    os.path.join(os.getcwd(), 'exports', 'FICHE_DE_PAIE.xlsx'),
    os.path.join(os.path.dirname(os.getcwd()), 'exports', 'FICHE_DE_PAIE.xlsx'),
    os.path.join(os.getcwd(), '..', 'exports', 'FICHE_DE_PAIE.xlsx'),
]

TEMPLATE_PATH = None
for p in _CANDIDATES:
    if os.path.exists(p):
        TEMPLATE_PATH = p
        break

class Command(BaseCommand):
    help = 'Generate a filled payslip XLSX from template for an employee (uses first employee if --employee not provided)'

    def add_arguments(self, parser):
        parser.add_argument('--employee', type=int, help='Employee id')
        parser.add_argument('--year', type=int, help='Year (defaults to current)')
        parser.add_argument('--month', type=int, help='Month (1-12, defaults to current)')
        parser.add_argument('--hours_sup_30', type=float, default=0.0)
        parser.add_argument('--hours_sup_50', type=float, default=0.0)
        parser.add_argument('--night_hours', type=float, default=0.0)
        parser.add_argument('--indemnites', type=float, default=0.0)

    def handle(self, *args, **options):
        if not os.path.exists(TEMPLATE_PATH):
            self.stderr.write(self.style.ERROR(f'Template not found at {TEMPLATE_PATH}'))
            return

        emp = None
        if options.get('employee'):
            emp = Employee.objects.filter(pk=options.get('employee')).first()
            if not emp:
                self.stderr.write(self.style.ERROR('Employee id provided not found'))
                return
        else:
            emp = Employee.objects.first()
            if not emp:
                self.stderr.write(self.style.ERROR('No employee found in DB'))
                return

        year = options.get('year') or timezone.now().year
        month = options.get('month') or timezone.now().month
        # compute period start/end
        start = date(year, month, 1)
        if month == 12:
            end = date(year + 1, 1, 1) - timezone.timedelta(days=1)
        else:
            end = date(year, month + 1, 1) - timezone.timedelta(days=1)

        wb = load_workbook(TEMPLATE_PATH)
        ws = wb['FICHE'] if 'FICHE' in wb.sheetnames else wb.active

        # mapping
        try:
            salary = float(getattr(emp, 'salary_base', 0) or 0)
        except Exception:
            salary = 0

        mapping = {
            'B16': f"{getattr(emp, 'last_name', '')} {getattr(emp, 'first_name', '')}".strip(),
            'B17': getattr(emp, 'matricule', ''),
            'B18': getattr(emp, 'function', ''),
            'B19': getattr(emp, 'cnaps_number', '') or 0,
            'B20': emp.hire_date.isoformat() if getattr(emp, 'hire_date', None) else '',
            'J16': salary,
            'F10': start.isoformat(),
            'N10': end.isoformat(),
            'H26': options.get('hours_sup_30'),
            'H27': options.get('hours_sup_50'),
            'H29': options.get('night_hours'),
            'J52': options.get('indemnites'),
        }

        for cell, val in mapping.items():
            try:
                ws[cell].value = val
            except Exception:
                pass

        # update _META
        if '_META' in wb.sheetnames:
            meta = wb['_META']
            # try to update the JSON string stored in cell A2 (handle Worksheet API correctly)
            try:
                import json
                # access A2 safely
                cell_a2 = meta['A2'] if 'A2' in meta else meta.cell(row=2, column=1)
                current = cell_a2.value or ''
                vars_obj = json.loads(current) if current else {}
                vars_obj.update({
                    'NOM': mapping['B16'],
                    'MATRICULE': mapping['B17'],
                    'FONCTION': mapping['B18'],
                    'CNAPS_NUM': mapping['B19'],
                    'DATE_EMBAUCHE': mapping['B20'],
                    'DATE_DEBUT': mapping['F10'],
                    'DATE_FIN': mapping['N10'],
                    'SALAIRE_BASE': mapping['J16'],
                })
                cell_a2.value = json.dumps(vars_obj)
            except Exception:
                pass

        # write output file
        out_dir = os.path.join(os.getcwd(), 'exports')
        os.makedirs(out_dir, exist_ok=True)
        fname = f"fiche_{emp.matricule}_{year}_{month:02d}.xlsx"
        out_path = os.path.join(out_dir, fname)
        wb.save(out_path)
        size = os.path.getsize(out_path)
        self.stdout.write(self.style.SUCCESS(f'Wrote filled payslip to {out_path} ({size} bytes)'))
