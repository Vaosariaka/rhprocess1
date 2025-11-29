from rest_framework import viewsets
from .models import (
    Category,
    Employee,
    Leave,
    Payroll,
    Competency,
    PerformanceReview,
    TrainingSuggestion,
    Message,
    ReplacementRequest,
    SuggestedReplacement,
    Absence,
    Presence,
    Contract,
    Alerte,
    LeaveHistory,
    Report,
    LeaveBalance,
)
from .serializers import CategorySerializer, EmployeeSerializer, LeaveSerializer, PayrollSerializer, CompetencySerializer
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status as drf_status
from .serializers import (
    ReplacementRequestSerializer,
    SuggestedReplacementSerializer,
    EmployeeCompetencySerializer,
    PerformanceReviewSerializer,
    TrainingSuggestionSerializer,
    MessageSerializer,
)
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from .permissions import IsHROrReadOnly, IsHRManager
try:
    from icalendar import Calendar, Event
    ICAL_AVAILABLE = True
except Exception:
    # icalendar might not be installed in all environments; fall back to
    # generating a minimal RFC5545-compatible .ics manually.
    Calendar = None
    Event = None
    ICAL_AVAILABLE = False

# DRF ViewSets (API)
class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer


class EmployeeViewSet(viewsets.ModelViewSet):
    queryset = Employee.objects.all().order_by('matricule')
    serializer_class = EmployeeSerializer


class LeaveViewSet(viewsets.ModelViewSet):
    queryset = Leave.objects.all().order_by('-start_date')
    serializer_class = LeaveSerializer


class PayrollViewSet(viewsets.ModelViewSet):
    queryset = Payroll.objects.all().order_by('-year', '-month')
    serializer_class = PayrollSerializer


class ReplacementRequestViewSet(viewsets.ModelViewSet):
    queryset = ReplacementRequest.objects.all().order_by('-created_at')
    serializer_class = ReplacementRequestSerializer
    permission_classes = [IsHROrReadOnly]

    def get_queryset(self):
        """Allow optional filtering by target employee, department or date range via query params.

        Supported query params: employee, target_employee, start_date, end_date, department
        Any filter errors are ignored to keep the endpoint robust.
        """
        qs = ReplacementRequest.objects.all().order_by('-created_at')
        params = getattr(self.request, 'query_params', {})
        emp = params.get('employee') or params.get('target_employee')
        dept = params.get('department')
        start = params.get('start_date')
        end = params.get('end_date')
        try:
            if emp:
                try:
                    qs = qs.filter(target_employee__pk=int(emp))
                except Exception:
                    qs = qs.filter(target_employee__matricule__icontains=str(emp))
        except Exception:
            pass
        try:
            if dept:
                qs = qs.filter(department__icontains=dept)
        except Exception:
            pass
        try:
            if start:
                qs = qs.filter(start_date__gte=start)
            if end:
                qs = qs.filter(end_date__lte=end)
        except Exception:
            pass
        return qs


class SuggestedReplacementViewSet(viewsets.ModelViewSet):
    queryset = SuggestedReplacement.objects.all().order_by('-created_at')
    serializer_class = SuggestedReplacementSerializer
    permission_classes = [IsHROrReadOnly]

    def get_queryset(self):
        """Support filtering by employee, suggested employee and date range via query params."""
        qs = SuggestedReplacement.objects.all().order_by('-created_at')
        params = getattr(self.request, 'query_params', {})
        emp = params.get('employee') or params.get('target_employee')
        sugg = params.get('suggested_employee') or params.get('candidate')
        start = params.get('start_date')
        end = params.get('end_date')
        try:
            if emp:
                try:
                    qs = qs.filter(target_employee__pk=int(emp))
                except Exception:
                    qs = qs.filter(target_employee__matricule__icontains=str(emp))
        except Exception:
            pass
        try:
            if sugg:
                try:
                    qs = qs.filter(suggested_employee__pk=int(sugg))
                except Exception:
                    qs = qs.filter(suggested_employee__last_name__icontains=str(sugg))
        except Exception:
            pass
        try:
            if start:
                qs = qs.filter(start_date__gte=start)
            if end:
                qs = qs.filter(end_date__lte=end)
        except Exception:
            pass
        return qs

    @action(detail=True, methods=['post'], permission_classes=[IsHRManager])
    def approve(self, request, pk=None):
        """Approve a suggested replacement via the API (HR/Manager only).

        POST /api/suggested-replacements/{id}/approve/
        """
        suggestion = self.get_object()
        try:
            suggestion.approve(request.user)
        except Exception as e:
            return Response({'error': str(e)}, status=drf_status.HTTP_500_INTERNAL_SERVER_ERROR)
        serializer = SuggestedReplacementSerializer(suggestion, context={'request': request})
        return Response(serializer.data)


class CompetencyViewSet(viewsets.ModelViewSet):
    queryset = Competency.objects.all().order_by('name')
    serializer_class = CompetencySerializer


class EmployeeCompetencyViewSet(viewsets.ModelViewSet):
    """CRUD for EmployeeCompetency mappings."""
    queryset = getattr(__import__('core.models', fromlist=['EmployeeCompetency']), 'EmployeeCompetency').objects.select_related('employee','competency').all().order_by('-created_at')
    serializer_class = EmployeeCompetencySerializer


class MatchCandidatesAPIView(APIView):
    """Match employees to a required competency set.

    POST JSON payload:
      {
        "competencies": [1,2,3]  # competency ids or names (strings), or objects with id/level
        "min_level": 3,          # optional, desired minimal level
        "max_results": 10
      }

    Returns: {candidates: [{id, matricule, name, match_count, score, matched_competencies: [{competency_id, required_level, employee_level}]}]}
    """
    def post(self, request, *args, **kwargs):
        """Enhanced matcher: supports competency-list and/or job_description text.

        Scoring strategy (simple heuristic):
        - competency matches: sum of employee competency levels for requested competencies
        - text similarity: cosine of simple term-frequency vectors between job_description and employee profile
        - final score = competency_score + int(text_similarity * 20)

        This keeps dependencies minimal (no sklearn). If no competencies provided, ranking falls back to text similarity.
        """
        data = request.data or {}
        raw = data.get('competencies') or []
        job_desc = (data.get('job_description') or data.get('description') or '').strip()
        min_level = int(data.get('min_level') or 0)
        max_results = int(data.get('max_results') or 20)

        # normalize competencies (ids/names)
        comp_ids = []
        comp_names = []
        for c in raw:
            if isinstance(c, int):
                comp_ids.append(c)
            elif isinstance(c, dict) and c.get('id'):
                try:
                    comp_ids.append(int(c.get('id')))
                except Exception:
                    pass
            elif isinstance(c, str):
                comp_names.append(c.strip())

        comps_q = Competency.objects.none()
        if comp_ids:
            comps_q = Competency.objects.filter(pk__in=comp_ids)
        if comp_names:
            comps_q = comps_q | Competency.objects.filter(name__in=comp_names)
        comps = list(comps_q)

        # prepare competency-based scores
        emp_scores = {}
        EC = getattr(__import__('core.models', fromlist=['EmployeeCompetency']), 'EmployeeCompetency')
        if comps:
            for ec in EC.objects.filter(competency__in=comps).select_related('employee', 'competency'):
                e = ec.employee
                ent = emp_scores.setdefault(e.pk, {'employee': e, 'matches': [], 'comp_score': 0})
                lvl = int(ec.level or 0)
                ent['matches'].append({'competency_id': ec.competency.pk, 'employee_level': lvl})
                ent['comp_score'] += lvl

        # text similarity helper (simple TF vector + cosine)
        import re, math
        def tokenize(text):
            if not text:
                return []
            text = text.lower()
            tokens = re.findall(r"\w+", text)
            return [t for t in tokens if len(t) > 1]

        def tf_vec(tokens):
            d = {}
            for t in tokens:
                d[t] = d.get(t, 0) + 1
            return d

        def cosine(a, b):
            if not a or not b:
                return 0.0
            num = 0
            for k, v in a.items():
                if k in b:
                    num += v * b[k]
            norm_a = math.sqrt(sum(v * v for v in a.values()))
            norm_b = math.sqrt(sum(v * v for v in b.values()))
            if norm_a == 0 or norm_b == 0:
                return 0.0
            return num / (norm_a * norm_b)

        job_tokens = tokenize(job_desc)
        job_tf = tf_vec(job_tokens)

        # For each active employee, compute combined score
        candidates = []
        emp_qs = Employee.objects.filter(is_active=True, archived=False)
        for e in emp_qs:
            comp_part = 0
            matches = []
            ent = emp_scores.get(e.pk)
            if ent:
                comp_part = int(ent.get('comp_score', 0))
                matches = ent.get('matches', [])

            # text part: build a profile text from available fields
            profile_parts = []
            try:
                profile_parts.append(str(getattr(e, 'function', '') or ''))
            except Exception:
                pass
            try:
                profile_parts.append(str(getattr(e, 'department', '') or ''))
            except Exception:
                pass
            # some projects have free-text fields like notes or profile_summary
            try:
                profile_parts.append(str(getattr(e, 'notes', '') or ''))
            except Exception:
                pass
            try:
                profile_parts.append(str(getattr(e, 'profile_summary', '') or ''))
            except Exception:
                pass
            prof_text = ' '.join(profile_parts).strip()
            prof_tf = tf_vec(tokenize(prof_text))
            text_sim = cosine(job_tf, prof_tf) if job_desc else 0.0

            total_score = comp_part + int(text_sim * 20)

            # if min_level is specified, ensure matched competencies meet it (if competencies requested)
            if comps and min_level > 0:
                # require that at least one matched competency has employee_level >= min_level
                ok = False
                for m in matches:
                    if m.get('employee_level', 0) >= min_level:
                        ok = True
                        break
                if not ok:
                    # skip candidate
                    continue

            candidates.append({
                'id': e.pk,
                'matricule': e.matricule or '',
                'name': f"{e.last_name} {e.first_name}".strip(),
                'match_count': len(matches),
                'score': int(total_score),
                'matched_competencies': matches,
                'text_similarity': round(text_sim, 3),
            })

        candidates = sorted(candidates, key=lambda x: (-x['score'], -x['match_count']))[:max_results]
        return Response({'candidates': candidates})


class CompetencyCartographyAPIView(APIView):
    """Return coverage metrics per competency: employee count, avg level and employees below threshold.

    GET params: min_level (optional, default 3) to list employees below that level.
    """
    def get(self, request, *args, **kwargs):
        min_level = int(request.GET.get('min_level') or 3)
        EC = getattr(__import__('core.models', fromlist=['EmployeeCompetency']), 'EmployeeCompetency')
        data = []
        for c in Competency.objects.all().order_by('name'):
            ecs = EC.objects.filter(competency=c).select_related('employee')
            count = ecs.count()
            levels = [int(x.level or 0) for x in ecs]
            avg = (sum(levels) / len(levels)) if levels else 0
            below = []
            for x in ecs:
                if int(x.level or 0) < min_level:
                    e = x.employee
                    below.append({'employee_id': e.pk, 'name': f"{e.last_name} {e.first_name}".strip(), 'level': int(x.level or 0)})
            data.append({'competency_id': c.pk, 'name': c.name, 'count': count, 'avg_level': round(avg, 2), 'below_min_level': below})
        return Response({'cartography': data})


class GenerateTrainingSuggestionsAPIView(APIView):
    """Generate training suggestions for employees with low competency levels.

    POST JSON payload: {"min_level": 3, "dry_run": true}
    Requires HR/manager permission for write actions.
    """
    permission_classes = [IsHRManager]

    def post(self, request, *args, **kwargs):
        data = request.data or {}
        min_level = int(data.get('min_level') or 3)
        dry_run = bool(data.get('dry_run') or data.get('dry') or False)
        created = []
        EC = getattr(__import__('core.models', fromlist=['EmployeeCompetency']), 'EmployeeCompetency')
        TS = TrainingSuggestion
        # For every competency, find employees with level < min_level or missing competency
        for c in Competency.objects.all():
            # employees with competency below threshold
            low_ecs = EC.objects.filter(competency=c).select_related('employee')
            for ec in low_ecs:
                if int(ec.level or 0) < min_level:
                    e = ec.employee
                    exists = TS.objects.filter(employee=e, competency=c, status='PENDING').exists()
                    if not exists:
                        if not dry_run:
                            try:
                                TS.objects.create(employee=e, competency=c, recommended_level=min_level, reason='level_below_threshold')
                            except Exception:
                                pass
                        created.append({'employee_id': e.pk, 'competency_id': c.pk, 'action': 'create' if not dry_run else 'dry'})
            # employees missing the competency entirely
            emp_with = set(EC.objects.filter(competency=c).values_list('employee_id', flat=True))
            for e in Employee.objects.filter(is_active=True, archived=False).exclude(pk__in=emp_with):
                exists = TS.objects.filter(employee=e, competency=c, status='PENDING').exists()
                if not exists:
                    if not dry_run:
                        try:
                            TS.objects.create(employee=e, competency=c, recommended_level=min_level, reason='missing_competency')
                        except Exception:
                            pass
                    created.append({'employee_id': e.pk, 'competency_id': c.pk, 'action': 'create' if not dry_run else 'dry'})

        return Response({'created_suggestions': len(created), 'details': created})


class PerformanceReviewViewSet(viewsets.ModelViewSet):
    queryset = PerformanceReview.objects.all().order_by('-review_date')
    serializer_class = PerformanceReviewSerializer


class TrainingSuggestionViewSet(viewsets.ModelViewSet):
    queryset = TrainingSuggestion.objects.all().order_by('-created_at')
    serializer_class = TrainingSuggestionSerializer


class MessageViewSet(viewsets.ModelViewSet):
    queryset = Message.objects.all().order_by('-created_at')
    serializer_class = MessageSerializer
    lookup_value_regex = r'\d+'


class SuggestReplacementAPIView(APIView):
    """API: suggest replacement candidates for a given period.

    POST payload (JSON): {"start_date": "YYYY-MM-DD", "end_date": "YYYY-MM-DD", "department": "...", "function": "...", "exclude_employee": id, "max_results": 10}

    Returns: {"candidates": [{id, matricule, name, department, function, email, score}, ...]}
    """
    def post(self, request, *args, **kwargs):
        data = request.data or {}
        start_str = data.get('start_date')
        end_str = data.get('end_date')
        dept = data.get('department') or data.get('department_hint')
        func = data.get('function') or data.get('function_hint')
        exclude_id = data.get('exclude_employee') or data.get('target_employee')
        try:
            from datetime import datetime
            start = datetime.strptime(start_str, '%Y-%m-%d').date() if start_str else None
            end = datetime.strptime(end_str, '%Y-%m-%d').date() if end_str else None
        except Exception:
            return Response({'error': 'Invalid date format. Use YYYY-MM-DD.'}, status=drf_status.HTTP_400_BAD_REQUEST)

        if not start or not end or start > end:
            return Response({'error': 'Provide valid start_date and end_date (start <= end).'}, status=drf_status.HTTP_400_BAD_REQUEST)

        qs = Employee.objects.filter(is_active=True, archived=False)
        if dept:
            qs = qs.filter(department__icontains=dept)
        if func:
            qs = qs.filter(function__icontains=func)
        if exclude_id:
            try:
                qs = qs.exclude(pk=int(exclude_id))
            except Exception:
                pass

        # simple candidate evaluation: exclude employees with approved leave overlapping or absences
        candidates = []
        max_results = int(data.get('max_results') or 10)
        for e in qs.order_by('last_name', 'first_name'):
            # exclude those with approved leave overlapping period
            on_leave = Leave.objects.filter(employee=e, status='APPROVED', start_date__lte=end, end_date__gte=start).exists()
            if on_leave:
                continue
            # exclude unjustified absences during the period
            has_absence = Absence.objects.filter(employee=e, date__gte=start, date__lte=end, justified=False).exists()
            if has_absence:
                continue
            # candidate ok
            score = 100
            candidates.append({
                'id': e.pk,
                'matricule': e.matricule or '',
                'name': f"{e.last_name} {e.first_name}".strip(),
                'department': e.department or '',
                'function': e.function or '',
                'email': e.email or '',
                'score': score,
            })
            if len(candidates) >= max_results:
                break

        return Response({'candidates': candidates})


class CreateReplacementRequestAPIView(APIView):
    """Create a ReplacementRequest via API (HR/manager usage).

    POST JSON: {"target_employee": id, "start_date": "YYYY-MM-DD", "end_date": "YYYY-MM-DD", "department_hint": "", "function_hint":"", "notes":""}
    """
    def post(self, request, *args, **kwargs):
        user = request.user if request.user.is_authenticated else None
        data = request.data or {}
        # Basic validation is handled by serializer
        serializer = ReplacementRequestSerializer(data=data)
        if serializer.is_valid():
            rr = serializer.save(requester=user)
            return Response({'result': 'ok', 'id': rr.id}, status=drf_status.HTTP_201_CREATED)
        return Response({'error': serializer.errors}, status=drf_status.HTTP_400_BAD_REQUEST)


def export_leaves_ical(request, pk=None):
    """Return an iCalendar (.ics) of upcoming approved leaves.

    If `pk` is provided, return leaves for that employee; otherwise return all approved leaves.
    """
    from datetime import datetime

    q = Leave.objects.filter(status='APPROVED').order_by('start_date')
    if pk:
        q = q.filter(employee__pk=pk)
    if ICAL_AVAILABLE:
        cal = Calendar()
        cal.add('prodid', '-//HRMS//Leaves Calendar//FR')
        cal.add('version', '2.0')
        for l in q:
            ev = Event()
            ev.add('summary', f"Congé: {l.employee.last_name} {l.employee.first_name} - {l.leave_type}")
            ev.add('dtstart', datetime(l.start_date.year, l.start_date.month, l.start_date.day))
            ev.add('dtend', datetime(l.end_date.year, l.end_date.month, l.end_date.day) + relativedelta(days=1))
            ev.add('description', l.note or '')
            ev.add('uid', f"leave-{l.pk}@hrms.local")
            cal.add_component(ev)
        resp = HttpResponse(cal.to_ical(), content_type='text/calendar')
    else:
        # Build a minimal ICS string by hand
        lines = [
            'BEGIN:VCALENDAR',
            'PRODID:-//HRMS//Leaves Calendar//FR',
            'VERSION:2.0',
        ]
        for l in q:
            uid = f"leave-{l.pk}@hrms.local"
            dtstart = l.start_date.strftime('%Y%m%d')
            # dtend is exclusive in iCal; add one day
            dtend_date = l.end_date + relativedelta(days=1)
            dtend = dtend_date.strftime('%Y%m%d')
            summary = f"Congé: {l.employee.last_name} {l.employee.first_name} - {l.leave_type}"
            desc = (l.note or '').replace('\n', '\\n')
            lines += [
                'BEGIN:VEVENT',
                f'UID:{uid}',
                f'DTSTAMP:{datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")}',
                f'DTSTART;VALUE=DATE:{dtstart}',
                f'DTEND;VALUE=DATE:{dtend}',
                f'SUMMARY:{summary}',
                f'DESCRIPTION:{desc}',
                'END:VEVENT',
            ]
        lines.append('END:VCALENDAR')
        resp = HttpResponse('\r\n'.join(lines).encode('utf-8'), content_type='text/calendar')
    filename = f"leaves_{pk or 'all'}.ics"
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


# Simple web views (Django templates)
from django.views.generic import ListView, DetailView
from django.views.generic import CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.http import HttpResponse, FileResponse
import openpyxl
from io import BytesIO
from decimal import Decimal, ROUND_HALF_UP
try:
    import pandas as pd
except Exception:
    pd = None
from django.shortcuts import render, redirect, get_object_or_404
from django.conf import settings
from django.db.models import Sum, Avg, Count
from .calculators import PayrollCalculator
from django.core.management import call_command
from io import StringIO
from pathlib import Path
import uuid
import os
from django.http import JsonResponse
import json
from pathlib import Path
from django.utils import timezone
from django.db import connection
from django.db.models import Q
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import Group
from django.views.decorators.http import require_POST
from django.core.mail import send_mail
from .models import Contract, Alerte, LeaveHistory, Presence, Absence
from .models import Contract, Alerte, LeaveHistory, Presence, Absence, SuggestedReplacement
from django.template.loader import render_to_string
from django.urls import reverse
import tempfile
import subprocess
import shutil
import logging
from .signals import write_employees_to_fiche
from datetime import datetime, timedelta, date
from dateutil.relativedelta import relativedelta
import re
import calendar
from rest_framework.parsers import MultiPartParser, FormParser
from .serializers import EmployeeSelfSerializer, MessageSerializer

MONTH_LABELS = [
    '', 'Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin',
    'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre'
]


def _local_today():
    try:
        return timezone.localdate()
    except Exception:
        return date.today()


def _format_currency(value):
    try:
        return f"{float(value):,.0f}".replace(',', ' ')
    except (TypeError, ValueError):
        return '0'


def _format_percentage(value, digits=1):
    try:
        return f"{float(value):.{digits}f}%"
    except (TypeError, ValueError):
        return '0%'


def _build_daily_status(today):
    statuses = {
        'working': {
            'label': "Au travail aujourd'hui",
            'color': '#16a34a',
            'employees': [],
            'ids': set(),
        },
        'paid_leave': {
            'label': 'En congé payé',
            'color': '#0284c7',
            'employees': [],
            'ids': set(),
        },
        'sick_leave': {
            'label': 'Congé maladie / maternité',
            'color': '#facc15',
            'employees': [],
            'ids': set(),
        },
        'unjustified_absence': {
            'label': 'Absence injustifiée',
            'color': '#dc2626',
            'employees': [],
            'ids': set(),
        },
    }

    def add_employee(status_key, employee, reason):
        if not employee or employee.id in statuses[status_key]['ids']:
            return
        statuses[status_key]['ids'].add(employee.id)
        statuses[status_key]['employees'].append({
            'id': employee.id,
            'matricule': employee.matricule,
            'name': employee.get_full_name() if hasattr(employee, 'get_full_name') else str(employee),
            'function': employee.function,
            'department': employee.department,
            'reason': reason,
        })

    leaves_today = Leave.objects.filter(status='APPROVED', start_date__lte=today, end_date__gte=today).select_related('employee')
    for leave in leaves_today:
        employee = leave.employee
        if not employee or not employee.is_active or employee.archived:
            continue
        bucket = 'paid_leave'
        if leave.leave_type in ['SICK', 'MAT']:
            bucket = 'sick_leave'
        add_employee(bucket, employee, leave.get_leave_type_display() if hasattr(leave, 'get_leave_type_display') else leave.leave_type)

    absences_today = Absence.objects.filter(date=today).select_related('employee')
    for absence in absences_today:
        employee = absence.employee
        if not employee or not employee.is_active or employee.archived:
            continue
        if not absence.justified:
            add_employee('unjustified_absence', employee, absence.reason or 'Absence injustifiée')

    leave_conflict_ids = statuses['paid_leave']['ids'] | statuses['sick_leave']['ids'] | statuses['unjustified_absence']['ids']

    presence_today = Presence.objects.filter(date=today).select_related('employee')
    for presence in presence_today:
        employee = presence.employee
        if not employee or not employee.is_active or employee.archived:
            continue
        if employee.id in leave_conflict_ids:
            continue
        if (presence.worked_minutes or 0) > 0 or presence.time_in or presence.time_out:
            hours = (presence.worked_minutes or 0) / 60.0
            reason = f"Présent ({hours:.1f} h)" if hours else 'Présent'
            add_employee('working', employee, reason)

    for key, data in statuses.items():
        data['count'] = len(data['employees'])
        data.pop('ids', None)

    return statuses




class EmployeeListView(ListView):
    model = Employee
    template_name = 'core/employee_list.html'
    context_object_name = 'employees'
    paginate_by = 40

    def get_queryset(self):
        search = (self.request.GET.get('q') or '').strip()
        self.search_query = search
        qs = (
            Employee.objects.select_related('category')
            .prefetch_related('contracts')
            .order_by('last_name', 'first_name')
        )
        if search:
            qs = qs.filter(
                Q(last_name__icontains=search)
                | Q(first_name__icontains=search)
                | Q(matricule__icontains=search)
                | Q(function__icontains=search)
                | Q(department__icontains=search)
            )
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['search_query'] = getattr(self, 'search_query', '')
        ctx['active_employee_count'] = Employee.objects.filter(is_active=True, archived=False).count()
        ctx['archived_employee_count'] = Employee.objects.filter(archived=True).count()
        paginator = ctx.get('paginator')
        ctx['result_count'] = paginator.count if paginator else len(ctx.get('employees', []))
        return ctx


class EmployeeDetailView(DetailView):
    model = Employee
    template_name = 'core/employee_detail.html'
    context_object_name = 'employee'

    def get_queryset(self):
        return (
            Employee.objects.select_related('category')
            .prefetch_related('contracts', 'documents', 'position_histories', 'contract_histories')
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        employee = self.object
        ctx['active_contract'] = employee.get_active_contract()
        ctx['contract_badge'] = employee.get_contract_badge()
        ctx['contract_status'] = employee.get_contract_status_banner()
        ctx['salary_summary'] = employee.get_salary_summary()
        ctx['upcoming_deadlines'] = employee.get_upcoming_deadlines(limit=3)
        ctx['position_history'] = employee.position_histories.order_by('-effective_date', '-created_at')
        ctx['contract_history'] = employee.contract_histories.order_by('-date_action')
        ctx['can_manage_contract'] = bool(getattr(self.request.user, 'is_staff', False))
        return ctx


class LeaveListView(ListView):
    model = Leave
    template_name = 'core/leave_list.html'
    context_object_name = 'leaves'
    paginate_by = 50
    def get_queryset(self):
        return Leave.objects.select_related('employee').order_by('-start_date')
    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        # allow template to show approve/reject buttons when user is manager
        user = getattr(self.request, 'user', None)
        ctx['can_approve'] = user_is_manager(user)
        return ctx


@login_required
@require_POST
def employee_contract_action(request, pk, action):
    employee = get_object_or_404(Employee, pk=pk)
    redirect_url = reverse('employee_detail', args=[employee.pk])
    if not getattr(request.user, 'is_staff', False):
        messages.error(request, "Vous n'avez pas les droits pour cette action.")
        return redirect(redirect_url)

    contract = employee.get_active_contract()
    if not contract:
        messages.error(request, "Aucun contrat actif pour cet employé.")
        return redirect(redirect_url)

    if action == 'renew_trial':
        if contract.type != 'ESSAI':
            messages.warning(request, "La période d'essai n'est plus active.")
        elif contract.renew_trial():
            messages.success(request, "Période d'essai renouvelée de 6 mois.")
        else:
            messages.error(request, "Impossible de renouveler l'essai (limite atteinte).")
    elif action == 'convert_cdi':
        if contract.type == 'CDI':
            messages.info(request, "Le contrat est déjà en CDI.")
        else:
            contract.convert_to_cdi()
            messages.success(request, "Contrat converti en CDI.")
    elif action == 'terminate':
        contract.terminate(date_termination=date.today(), reason='Terminé via portail RH')
        messages.success(request, "Contrat marqué comme terminé.")
    else:
        messages.error(request, "Action non reconnue.")
    return redirect(redirect_url)

def calculate_seniority(hire_date):
    """Retourne l'ancienneté au format exact du modèle : '14 an(s) 7 mois et 12 jour(s)'"""
    if not hire_date:
        return ""
    today = datetime.today().date()
    delta = relativedelta(today, hire_date)
    years = delta.years
    months = delta.months
    days = delta.days
    parts = []
    if years:
        parts.append(f"{years} an{'s' if years > 1 else ''}")
    if months:
        parts.append(f"{months} mois")
    if days:
        parts.append(f"{days} jour{'s' if days > 1 else ''}")
    return " ".join(parts) if parts else "0 jour"


def format_excel_date(date_obj):
    """Convertit une date Python en nombre Excel (comme 45961)"""
    if not date_obj:
        return ""
    try:
        return (date_obj - datetime(1899, 12, 30).date()).days
    except:
        return ""

def export_employees_xlsx(request):
    """Export employees to an Excel file generated on the fly."""
    headers = ['Matricule', 'Last Name', 'First Name', 'CNAPS', 'Category', 'Function', 'Hire Date', 'Salary Base']
    non_empty = request.GET.get('non_empty') in ('1', 'true', 'True')
    data = []
    for e in Employee.objects.select_related('category').all().order_by('matricule'):
        matr = (e.matricule or '')
        if non_empty and (not matr or 'MATR' in matr.upper() or 'N\u00B0MATR' in matr.upper()):
            continue
        row = {
            'Matricule': e.matricule or '',
            'Last Name': e.last_name or '',
            'First Name': e.first_name or '',
            'CNAPS': e.cnaps_number or '',
            'Category': str(e.category) if e.category else '',
            'Function': e.function or '',
            'Hire Date': e.hire_date.isoformat() if e.hire_date else '',
            'Salary Base': float(e.salary_base or 0),
        }
        if non_empty:
            vals = [row[k] for k in headers[1:]]
            if all((v is None or (isinstance(v, str) and v.strip() == '') or v == 0) for v in vals):
                continue
        data.append(row)

    # create excel using pandas if available (clean), otherwise openpyxl
    if pd is not None:
        df = pd.DataFrame(data, columns=headers)
        bio = BytesIO()
        with pd.ExcelWriter(bio, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Employees')
        bio.seek(0)
        resp = HttpResponse(bio.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="employees.xlsx"'
        return resp
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Employees'
        ws.append(headers)
        for r in data:
            ws.append([r[h] for h in headers])
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        resp = HttpResponse(bio.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="employees.xlsx"'
        return resp

def export_employee_fiche(request, pk):
    """Génère la fiche de paie EXACTEMENT comme ton modèle Excel"""
    try:
        employee = Employee.objects.select_related('category').get(pk=pk)
    except Employee.DoesNotExist:
        return HttpResponse('Employé introuvable', status=404)

    # Récupérer la paie du mois demandé ou la plus récente
    year_q = request.GET.get('year')
    month_q = request.GET.get('month')
    payroll = None
    if year_q and month_q:
        try:
            y, m = int(year_q), int(month_q)
            payroll = Payroll.objects.filter(employee=employee, year=y, month=m).first()
        except:
            pass
    if not payroll:
        payroll = Payroll.objects.filter(employee=employee).order_by('-year', '-month').first()

    # Données calculées
    hire_date_excel = format_excel_date(employee.hire_date)
    seniority_text = calculate_seniority(employee.hire_date)

    # Use PayrollCalculator to compute all payroll elements in one place
    salaire_brut = float(employee.salary_base or 0)
    calculs = PayrollCalculator.generer_fiche_paie_complete(salaire_brut)

    base_salary = salaire_brut
    daily_rate = calculs.get('taux_journalier', 0)
    hourly_rate = calculs.get('taux_horaire', 0)

    # Paie du mois (computed)
    gross = calculs.get('salaire_brut', salaire_brut)
    cnaps_1 = calculs.get('cnaps_salarie', 0)
    sanitaire = calculs.get('sanitaire_salarie', 0)
    irsa = calculs.get('irsa', 0)
    total_retenues = calculs.get('total_retenues', 0)
    net = calculs.get('salaire_net', gross - total_retenues)
    base_imposable = calculs.get('base_imposable_irsa', 0)
    irsa_parts = []

    # Chargement du modèle
    base = Path(settings.BASE_DIR)
    template_paths = [
        base.parent / 'FICHE DE PAIE .xlsx',
        base.parent / 'exports' / 'FICHE DE PAIE .xlsx',
        base / 'FICHE DE PAIE .xlsx',
    ]
    template_path = None
    for p in template_paths:
        if p.exists():
            template_path = p
            break

    if not template_path:
        return HttpResponse("Modèle 'FICHE DE PAIE .xlsx' introuvable", status=500)

    # Ouvrir et modifier
    wb = openpyxl.load_workbook(filename=str(template_path))
    ws = wb.active

    # === BEGIN: fill exact template cells per full fiche specification ===
    try:
        # Section 1: En-tête et métadonnées
        if payroll:
            period_start = datetime(payroll.year, payroll.month, 1).date()
            period_end = (period_start + relativedelta(months=1)) - relativedelta(days=1)
        else:
            period_start = None
            period_end = None

        # L1 / U1
        try:
            ws['L1'] = 'nb jour/mois'
            ws['U1'] = '=365/12'
        except Exception:
            pass

        # L2 block
        try:
            ws['M2'] = 'Clôture'
            ws['O2'] = 'Début'
            ws['P2'] = period_start if period_start else ''
            ws['S2'] = 'Ancienneté en jours'
            ws['T2'] = 'Partie en années'
            ws['U2'] = 'Partie en mois'
            ws['V2'] = 'Partie en jours'
            ws['W2'] = 'Ancienneté'
        except Exception:
            pass

        # L10: period end and ancienneté breakdown
        try:
            ws['M10'] = period_end if period_end else ''
            ws['O10'] = 'Fin'
            ws['P10'] = period_end if period_end else ''
            # ancienneté en jours : date_fin - date_embauche + 1
            if employee.hire_date and period_end:
                anciennete_jours = (period_end - employee.hire_date).days + 1
            else:
                anciennete_jours = 0
            ws['S10'] = int(anciennete_jours)
            t_years = int(anciennete_jours / 365) if anciennete_jours else 0
            ws['T10'] = int(t_years)
            t_months = int((anciennete_jours - (t_years * 365)) / 30.4167) if anciennete_jours else 0
            ws['U10'] = int(t_months)
            t_days_rest = anciennete_jours - (t_years * 365) - int(t_months * 30.4167) if anciennete_jours else 0
            ws['V10'] = int(t_days_rest)
            ws['W10'] = f"{t_years} an(s) {t_months} mois {int(t_days_rest)} jour(s)"
        except Exception:   
            pass

        # Section 2: Informations employé
        try:
            ws['C12'] = 'FICHE DE PAIE'

            ws['D14'] = 'Nom et Prénoms :'
            ws['F14'] = f"{employee.last_name} {employee.first_name}".strip()

            ws['D15'] = 'Matricule :'
            ws['F15'] = employee.matricule or ''
            ws['H15'] = 'Classification :'
            ws['J15'] = str(employee.category) if employee.category else ''

            ws['D16'] = 'Fonction :'
            ws['F16'] = employee.function or ''
            ws['H16'] = 'Salaire de base :'
            ws['J16'] = float(base_salary)
            try:
                ws['J16'].number_format = '#,##0'
            except Exception:
                pass

            ws['D17'] = 'N° CNaPS :'
            ws['F17'] = employee.cnaps_number or ''
            ws['H17'] = 'Taux journaliers :'
            # J17 = formula (template already has it, don't overwrite)

            ws['D18'] = "Date d'embauche :"
            ws['F18'] = employee.hire_date if employee.hire_date else ''
            ws['H18'] = 'Taux horaires :'
            # J18 = formula (template already has it, don't overwrite)

            ws['D19'] = 'Ancienneté :'
            ws['F19'] = seniority_text
            ws['H19'] = 'Indice :'
            # J19 = formula (template already has it, don't overwrite)
        except Exception:
            pass

        # Section 3: Base congés (M16..M19)
        try:
            ws['M16'] = float(base_salary)
            ws['M17'] = round(ws['M16'].value / 30, 0) if ws['M16'].value not in (None, '') else 0
            ws['M18'] = round(ws['M16'].value / 173.33, 0) if ws['M16'].value not in (None, '') else 0
            try:
                ws['M19'] = round(ws['M17'].value / 1.334, 0)
            except Exception:
                ws['M19'] = ''
        except Exception:
            pass

        # Section 4: Tableau des gains (headers + period line)
        try:
            # Headers are already in template, only write period text
            # Period text DD/MM/AA
            if period_start and period_end:
                period_text = f"Salaire du {period_start.strftime('%d/%m/%y')} au {period_end.strftime('%d/%m/%y')}"
            else:
                period_text = ''
            ws['D23'] = period_text

            # Nombre jours travaillés (try common field names on payroll)
            def get_payroll_field(p, candidates, default=0):
                for n in candidates:
                    if hasattr(p, n):
                        return getattr(p, n) or default
                return default

            jours_travailles = get_payroll_field(payroll, ['days_worked', 'worked_days', 'nbr_days', 'nb_days', 'number_of_days'], 0) if payroll else 0
            ws['H23'] = int(jours_travailles) if jours_travailles else '1 mois'
            # taux journalier in I23, montant in J23 (template has formulas)
            # We only need to set H23 (nombre), template formulas will handle rest

            # Absences déductibles: set H24 (nombre) only
            # Template already has I24=$J$17 and J24=H24*I24

            # Heures sup: set H27..H31 (nombre) only, template has formulas
            # We can optionally write nombre values if payroll has them00
            # For now, leave template formulas intact

            # Droits de congés: set H34 (nombre) only
            # Template already has formulas

            # Don't overwrite formulas - template already has them!
        except Exception:
            pass

        # Section 5: Retenues
        try:
            # Write M40 plafond CNAPS (from parameters)
            try:
                ws['M40'] = PayrollCalculator.PLAFOND_CNAPS * PayrollCalculator.TAUX_CNAPS_SALARIE
            except Exception:
                ws['M40'] = PayrollCalculator.PLAFOND_CNAPS

            # Template has J40 (CNaPS) and J41 (Sanitaire) formulas - keep them

            # IRSA: Template does NOT have formulas in J42-J47
            # Use the calculator's base_imposable and compute tranche values deterministically
            tranche_bounds = [350000, 400000, 500000, 600000, 4000000]
            imposable = max(base_imposable if base_imposable else (gross - cnaps_1 - sanitaire), 0)
            total_irsa = int(round(irsa if irsa is not None else 0))

            # Extend bounds with +inf so each template row matches one bracket
            extended_bounds = tranche_bounds + [float('inf')]
            irsa_values = [0] * len(extended_bounds)

            if total_irsa > 0 and imposable > 0:
                active_index = next((i for i, bound in enumerate(extended_bounds) if imposable <= bound), len(extended_bounds) - 1)
                irsa_values[active_index] = total_irsa

            # Write IRSA values in J42-J47
            for idx, val in enumerate(irsa_values, start=42):
                try:
                    ws[f'J{idx}'] = val
                    ws[f'J{idx}'].number_format = '#,##0'
                except Exception:
                    pass

            # J49 = Total IRSA (sum of tranches)
            ws['J49'] = '=SUM(J42:J47)'

            # J51 = Total retenues (CNaPS + Sanitaire + IRSA)
            ws['J51'] = '=J49+J40+J41'
        except Exception:
            pass

        # Section 6: Net à payer
        try:
            # Nettoyer l'ancienne cellule (ligne 53) pour éviter un doublon
            ws['H53'] = ''
            ws['I53'] = ''
            ws['J53'] = ''

            # J54 = Net à payer = Brut - Retenues + Autres (selon le modèle)
            ws['H54'] = ''
            ws['I54'] = 'Net à payer'
            ws['J54'] = '=J38-J51+J52'
        except Exception:
            pass

        # Section 7: Informations complémentaires
        try:
            # Template already has labels and formulas, don't overwrite
            # F59 = formula for montant imposable (already in template)
            # F62 = mode paiement (user can fill manually or leave empty)
            pass
        except Exception:
            pass
    except Exception:
        # be defensive: don't break the exporter if any cell write fails
        pass
    # === END: full fiche mapping ===
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"fiche_{employee.matricule or employee.pk}_{payroll.year}_{payroll.month}.xlsx" if payroll else f"fiche_{employee.matricule or employee.pk}.xlsx"

    response = HttpResponse(
        bio.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

def export_all_fiches(request):
    """Export all employees fiches into a single workbook (one sheet per employee)."""
    employees = Employee.objects.all().order_by('matricule')
    bio = BytesIO()
    if pd is not None:
        with pd.ExcelWriter(bio, engine='openpyxl') as writer:
            for e in employees:
                p = Payroll.objects.filter(employee=e).order_by('-year', '-month').first()
                rows = [
                    ('Matricule', e.matricule or ''),
                    ('Last name', e.last_name or ''),
                    ('First name', e.first_name or ''),
                    ('Category', str(e.category) if e.category else ''),
                    ('Function', e.function or ''),
                    ('Hire date', e.hire_date.isoformat() if e.hire_date else ''),
                    ('Salary base', float(e.salary_base or 0)),
                ]
                if p:
                    rows += [
                        ('Payroll year', p.year),
                        ('Payroll month', p.month),
                        ('Gross', float(p.gross_salary or 0)),
                        ('Net', float(p.net_salary or 0)),
                    ]
                df = pd.DataFrame(rows, columns=['Field', 'Value'])
                sheet_name = str(e.matricule or f'emp_{e.pk}')[:31]
                df.to_excel(writer, index=False, sheet_name=sheet_name)
        bio.seek(0)
        resp = HttpResponse(bio.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="all_fiches.xlsx"'
        return resp


    else:
        wb = openpyxl.Workbook()
        first = True
        for e in employees:
            if first:
                ws = wb.active
                ws.title = str(e.matricule or f'emp_{e.pk}')[:31]
                first = False
            else:
                ws = wb.create_sheet(title=str(e.matricule or f'emp_{e.pk}')[:31])
            rows = [
                ('Matricule', e.matricule or ''),
                ('Last name', e.last_name or ''),
                ('First name', e.first_name or ''),
                ('Category', str(e.category) if e.category else ''),
                ('Function', e.function or ''),
                ('Hire date', e.hire_date.isoformat() if e.hire_date else ''),
                ('Salary base', float(e.salary_base or 0)),
            ]
            p = Payroll.objects.filter(employee=e).order_by('-year', '-month').first()
            if p:
                rows += [
                    ('Payroll year', p.year),
                    ('Payroll month', p.month),
                    ('Gross', float(p.gross_salary or 0)),
                    ('Net', float(p.net_salary or 0)),
                ]
            ws.append(['Field', 'Value'])
            for r in rows:
                ws.append([r[0], r[1]])
        wb.save(bio)
        bio.seek(0)
        resp = HttpResponse(bio.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="all_fiches.xlsx"'
        return resp


def employees_with_payments(request):
    """Return all employees with their latest payroll (if any).
    JSON structure: { data: [ { employee_id, matricule, first_name, last_name, category, salary_base, latest_payroll: {year,month,gross_salary,net_salary} | null }, ... ] }
    """
    employees = Employee.objects.all().order_by('matricule')
    data = []
    for e in employees:
        p = Payroll.objects.filter(employee=e).order_by('-year', '-month').first()
        latest = None
        if p:
            latest = {
                'id': p.pk,
                'year': p.year,
                'month': p.month,
                'gross_salary': float(p.gross_salary or 0),
                'net_salary': float(p.net_salary or 0),
            }
        data.append({
            'employee_id': int(e.pk),
            'matricule': e.matricule or '',
            'first_name': e.first_name or '',
            'last_name': e.last_name or '',
            'category': str(e.category) if e.category else '',
            'salary_base': float(e.salary_base or 0),
            'latest_payroll': latest,
        })

    return JsonResponse({'data': data})


def chatbot_view(request):
    """Simple FAQ chatbot endpoint.
    GET: return list of available FAQ topics
    POST: JSON {"question": "..."} -> returns JSON {"answer": "...", "matched": true/false}
    """
    base = Path(__file__).resolve().parent
    faq_path = base / 'faq.json'
    faq = []
    try:
        if faq_path.exists():
            with open(faq_path, 'r', encoding='utf-8') as fh:
                faq = json.load(fh)
    except Exception:
        faq = []

    if request.method == 'GET':
        # return topics only (short)
        topics = [item.get('question') for item in faq]
        return JsonResponse({'topics': topics})

    # POST: expect JSON payload
    try:
        payload = json.loads(request.body.decode('utf-8') or '{}')
    except Exception:
        payload = {}
    question = (payload.get('question') or '').strip()
    if not question:
        return JsonResponse({'answer': "Posez une question sur la paie ou les congés, par exemple 'Comment poser un congé ?'", 'matched': False})

    q_lower = question.lower()

    # optional context: employee_id (string or int) and optional date
    employee_obj = None
    emp_id = payload.get('employee_id') or payload.get('employee')
    if emp_id:
        try:
            employee_obj = Employee.objects.get(pk=int(emp_id))
        except Exception:
            employee_obj = None
    # support multiple or "all" employees
    employees_list = None
    try:
        raw_emps = payload.get('employees') or payload.get('employee_ids')
        if raw_emps:
            # accept list or comma-separated string
            if isinstance(raw_emps, list):
                ids = [int(x) for x in raw_emps if x]
            else:
                ids = [int(x.strip()) for x in str(raw_emps).split(',') if x.strip()]
            employees_list = list(Employee.objects.filter(pk__in=ids))
        else:
            if isinstance(emp_id, str) and emp_id.lower() in ('all', 'tous', 'tout'):
                employees_list = list(Employee.objects.all())
            elif emp_id and isinstance(emp_id, str) and ',' in emp_id:
                try:
                    ids = [int(x.strip()) for x in emp_id.split(',') if x.strip()]
                    employees_list = list(Employee.objects.filter(pk__in=ids))
                except Exception:
                    employees_list = None
    except Exception:
        employees_list = None
    # try to parse an optional date (ISO yyyy-mm-dd or dd/mm/yyyy)
    date_obj = None
    date_str = payload.get('date') or payload.get('date_str')
    if date_str:
        try:
            # try ISO first
            date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
        except Exception:
            try:
                date_obj = datetime.strptime(date_str, '%d/%m/%Y').date()
            except Exception:
                date_obj = None

    def format_int(value):
        try:
            return f"{int(value):,}".replace(',', ' ')
        except (TypeError, ValueError):
            return str(value)

    def format_date(value):
        if not value:
            return '—'
        try:
            return value.strftime('%d/%m/%Y')
        except Exception:
            try:
                return value.isoformat()
            except Exception:
                return str(value)

    contract_label_map = dict(Contract.CONTRACT_TYPES)
    base_employee_qs = Employee.objects.filter(archived=False)
    month_alias_map = {
        'janvier': 1, 'janv': 1, 'jan': 1, 'january': 1, 'janvr': 1,
        'février': 2, 'fevrier': 2, 'févr': 2, 'feb': 2, 'february': 2, 'fevr': 2,
        'mars': 3, 'march': 3,
        'avril': 4, 'avr': 4, 'april': 4,
        'mai': 5, 'may': 5,
        'juin': 6, 'june': 6,
        'juillet': 7, 'juil': 7, 'july': 7,
        'août': 8, 'aout': 8, 'august': 8, 'aug': 8,
        'septembre': 9, 'sept': 9, 'september': 9, 'septembr': 9,
        'octobre': 10, 'oct': 10, 'october': 10, 'octobr': 10,
        'novembre': 11, 'nov': 11, 'november': 11, 'novembr': 11,
        'décembre': 12, 'decembre': 12, 'dec': 12, 'december': 12, 'decembr': 12,
    }

    def _month_range(year, month):
        last_day = calendar.monthrange(year, month)[1]
        return date(year, month, 1), date(year, month, last_day)

    def _is_full_month(period_tuple):
        if not period_tuple:
            return False
        start, end, _ = period_tuple
        if not start or not end:
            return False
        same_month = start.month == end.month and start.year == end.year
        enough_days = (end - start).days >= 27
        return start.day == 1 and same_month and enough_days

    def resolve_period_from_question(question_text, explicit_date=None, default_days=None):
        """Return (start_date, end_date, label) inferred from text/date."""
        today_local = _local_today()
        if explicit_date:
            label = explicit_date.strftime('%d/%m/%Y')
            return explicit_date, explicit_date, label
        text = (question_text or '').strip()
        if not text:
            return None
        direction = 0
        lowered = text
        if any(word in lowered for word in ['prochain', 'suivant', 'à venir']):
            direction = 1
        elif any(word in lowered for word in ['dernier', 'dernière', 'precedent', 'précédent', 'passé', 'passe']):
            direction = -1
        year_match = re.search(r'(20[0-9]{2})', lowered)
        explicit_year = int(year_match.group(1)) if year_match else None
        detected_month = None
        for alias, idx in month_alias_map.items():
            if alias in lowered:
                detected_month = idx
                break
        if detected_month:
            target_year = explicit_year or today_local.year
            if explicit_year is None:
                if direction == -1 and detected_month >= today_local.month:
                    target_year -= 1
                elif direction == 1 and detected_month <= today_local.month:
                    target_year += 1
            start, end = _month_range(target_year, detected_month)
            label = f"{MONTH_LABELS[detected_month]} {target_year}"
            return start, end, label
        if any(kw in lowered for kw in ['mois prochain', 'mois-prochain']):
            month = today_local.month + 1
            year = today_local.year
            if month > 12:
                month = 1
                year += 1
            start, end = _month_range(year, month)
            return start, end, f"{MONTH_LABELS[month]} {year}"
        if any(kw in lowered for kw in ['ce mois', 'ce mois-ci', 'mois en cours']):
            start, end = _month_range(today_local.year, today_local.month)
            return start, end, f"{MONTH_LABELS[today_local.month]} {today_local.year}"
        if any(kw in lowered for kw in ['mois dernier', 'mois passé', 'mois precedent', 'mois précédent']):
            month = today_local.month - 1 or 12
            year = today_local.year - 1 if today_local.month == 1 else today_local.year
            start, end = _month_range(year, month)
            return start, end, f"{MONTH_LABELS[month]} {year}"
        if 'semaine prochaine' in lowered:
            start = today_local + timedelta(days=7 - today_local.weekday())
            end = start + timedelta(days=6)
            return start, end, "Semaine prochaine"
        if any(kw in lowered for kw in ['cette semaine', '7 derniers jours', 'semaine en cours']):
            start = today_local - timedelta(days=today_local.weekday())
            end = start + timedelta(days=6)
            return start, end, "Semaine en cours"
        if any(kw in lowered for kw in ['30 derniers jours', 'dernier mois', 'sur 30 jours']):
            start = today_local - timedelta(days=30)
            return start, today_local, '30 derniers jours'
        if default_days:
            start = today_local - timedelta(days=default_days)
            return start, today_local, f'{default_days} derniers jours'
        return None

    today = _local_today()

    def _with_answer(answer, insights=None, data=None):
        payload = {'answer': answer, 'matched': True}
        if insights:
            payload['insights'] = insights
        if data is not None:
            payload['data'] = data
        return JsonResponse(payload)

    if any(k in q_lower for k in ['congé', 'conge', 'conges', 'congés', 'vacances', 'permission']):
        list_keywords = ['qui', 'liste', 'lesquels', 'lesquelles', 'montre', 'affiche']
        wants_people = any(word in q_lower for word in list_keywords)
        if wants_people:
            period_hint = resolve_period_from_question(q_lower, date_obj)
            if date_obj and not period_hint:
                period_hint = (date_obj, date_obj, format_date(date_obj))
            if not period_hint and 'aujourd' in q_lower:
                period_hint = (today, today, "Aujourd'hui")
            if not period_hint and 'demain' in q_lower:
                target = today + timedelta(days=1)
                period_hint = (target, target, "Demain")
            if not period_hint:
                start_default, end_default = _month_range(today.year, today.month)
                period_hint = (start_default, end_default, f"{MONTH_LABELS[today.month]} {today.year}")
            start_date, end_date, label = period_hint
            leaves_qs = (
                Leave.objects.filter(status='APPROVED', start_date__lte=end_date, end_date__gte=start_date)
                .select_related('employee')
                .order_by('start_date')
            )
            leaves_sample = list(leaves_qs[:50])
            if not leaves_sample:
                return _with_answer(f"Aucun congé approuvé sur {label}.")
            total_leaves = leaves_qs.count()
            distinct_employees = leaves_qs.values('employee').distinct().count()
            data = [
                {
                    'employee': leave.employee.get_full_name() if hasattr(leave.employee, 'get_full_name') else str(leave.employee),
                    'start_date': leave.start_date.isoformat() if leave.start_date else None,
                    'end_date': leave.end_date.isoformat() if leave.end_date else None,
                    'type': leave.leave_type,
                    'status': leave.status,
                    'days': leave.days,
                }
                for leave in leaves_sample
            ]
            preview_names = []
            seen_names = set()
            for item in data:
                name = item['employee']
                if name and name not in seen_names:
                    preview_names.append(name)
                    seen_names.add(name)
                if len(preview_names) >= 4:
                    break
            answer = f"{format_int(distinct_employees)} collaborateur(s) ont un congé approuvé sur {label}."
            if preview_names:
                answer += " Exemples : " + ', '.join(preview_names) + '. '
            insights = [
                f"Période: {label}",
                f"Congés approuvés: {format_int(total_leaves)}",
                f"Employés concernés: {format_int(distinct_employees)}",
            ]
            return _with_answer(answer.strip(), insights, data)

        pending = Leave.objects.filter(status='PENDING').count()
        on_leave_today = Leave.objects.filter(status='APPROVED', start_date__lte=today, end_date__gte=today).count()
        upcoming = list(
            Leave.objects.filter(status='APPROVED', start_date__gt=today)
            .select_related('employee')
            .order_by('start_date')[:5]
        )
        upcoming_data = [
            {
                'employee': leave.employee.get_full_name() if hasattr(leave.employee, 'get_full_name') else str(leave.employee),
                'start_date': leave.start_date.isoformat() if leave.start_date else None,
                'end_date': leave.end_date.isoformat() if leave.end_date else None,
                'type': leave.leave_type,
                'status': leave.status,
            }
            for leave in upcoming
        ]
        insights = [
            f"Congés approuvés en cours: {format_int(on_leave_today)}",
            f"Demandes en attente: {format_int(pending)}",
            f"Prochains départs ({len(upcoming_data)} suivis): " + ', '.join(
                f"{item['employee']} ({format_date(datetime.fromisoformat(item['start_date']).date() if item['start_date'] else None)})"
                for item in upcoming_data
            ) if upcoming_data else 'Aucun congé programmé dans les prochains jours.',
        ]
        data = {
            'pending': pending,
            'on_leave_today': on_leave_today,
            'upcoming': upcoming_data,
        }
        return _with_answer("Synthèse congés générée à partir des données live.", insights, data)

    if any(k in q_lower for k in ['absence', 'absences', 'absent', 'retard collectif']):
        since = today - timedelta(days=30)
        unjustified = Absence.objects.filter(justified=False, date__gte=since).count()
        total = Absence.objects.filter(date__gte=since).count()
        top_absentees = list(
            Absence.objects.filter(date__gte=since)
            .values('employee__last_name', 'employee__first_name')
            .annotate(total=Count('id'))
            .order_by('-total')[:3]
        )
        insights = [
            f"Absences enregistrées (30 jours): {format_int(total)}",
            f"Dont injustifiées: {format_int(unjustified)}",
        ]
        if top_absentees:
            insights.append(
                "Top 3: " + ', '.join(
                    f"{item['employee__last_name']} {item['employee__first_name']} ({item['total']})" for item in top_absentees
                )
            )
        data = {
            'window_start': since.isoformat(),
            'window_end': today.isoformat(),
            'total_absences': total,
            'unjustified': unjustified,
            'top_absentees': [
                {
                    'name': f"{item['employee__last_name']} {item['employee__first_name']}",
                    'count': item['total'],
                }
                for item in top_absentees
            ],
        }
        return _with_answer("Absences consolidées sur les 30 derniers jours.", insights, data)

        if any(k in q_lower for k in ['paie', 'payroll', 'salaire', 'bulletin']):
            wants_remaining = any(kw in q_lower for kw in ['reste', 'restant', 'reste à payer', 'encore payer', 'à payer'])
            wants_month = any(kw in q_lower for kw in ['quel mois', 'quel mois', 'sur', 'pour', 'mois'])
            period_hint = resolve_period_from_question(q_lower, date_obj, default_days=30 if not wants_month else None)
            cohort = None
            label = None
            if period_hint and _is_full_month(period_hint):
                start_date, end_date, label = period_hint
                target_month = start_date.month
                target_year = start_date.year
                cohort = Payroll.objects.filter(year=target_year, month=target_month)
            else:
                last_pay = Payroll.objects.order_by('-year', '-month').first()
                if last_pay:
                    label = f"{MONTH_LABELS[last_pay.month]} {last_pay.year}"
                    cohort = Payroll.objects.filter(year=last_pay.year, month=last_pay.month)
            if not cohort:
                return _with_answer("Aucune paie enregistrée pour le moment.")
            totals = cohort.aggregate(total_gross=Sum('gross_salary'), total_net=Sum('net_salary'))
            paid_employees = cohort.count()
            gross_val = float(totals['total_gross'] or 0)
            net_val = float(totals['total_net'] or 0)
            answer = f"Paie {label or 'la plus récente'} : {format_int(paid_employees)} salarié(s) payés. Total net { _format_currency(net_val)} Ar."
            insights = [
                f"Période: {label}",
                f"Total brut: {_format_currency(gross_val)} Ar",
                f"Total net: {_format_currency(net_val)} Ar",
            ]
            if wants_remaining:
                active_emps = base_employee_qs.count()
                remaining = max(active_emps - paid_employees, 0)
                insights.append(f"Salariés restants à payer (base active): {format_int(remaining)}")
            data = {
                'period': label,
                'paid_employees': paid_employees,
                'total_gross': gross_val,
                'total_net': net_val,
            }
            if wants_remaining:
                data['remaining_to_pay'] = remaining
            return _with_answer("Synthèse de la campagne de paie.", insights, data)

        if any(k in q_lower for k in ['contrat', 'cdi', 'cdd', 'essai']):
            active_contracts = Contract.objects.filter(active=True, employee__archived=False)
            breakdown_qs = active_contracts.values('type').annotate(total=Count('id')).order_by('-total')
            breakdown = [
                {
                    'code': item['type'],
                    'label': contract_label_map.get(item['type'], item['type']),
                    'total': item['total'],
                }
                for item in breakdown_qs
            ]
            upcoming_trials = list(
                active_contracts.filter(type='ESSAI', date_end__isnull=False, date_end__lte=today + timedelta(days=30))
                .select_related('employee')
                .order_by('date_end')[:5]
            )
            insights = [
                f"Contrats actifs: {format_int(active_contracts.count())}",
                "Répartition: " + ', '.join(f"{item['label']} ({item['total']})" for item in breakdown[:3]) if breakdown else 'Pas de données contrat.',
            ]
            if upcoming_trials:
                insights.append(
                    "Essais à surveiller: " + ', '.join(
                        f"{c.employee.last_name} {c.employee.first_name} ({format_date(c.date_end)})" for c in upcoming_trials
                    )
                )
            data = {
                'breakdown': breakdown,
                'trial_periods_ending': [
                    {
                        'employee': f"{c.employee.last_name} {c.employee.first_name}",
                        'date_end': c.date_end.isoformat() if c.date_end else None,
                    }
                    for c in upcoming_trials
                ],
            }
            return _with_answer("Contrats RH consolidés.", insights, data)

        effectif_keywords = ['effectif', 'effectifs', "nombre d'employ", 'combien d\'employ', 'headcount', 'employés']
        if any(k in q_lower for k in effectif_keywords):
            total = base_employee_qs.count()
            active = base_employee_qs.filter(is_active=True).count()
            inactive = total - active
            archived = Employee.objects.filter(archived=True).count()
            gender_map = {'M': 'Hommes', 'F': 'Femmes', 'O': 'Autres', None: 'Non renseigné'}
            gender_breakdown = [
                {
                    'code': item['gender'],
                    'label': gender_map.get(item['gender'], gender_map[None]),
                    'total': item['total'],
                }
                for item in base_employee_qs.values('gender').annotate(total=Count('id'))
            ]
            contract_breakdown = [
                {
                    'code': item['type'],
                    'label': contract_label_map.get(item['type'], item['type']),
                    'total': item['total'],
                }
                for item in Contract.objects.filter(active=True, employee__archived=False).values('type').annotate(total=Count('id')).order_by('-total')
            ]
            recent_hires = [
                {
                    'name': f"{emp.last_name} {emp.first_name}",
                    'hire_date': emp.hire_date.isoformat() if emp.hire_date else None,
                    'department': emp.department,
                }
                for emp in base_employee_qs.filter(hire_date__isnull=False).order_by('-hire_date')[:3]
            ]
            insights = [
                f"Effectif total (hors archivés): {format_int(total)}",
                f"Actifs: {format_int(active)} — Inactifs: {format_int(inactive)}",
            ]
            if archived:
                insights.append(f"Dossiers archivés: {format_int(archived)}")
            if contract_breakdown:
                top_contracts = ', '.join(f"{item['label']}: {item['total']}" for item in contract_breakdown[:3])
                insights.append(f"Principaux contrats: {top_contracts}")
            data = {
                'total': total,
                'active': active,
                'inactive': inactive,
                'archived': archived,
                'gender_breakdown': gender_breakdown,
                'contract_breakdown': contract_breakdown,
                'recent_hires': recent_hires,
            }
            return _with_answer("Effectif consolidé depuis la base RH.", insights, data)

        if 'departement' in q_lower or 'département' in q_lower or 'service' in q_lower:
            top_departments = [
                {
                    'name': item['department'] or 'Non renseigné',
                    'total': item['total'],
                }
                for item in base_employee_qs.values('department').annotate(total=Count('id')).order_by('-total')[:5]
            ]
            top_services = [
                {
                    'name': item['service'] or 'Non renseigné',
                    'total': item['total'],
                }
                for item in base_employee_qs.values('service').annotate(total=Count('id')).order_by('-total')[:5]
            ]
            insights = []
            if top_departments:
                insights.append("Top départements: " + ', '.join(f"{d['name']} ({d['total']})" for d in top_departments[:3]))
            if top_services:
                insights.append("Top services: " + ', '.join(f"{s['name']} ({s['total']})" for s in top_services[:3]))
            data = {
                'departments': top_departments,
                'services': top_services,
            }
            return _with_answer("Répartition par département/service.", insights or None, data)

        if any(k in q_lower for k in ['alerte', 'incident', 'risque', 'anomalie']):
            open_alerts = Alerte.objects.filter(statut='OPEN').order_by('date_creation')
            insights = [f"Alertes ouvertes: {format_int(open_alerts.count())}"]
            upcoming = [
                {
                    'message': alert.message,
                    'created_at': alert.date_creation.isoformat() if alert.date_creation else None,
                    'priority': getattr(alert, 'priority', None),
                }
                for alert in open_alerts[:5]
            ]
            if upcoming:
                insights.append("Principales alertes: " + '; '.join(item['message'] for item in upcoming))
            data = {'alerts': upcoming}
            return _with_answer("Alertes RH en cours.", insights, data)

        if any(k in q_lower for k in ['formation', 'compétence', 'competence', 'training']):
            total_suggestions = TrainingSuggestion.objects.count()
            recent = [
                {
                    'title': sugg.title,
                    'employee': str(sugg.employee) if sugg.employee else None,
                    'created_at': sugg.created_at.isoformat() if sugg.created_at else None,
                }
                for sugg in TrainingSuggestion.objects.order_by('-created_at')[:5]
            ]
            insights = [
                f"Suggestions de formation enregistrées: {format_int(total_suggestions)}",
                f"Dernières idées: {', '.join(item['title'] for item in recent)}" if recent else 'Aucune suggestion récente.',
            ]
            data = {
                'total_suggestions': total_suggestions,
                'recent': recent,
            }
            return _with_answer("Formations & compétences (source base RH).", insights, data)

    # exact match first
    for item in faq:
        q = (item.get('question') or '').strip().lower()
        if q and q == q_lower:
            ans = item.get('answer', '')
            if employee_obj:
                # personalize short FAQ answers when possible
                ans = f"Pour {employee_obj.last_name} {employee_obj.first_name} : {ans}"
            return JsonResponse({'answer': ans, 'matched': True})

    # keyword matching fallback
    for item in faq:
        keywords = item.get('keywords', [])
        for kw in keywords:
            if kw and kw in q_lower:
                return JsonResponse({'answer': item.get('answer', ''), 'matched': True})

    # last resort: substring match on question text
    for item in faq:
        if item.get('question') and item.get('question').lower() in q_lower:
            ans = item.get('answer', '')
            if employee_obj:
                ans = f"Pour {employee_obj.last_name} {employee_obj.first_name} : {ans}"
            return JsonResponse({'answer': ans, 'matched': True})

    # If no FAQ matched, try a few simple employee-aware heuristics
    # If multiple employees requested, handle aggregated queries (leave balances etc.)
    if employees_list:
        # keywords for leave balance
        if any(k in q_lower for k in ['solde', 'solde des congés', 'congé', 'conge', 'congés', 'conges']):
            lines = []
            as_of = date_obj or datetime.today().date()
            # limit to avoid huge responses
            limit = 200
            for e in employees_list[:limit]:
                try:
                    avail = employee_available_leave(e, as_of_date=as_of)
                    lines.append(f"{e.last_name} {e.first_name}: {avail:.2f} jour(s)")
                except Exception:
                    lines.append(f"{e.last_name} {e.first_name}: solde inconnu")
            if len(employees_list) > limit:
                lines.append(f"...et {len(employees_list)-limit} autres employés")
            return JsonResponse({'answer': ' / '.join(lines), 'matched': True})

        # keywords for absences summary for multiple employees
        if any(k in q_lower for k in ['absence', 'absences', 'absent']):
            now = datetime.today().date()
            one_year_ago = now - timedelta(days=365)
            limit = 200
            data = []
            lines = []
            for e in employees_list[:limit]:
                try:
                    abs_count = Absence.objects.filter(employee=e, date__gte=one_year_ago).count()
                    recent = Absence.objects.filter(employee=e).order_by('-date')[:3]
                    recent_list = [a.date.isoformat() for a in recent]
                    data.append({
                        'employee_id': e.pk,
                        'name': f"{e.last_name} {e.first_name}",
                        'matricule': e.matricule,
                        'absences_last_12_months': int(abs_count),
                        'recent_absence_dates': recent_list,
                    })
                    lines.append(f"{e.last_name} {e.first_name}: {abs_count} absences (dernieres: {', '.join(recent_list)})")
                except Exception:
                    lines.append(f"{e.last_name} {e.first_name}: données absence indisponibles")
            if len(employees_list) > limit:
                lines.append(f"...et {len(employees_list)-limit} autres employés")
            return JsonResponse({'answer': ' / '.join(lines), 'matched': True, 'data': data})

        # full profile / all available data for multiple employees
        if any(k in q_lower for k in ['tous', 'toutes', 'toutes les données', 'toutes les informations', 'informations', 'profil']):
            now = datetime.today().date()
            one_year_ago = now - timedelta(days=365)
            limit = 100
            data = []
            for e in employees_list[:limit]:
                try:
                    leave_balance = employee_available_leave(e, as_of_date=now)
                except Exception:
                    leave_balance = None
                last_pay = Payroll.objects.filter(employee=e).order_by('-year', '-month').first()
                last_pay_obj = None
                if last_pay:
                    last_pay_obj = {
                        'year': last_pay.year,
                        'month': last_pay.month,
                        'gross': float(last_pay.gross_salary or 0),
                        'net': float(last_pay.net_salary or 0),
                    }
                upcoming = list(Leave.objects.filter(employee=e, start_date__gte=now).order_by('start_date')[:5].values('start_date','end_date','leave_type','status'))
                abs_count = Absence.objects.filter(employee=e, date__gte=one_year_ago).count()
                lates_qs = Presence.objects.filter(employee=e, date__gte=(now - timedelta(days=30)), minutes_late__gt=0)
                lates_count = lates_qs.count()
                lates_minutes = lates_qs.aggregate(total=Sum('minutes_late')).get('total') or 0
                contracts = list(e.contracts.order_by('-date_start').values('type','date_start','date_end','active')[:3])
                data.append({
                    'employee_id': e.pk,
                    'matricule': e.matricule,
                    'name': f"{e.last_name} {e.first_name}",
                    'email': e.email,
                    'hire_date': e.hire_date.isoformat() if e.hire_date else None,
                    'is_active': bool(e.is_active),
                    'department': e.department,
                    'function': e.function,
                    'leave_balance': float(leave_balance) if leave_balance is not None else None,
                    'upcoming_leaves': upcoming,
                    'last_payroll': last_pay_obj,
                    'absences_last_12_months': int(abs_count),
                    'lates_30_days': int(lates_count),
                    'lates_minutes_30_days': int(lates_minutes),
                    'recent_contracts': contracts,
                })
            summary = f"Profil généré pour {len(data)} employé(s)."
            return JsonResponse({'answer': summary, 'matched': True, 'data': data})

    if employee_obj:
        # If the user asks about leave balance for the selected employee, return a
        # clear, user-facing message that points them to their employee fiche.
        if any(k in q_lower for k in ['solde', 'solde des congés', 'congé', 'conge', 'congés', 'conges']):
                try:
                    name = f"{employee_obj.first_name} {employee_obj.last_name}".strip()
                except Exception:
                    name = getattr(employee_obj, 'last_name', '') or getattr(employee_obj, 'first_name', '') or 'Employé'
                # Try to compute available leave dynamically, fall back to polite guidance
                as_of = date_obj or datetime.today().date()
                avail = None
                try:
                    avail = employee_available_leave(employee_obj, as_of_date=as_of)
                except Exception:
                    avail = None
                url = ''
                try:
                    url = reverse('export_employee_fiche', args=[employee_obj.pk])
                except Exception:
                    url = ''

                if avail is not None:
                    try:
                        # format numeric balances nicely
                        if isinstance(avail, Decimal):
                            avail_val = float(avail)
                        else:
                            avail_val = float(avail)
                        balance_text = f"{avail_val:.2f} jour(s)"
                    except Exception:
                        balance_text = str(avail)
                    parts = [f"Pour {name} : Solde des congés : {balance_text}."]
                    if url:
                        parts.append(f"Télécharger la fiche: {url}")
                    return JsonResponse({'answer': ' '.join(parts), 'matched': True, 'data': {'leave_balance': avail_val if isinstance(avail, (int, float, Decimal)) else None}})
                # Fallback message when balance unavailable
                if url:
                    return JsonResponse({'answer': f"Pour {name} : Le solde des congés n'est pas disponible immédiatement. Vous pouvez consulter votre fiche employé: {url} ou contacter l'administrateur RH.", 'matched': True})
                return JsonResponse({'answer': f"Pour {name} : Le solde des congés est visible dans votre fiche employé sous 'Congés'. Si vous ne le voyez pas, contactez l'administrateur RH.", 'matched': True})
        # employment status / still employed?
        if any(k in q_lower for k in ['encore', 'toujours', 'est-il', 'est elle', 'est-ce', 'travaille', 'employé', 'présent', 'present']):
            try:
                active_contract = Contract.objects.filter(employee=employee_obj, active=True).exists()
            except Exception:
                active_contract = False
            if active_contract:
                return JsonResponse({'answer': f"Oui — {employee_obj.first_name} {employee_obj.last_name} a un contrat actif.", 'matched': True})
            # fallback: check recent payroll
            last_pay = Payroll.objects.filter(employee=employee_obj).order_by('-year', '-month').first()
            if last_pay:
                return JsonResponse({'answer': f"La dernière paie enregistrée pour {employee_obj.first_name} {employee_obj.last_name} est {last_pay.month}/{last_pay.year}.", 'matched': True})
            return JsonResponse({'answer': f"Aucun contrat actif ou paie récente trouvé pour {employee_obj.first_name} {employee_obj.last_name}. Il se peut qu'il/elle ne soit plus employé(e).", 'matched': False})

        # lateness / retard
        if 'retard' in q_lower or 'en retard' in q_lower or 'tard' in q_lower:
            if date_obj:
                pres = Presence.objects.filter(employee=employee_obj, date=date_obj).first()
                if pres and getattr(pres, 'minutes_late', 0) and pres.minutes_late > 0:
                    return JsonResponse({'answer': f"Le {date_obj.isoformat()}, {employee_obj.first_name} {employee_obj.last_name} a été en retard de {int(pres.minutes_late)} minute(s).", 'matched': True})
                return JsonResponse({'answer': f"Aucun retard enregistré pour {employee_obj.first_name} {employee_obj.last_name} le {date_obj.isoformat()}.", 'matched': True})
            else:
                # summarize last 30 days
                since = (datetime.today() - timedelta(days=30)).date()
                late_qs = Presence.objects.filter(employee=employee_obj, date__gte=since, minutes_late__gt=0)
                count = late_qs.count()
                total = late_qs.aggregate(total=Sum('minutes_late')).get('total') or 0
                return JsonResponse({'answer': f"Dans les 30 derniers jours: {count} jour(s) avec retard, total {int(total)} minute(s).", 'matched': True})

        # leave on a given date
        if 'congé' in q_lower or 'conge' in q_lower or 'vacances' in q_lower or 'permission' in q_lower:
            if date_obj:
                leaves = Leave.objects.filter(employee=employee_obj, start_date__lte=date_obj, end_date__gte=date_obj)
                if leaves.exists():
                    items = []
                    for l in leaves:
                        items.append(f"{l.leave_type} ({l.status}) du {l.start_date} au {l.end_date}")
                    return JsonResponse({'answer': f"Oui — {employee_obj.first_name} {employee_obj.last_name} est en congé: {'; '.join(items)}", 'matched': True})
                return JsonResponse({'answer': f"Non — {employee_obj.first_name} {employee_obj.last_name} n'est pas en congé le {date_obj.isoformat()}.", 'matched': True})
            else:
                # upcoming leaves
                today_d = datetime.today().date()
                upcoming = Leave.objects.filter(employee=employee_obj, start_date__gte=today_d).order_by('start_date')[:5]
                if upcoming.exists():
                    items = [f"{l.leave_type}: {l.start_date} → {l.end_date} ({l.status})" for l in upcoming]
                    return JsonResponse({'answer': f"Congés à venir: {'; '.join(items)}", 'matched': True})
                return JsonResponse({'answer': "Aucun congé à venir trouvé pour cet employé.", 'matched': True})
        # salary related
        if 'salaire' in q_lower or 'salaire de base' in q_lower or 'net' in q_lower:
            p = Payroll.objects.filter(employee=employee_obj).order_by('-year', '-month').first()
            if p:
                gross = float(p.gross_salary or 0)
                net = float(p.net_salary or 0)
                base_salary = float(p.salary_base or employee_obj.salary_base or 0)
                url = ''
                try:
                    url = reverse('export_employee_fiche', args=[employee_obj.pk])
                except Exception:
                    url = ''
                parts = [f"Salaire brut: {int(round(gross))}", f"Net: {int(round(net))}", f"Salaire base: {int(round(base_salary))}"]
                if url:
                    parts.append(f"Télécharger la fiche: {url}")
                return JsonResponse({'answer': ' — '.join(parts), 'matched': True})
            else:
                return JsonResponse({'answer': "Aucune paie trouvée pour cet employé. Vous pouvez générer la fiche de paie via l'interface employé.", 'matched': False})

        # fiche de paie / export
        if 'fiche' in q_lower or 'fiche de paie' in q_lower:
            try:
                url = reverse('export_employee_fiche', args=[employee_obj.pk])
                return JsonResponse({'answer': f"Vous pouvez télécharger la fiche de paie: {url}", 'matched': True})
            except Exception:
                return JsonResponse({'answer': "La fiche n'est pas disponible pour le moment.", 'matched': False})

    # default fallback
    fallback = "Désolé, je n'ai pas trouvé de réponse exacte. Voici quelques sujets possibles: " + ", ".join([i.get('question') for i in faq[:6]])
    return JsonResponse({'answer': fallback, 'matched': False})

def export_leaves_xlsx(request):
    """Export leaves to an Excel file generated on the fly."""
    headers = ['Employee (matricule)', 'Employee Name', 'Type', 'Start Date', 'End Date', 'Days', 'Status', 'Note']
    non_empty = request.GET.get('non_empty') in ('1', 'true', 'True')
    data = []
    for l in Leave.objects.select_related('employee').all().order_by('-start_date'):
        row = {
            'Employee (matricule)': l.employee.matricule if l.employee else '',
            'Employee Name': f"{l.employee.last_name} {l.employee.first_name}" if l.employee else '',
            'Type': l.leave_type,
            'Start Date': l.start_date.isoformat() if l.start_date else '',
            'End Date': l.end_date.isoformat() if l.end_date else '',
            'Days': l.days,
            'Status': l.status,
            'Note': l.note or '',
        }
        if non_empty:
            if all((v is None or (isinstance(v, str) and v.strip() == '') or v == 0) for v in row.values()):
                continue
        data.append(row)

    if pd is not None:
        df = pd.DataFrame(data, columns=headers)
        bio = BytesIO()
        with pd.ExcelWriter(bio, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Leaves')
        bio.seek(0)
        resp = HttpResponse(bio.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="leaves.xlsx"'
        return resp

    # fallback when pandas not available: build with openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Leaves'
    ws.append(headers)
    for r in data:
        ws.append([r[h] for h in headers])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    resp = HttpResponse(bio.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="leaves.xlsx"'
    return resp


@login_required
def planner_calendar(request):
    """Render a minimal calendar UI that loads approved leaves from the API.

    Uses FullCalendar (via CDN) to show leave events fetched from /api/leaves/.
    """
    return render(request, 'core/planner_calendar.html', {})


@login_required
def suggest_replacement_page(request):
    """Render a simple suggestion form that lets the user create a ReplacementRequest
    and query the suggestion API, then save suggestions.
    """
    return render(request, 'core/suggest_replacement.html', {})


@login_required
def stats_charts(request):
    """Render a basic charts page that fetches stats endpoints and draws charts (Chart.js)."""
    return render(request, 'core/stats_charts.html', {})


@login_required
def messages_page(request):
    """Simple messaging UI that uses the Message API."""
    return render(request, 'core/messages.html', {})


@login_required
def performance_page(request):
    """List and create PerformanceReview via API."""
    return render(request, 'core/performance.html', {})


@login_required
def training_page(request):
    """List and create TrainingSuggestion via API."""
    # Render the training suggestions UI (client calls the API endpoint)
    return render(request, 'core/training_suggestions.html', {})


@login_required
def competency_cartography_page(request):
    """Render competency cartography UI."""
    return render(request, 'core/competency_cartography.html', {})


@login_required
def match_candidates_page(request):
    """Render simple candidate matching UI."""
    return render(request, 'core/match_candidates.html', {})


@login_required
def reports_page(request):
    """Render a simple reports listing page showing generated reports (xlsx/pdf).

    Access restricted to staff or HR/Manager groups.
    """
    user = request.user
    allowed = user.is_staff
    try:
        allowed = allowed or user.groups.filter(name__in=['HR', 'Manager']).exists()
    except Exception:
        allowed = allowed

    if not allowed:
        return HttpResponse('Accès refusé', status=403)

    reports = Report.objects.order_by('-created_at')[:200]
    return render(request, 'core/reports.html', {'reports': reports})


@login_required
def report_download(request, pk):
    """Serve the XLSX (or PDF) file attached to a Report record. Only staff/HR/Manager can download."""
    user = request.user
    allowed = user.is_staff
    try:
        allowed = allowed or user.groups.filter(name__in=['HR', 'Manager']).exists()
    except Exception:
        allowed = allowed
    if not allowed:
        return HttpResponse('Accès refusé', status=403)

    try:
        rep = Report.objects.get(pk=pk)
    except Report.DoesNotExist:
        return HttpResponse('Rapport introuvable', status=404)

    # Helper to resolve stored file path and transparently fix legacy rows
    def resolve_report_path(path_str, update_field):
        if not path_str:
            return None
        candidate = Path(path_str)
        if candidate.exists():
            return candidate

        # fallbacks: look for the same filename in known exports locations
        search_roots = [
            Path(settings.BASE_DIR) / 'exports',
            Path(settings.BASE_DIR) / 'exports' / 'reports',
            Path(settings.BASE_DIR).parent / 'exports',
            Path(settings.BASE_DIR).parent / 'exports' / 'reports',
        ]
        for root in search_roots:
            alt = root / candidate.name
            if alt.exists():
                try:
                    setattr(rep, update_field, str(alt))
                    rep.save(update_fields=[update_field])
                except Exception:
                    pass
                return alt
        return None

    file_path = resolve_report_path(rep.xlsx_path, 'xlsx_path') or resolve_report_path(rep.pdf_path, 'pdf_path')

    if not file_path:
        return HttpResponse('Fichier introuvable sur le serveur', status=404)

    # Choose content type based on suffix
    suffix = file_path.suffix.lower()
    if suffix in ['.xlsx', '.xls']:
        ct = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    elif suffix == '.csv':
        ct = 'text/csv'
    elif suffix == '.pdf':
        ct = 'application/pdf'
    else:
        ct = 'application/octet-stream'
    try:
        file_handle = open(file_path, 'rb')
        response = FileResponse(file_handle, content_type=ct)
        response['Content-Disposition'] = f'attachment; filename="{file_path.name}"'
        return response
    except Exception as exc:
        logger = logging.getLogger(__name__)
        logger.exception('Failed to serve report %s: %s', pk, exc)
        return HttpResponse('Erreur lors de la lecture du fichier', status=500)


class ReportsListAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def _is_hr_user(self, user):
        is_hr = getattr(user, 'is_staff', False)
        try:
            is_hr = is_hr or user.groups.filter(name__in=['HR', 'Manager']).exists()
        except Exception:
            pass
        return is_hr

    def get(self, request, *args, **kwargs):
        if not self._is_hr_user(request.user):
            return Response({'error': 'forbidden'}, status=drf_status.HTTP_403_FORBIDDEN)

        qs = Report.objects.order_by('-created_at')[:200]
        data = []
        for r in qs:
            data.append({
                'id': r.id,
                'name': r.name,
                'xlsx_path': r.xlsx_path,
                'pdf_path': r.pdf_path,
                'created_at': r.created_at.isoformat(),
                'created_by': getattr(r.created_by, 'username', None),
                'notes': r.notes,
            })
        return Response({'reports': data})


class PerformanceRunAPIView(APIView):
    """API endpoint to trigger compute_performance_scores management command on demand.

    Permissions: only staff users (HR/managers) can run this.
    This runs the existing management command synchronously and returns a simple status.
    For production scheduling, run the management command via cron/Celery/beat.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        user = request.user
        allowed = getattr(user, 'is_staff', False)
        try:
            allowed = allowed or user.groups.filter(name__in=['HR', 'Manager']).exists()
        except Exception:
            pass
        if not allowed:
            return Response({'detail': 'Permission denied'}, status=drf_status.HTTP_403_FORBIDDEN)

        try:
            # Synchronously call management command
            call_command('compute_performance_scores')
            return Response({'detail': 'compute_performance_scores executed. Check reports list.'})
        except Exception as e:
            return Response({'detail': 'Error running compute_performance_scores', 'error': str(e)}, status=drf_status.HTTP_500_INTERNAL_SERVER_ERROR)


@login_required
def self_service_profile(request):
    """Allow a logged-in user to edit their own Employee profile.

    Resolution strategy for mapping User -> Employee:
    1. Try to find Employee with email == request.user.email
    2. Fallback: try matricule == request.user.username
    3. If none found, render a help message instructing to contact HR/admin.
    """
    user = getattr(request, 'user', None)
    emp = None
    if user and user.is_authenticated:
        # try match by email first
        try:
            if getattr(user, 'email', None):
                emp = Employee.objects.filter(email__iexact=user.email).first()
        except Exception:
            emp = None
        # fallback to matricule == username
        if emp is None:
            try:
                emp = Employee.objects.filter(matricule__iexact=str(user.username)).first()
            except Exception:
                emp = None

    # If still not found, show a friendly page explaining how to link accounts
    if emp is None:
        return render(request, 'core/self_service_profile.html', {
            'not_linked': True,
            'user': user,
        })

    from .forms import EmployeeSelfServiceForm

    updated = False
    if request.method == 'POST':
        form = EmployeeSelfServiceForm(request.POST, request.FILES, instance=emp)
        if form.is_valid():
            form.save()
            return redirect(f"{reverse('self_service_profile')}?saved=1")
    else:
        updated = request.GET.get('saved') == '1'
        form = EmployeeSelfServiceForm(instance=emp)

    return render(request, 'core/self_service_profile.html', {
        'form': form,
        'employee': emp,
        'updated': updated,
    })


class EmployeeSelfAPIView(APIView):
    """API endpoint for the authenticated user to view and update their own Employee record."""
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def _resolve_employee(self, user):
        emp = None
        try:
            if getattr(user, 'email', None):
                emp = Employee.objects.filter(email__iexact=user.email).first()
        except Exception:
            emp = None
        if emp is None:
            try:
                emp = Employee.objects.filter(matricule__iexact=str(user.username)).first()
            except Exception:
                emp = None
        return emp

    def get(self, request, *args, **kwargs):
        user = request.user
        emp = self._resolve_employee(user)
        if emp is None:
            return Response({'error': 'Employee profile not linked to your account.'}, status=drf_status.HTTP_404_NOT_FOUND)
        serializer = EmployeeSelfSerializer(emp, context={'request': request})
        return Response(serializer.data)

    def patch(self, request, *args, **kwargs):
        user = request.user
        emp = self._resolve_employee(user)
        if emp is None:
            return Response({'error': 'Employee profile not linked to your account.'}, status=drf_status.HTTP_404_NOT_FOUND)
        serializer = EmployeeSelfSerializer(emp, data=request.data, partial=True, context={'request': request})
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response({'error': serializer.errors}, status=drf_status.HTTP_400_BAD_REQUEST)


class MessagesInboxAPIView(APIView):
    """List messages for the authenticated user (employees) or all when HR."""
    permission_classes = [IsAuthenticated]

    def _resolve_employee(self, user):
        try:
            if getattr(user, 'email', None):
                emp = Employee.objects.filter(email__iexact=user.email).first()
                if emp:
                    return emp
        except Exception:
            pass
        try:
            return Employee.objects.filter(matricule__iexact=str(user.username)).first()
        except Exception:
            return None

    def _is_hr_user(self, user):
        is_hr = getattr(user, 'is_staff', False)
        try:
            is_hr = is_hr or user.groups.filter(name__in=['HR', 'Manager']).exists()
        except Exception:
            pass
        return is_hr

    def get(self, request, *args, **kwargs):
        emp = self._resolve_employee(request.user)
        is_hr = self._is_hr_user(request.user)
        if emp is None and not is_hr:
            return Response({'error': 'Aucun dossier employé n’est lié à votre compte. Merci de contacter le service RH pour le rattacher.'}, status=drf_status.HTTP_404_NOT_FOUND)

        if is_hr:
            queryset = Message.objects.select_related('recipient').order_by('-created_at')[:200]
            mode = 'hr'
        else:
            queryset = Message.objects.filter(recipient=emp).order_by('-created_at')[:200]
            mode = 'employee'

        msgs = list(queryset)
        unread_ids = [m.pk for m in msgs if not m.read]
        if unread_ids and mode == 'employee':
            Message.objects.filter(pk__in=unread_ids).update(read=True)

        data = []
        for m in msgs:
            data.append({
                'id': m.pk,
                'subject': m.subject,
                'body': m.body,
                'read': m.read or (mode == 'employee' and m.pk in unread_ids),
                'created_at': m.created_at.isoformat(),
                'sender': getattr(m.sender, 'username', None),
                'recipient': str(m.recipient) if mode == 'hr' and m.recipient else None,
            })
        return Response({'messages': data, 'mode': mode})


class MessageRecipientsAPIView(APIView):
    """Return a list of employees that the current user can contact."""
    permission_classes = [IsAuthenticated]

    def get_queryset(self, user):
        qs = Employee.objects.filter(is_active=True, archived=False)
        is_hr = getattr(user, 'is_staff', False)
        try:
            is_hr = is_hr or user.groups.filter(name__in=['HR', 'Manager']).exists()
        except Exception:
            pass
        if not is_hr:
            qs = qs.filter(
                Q(department__icontains='RH') |
                Q(function__icontains='RH') |
                Q(service__icontains='RH')
            )
        return qs.order_by('last_name', 'first_name')[:200], is_hr

    def get(self, request, *args, **kwargs):
        try:
            recips, is_hr = self.get_queryset(request.user)
        except Exception as exc:
            logger = logging.getLogger(__name__)
            logger.exception('Failed to load message recipients for %s: %s', request.user, exc)
            return Response({'recipients': [], 'help_text': "Impossible de charger la liste des destinataires.", 'hr_view': False})

        data = []
        for e in recips:
            data.append({
                'id': e.pk,
                'name': f"{e.last_name} {e.first_name}".strip(),
                'department': e.department,
                'function': e.function,
            })

        help_text = "Sélectionnez un collaborateur pour lui adresser un message RH." if is_hr else "Seuls les contacts RH apparaissent dans cette liste."
        if not data:
            help_text = "Aucun collaborateur RH actif n'est encore paramétré dans le système."

        return Response({'recipients': data, 'help_text': help_text, 'hr_view': is_hr})


class MessageSendAPIView(APIView):
    """Allow authenticated users to send a message to an employee (HR or employees)."""
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        data = request.data or {}
        recipient_id = data.get('recipient') or data.get('recipient_id')
        subject = (data.get('subject') or '').strip()
        body = (data.get('body') or '').strip()
        sender = request.user

        if not recipient_id:
            return Response({'error': 'recipient is required (employee id)'}, status=drf_status.HTTP_400_BAD_REQUEST)
        if not subject:
            return Response({'error': 'Le sujet est obligatoire.'}, status=drf_status.HTTP_400_BAD_REQUEST)
        if not body:
            return Response({'error': 'Le message est obligatoire.'}, status=drf_status.HTTP_400_BAD_REQUEST)

        def allowed_recipients(user):
            qs = Employee.objects.filter(is_active=True, archived=False)
            is_hr = getattr(user, 'is_staff', False)
            try:
                is_hr = is_hr or user.groups.filter(name__in=['HR', 'Manager']).exists()
            except Exception:
                pass
            if not is_hr:
                qs = qs.filter(
                    Q(department__icontains='RH') |
                    Q(function__icontains='RH') |
                    Q(service__icontains='RH')
                )
            return qs

        try:
            recipient = allowed_recipients(sender).get(pk=int(recipient_id))
        except (Employee.DoesNotExist, ValueError, TypeError):
            return Response({'error': 'recipient not found or not allowed'}, status=drf_status.HTTP_404_NOT_FOUND)

        message = Message.objects.create(sender=sender, recipient=recipient, subject=subject, body=body)
        return Response({'result': 'ok', 'id': message.pk, 'subject': message.subject})


def _create_fiche_xlsx_bytes(employee, payroll):
    """Create XLSX bytes for a given employee+payroll using the same template logic
    as `export_employee_fiche`. Returns bytes of the workbook.
    """
    base = Path(settings.BASE_DIR)
    # prefer the same search logic as export_employee_fiche
    possible_templates = [
        base.parent / 'FICHE DE PAIE .xlsx',
        base.parent / 'exports' / 'FICHE_DE_PAIE.xlsx',
        base.parent / 'exports' / 'FICHE DE PAIE .xlsx',
        base.parent / 'exports' / 'FICHE_DE_PAIE.xls',
    ]
    template_path = None
    for candidate in possible_templates:
        if candidate.exists():
            template_path = candidate
            break

    def make_mapping_local(e, p):
        return {
            'MATRICULE': e.matricule or '',
            'LAST_NAME': e.last_name or '',
            'FIRST_NAME': e.first_name or '',
            'CATEGORY': str(e.category) if e.category else '',
            'FUNCTION': e.function or '',
            'HIRE_DATE': e.hire_date.isoformat() if e.hire_date else '',
            'SALARY_BASE': float(e.salary_base or 0),
            'CNAPS': e.cnaps_number or '',
            'PAYROLL_YEAR': p.year if p else '',
            'PAYROLL_MONTH': p.month if p else '',
            'GROSS': float(p.gross_salary or 0) if p else '',
            'NET': float(p.net_salary or 0) if p else '',
        }

    mapping = make_mapping_local(employee, payroll)
    bio = BytesIO()
    if template_path.exists():
        try:
            wb = openpyxl.load_workbook(filename=str(template_path))
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str) and cell.value:
                            text = cell.value
                            for key, val in mapping.items():
                                text = text.replace(f'{{{{{key}}}}}', str(val))
                                text = text.replace(f'{{{key}}}', str(val))
                                text = text.replace(key, str(val))
                            cell.value = text
            wb.save(bio)
            bio.seek(0)
            return bio.getvalue()
        except Exception:
            # will fallback to simple sheet generation below
            pass

    # fallback simple generation
    rows = [
        ('Matricule', employee.matricule or ''),
        ('Last name', employee.last_name or ''),
        ('First name', employee.first_name or ''),
        ('Category', str(employee.category) if employee.category else ''),
        ('Function', employee.function or ''),
        ('Hire date', employee.hire_date.isoformat() if employee.hire_date else ''),
        ('Salary base', float(employee.salary_base or 0)),
    ]
    if payroll:
        rows += [
            ('Payroll year', payroll.year),
            ('Payroll month', payroll.month),
            ('Gross', float(payroll.gross_salary or 0)),
            ('Net', float(payroll.net_salary or 0)),
        ]

    if pd is not None:
        df = pd.DataFrame(rows, columns=['Field', 'Value'])
        with pd.ExcelWriter(bio, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Fiche')
        bio.seek(0)
        return bio.getvalue()
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Fiche'
        ws.append(['Field', 'Value'])
        for r in rows:
            ws.append([r[0], r[1]])
        wb.save(bio)
        bio.seek(0)
        return bio.getvalue()


def _presence_summary(employee, year, month):
    stats = {
        'days_worked': 0,
        'worked_minutes': 0,
        'overtime_minutes': 0,
        'night_minutes': 0,
        'sunday_minutes': 0,
        'holiday_minutes': 0,
    }
    try:
        presences = Presence.objects.filter(employee=employee, date__year=year, date__month=month)
        stats['days_worked'] = presences.filter(worked_minutes__gt=0).count()
        stats['worked_minutes'] = sum(p.worked_minutes or 0 for p in presences)
        stats['overtime_minutes'] = sum(p.overtime_minutes or 0 for p in presences)
        stats['night_minutes'] = sum(p.night_minutes or 0 for p in presences)
        stats['sunday_minutes'] = sum(p.sunday_minutes or 0 for p in presences)
        stats['holiday_minutes'] = sum(p.holiday_minutes or 0 for p in presences)
    except Exception:
        pass
    return stats


def _absence_days(employee, year, month):
    try:
        return Absence.objects.filter(employee=employee, date__year=year, date__month=month, justified=False).count()
    except Exception:
        return 0


def _compute_irsa_tranches(base_imposable):
    tranche_bounds = [350000, 400000, 500000, 600000, 4000000]
    tranche_rates = [0.00, 0.05, 0.10, 0.15, 0.20, 0.25]
    imposable = max(base_imposable or 0, 0)
    prev = 0
    values = []
    for i, bound in enumerate(tranche_bounds):
        low = prev
        high = bound
        portion = max(0, min(imposable, high) - low)
        rate = tranche_rates[i]
        values.append(int(round(portion * rate)))
        prev = bound
    last_rate = tranche_rates[len(tranche_bounds)]
    portion = max(0, imposable - prev)
    values.append(int(round(portion * last_rate)))
    total = sum(values)
    return values, total

def export_payroll_pdf(request, pk):
    try:
        payroll = Payroll.objects.get(pk=pk)
        employee = payroll.employee
    except Payroll.DoesNotExist:
        return HttpResponse('Paie introuvable', status=404)

    # Determine requested format: pdf (default), xlsx, or both
    req_format = request.GET.get('format', 'pdf').lower()

    # --- Build XLSX bytes inline (same mapping as export_employee_fiche) ---
    period_start = datetime(payroll.year, payroll.month, 1).date()
    period_end = (period_start + relativedelta(months=1)) - relativedelta(days=1)
    seniority_text = calculate_seniority(employee.hire_date)

    base_salary = float(payroll.salary_base or employee.salary_base or 0)
    breakdown = payroll.computed_breakdown or {}
    gross = float(breakdown.get('gross') or payroll.gross_salary or base_salary)
    cnaps_employee = float(breakdown.get('cnaps_employee') or PayrollCalculator.calculer_cnaps_salarie(gross))
    sanitaire = float(breakdown.get('ostie') or PayrollCalculator.calculer_sanitaire_salarie(gross))
    absence_deduction = abs(float(breakdown.get('absence_deduction') or 0))
    deductions = float(breakdown.get('deductions') or (cnaps_employee + sanitaire + absence_deduction))
    net = float(breakdown.get('net') or (gross - deductions))
    details = breakdown.get('details', {}) if isinstance(breakdown, dict) else {}
    overtime_amount = float(details.get('overtime_pay', 0))
    night_amount = float(details.get('night_premium', 0))
    sunday_amount = float(details.get('sunday_premium', 0))
    holiday_amount = float(details.get('holiday_premium', 0))
    autres_indemnites = float(details.get('allowances', 0))

    presence_stats = _presence_summary(employee, payroll.year, payroll.month)
    absence_days = _absence_days(employee, payroll.year, payroll.month)
    default_days = PayrollCalculator.get_jours_mensuels()
    days_worked = presence_stats['days_worked'] or getattr(payroll, 'days_worked', None) or default_days
    worked_hours = round((presence_stats['worked_minutes'] or 0) / 60.0, 2)
    overtime_hours = round((presence_stats['overtime_minutes'] or 0) / 60.0, 2)
    night_hours = round((presence_stats['night_minutes'] or 0) / 60.0, 2)
    sunday_hours = round((presence_stats['sunday_minutes'] or 0) / 60.0, 2)
    holiday_hours = round((presence_stats['holiday_minutes'] or 0) / 60.0, 2)

    daily_rate_contract = PayrollCalculator.calculer_taux_journalier(base_salary)
    hourly_rate = PayrollCalculator.calculer_taux_horaire(base_salary)
    indice = int(round(hourly_rate / 1.334, 0)) if hourly_rate else 0

    base_imposable = gross - cnaps_employee - sanitaire
    irsa_values, irsa_total = _compute_irsa_tranches(base_imposable)
    total_retenues = cnaps_employee + sanitaire + irsa_total
    montant_imposable = max(base_imposable, 0)

    # Load the template when available, otherwise create a fresh workbook
    base = Path(settings.BASE_DIR)
    template_paths = [
        base.parent / 'FICHE DE PAIE .xlsx',
        base.parent / 'exports' / 'FICHE DE PAIE .xlsx',
        base / 'FICHE DE PAIE .xlsx',
    ]
    template_path = None
    for p in template_paths:
        if p.exists():
            template_path = p
            break

    if template_path and template_path.exists():
        wb = openpyxl.load_workbook(filename=str(template_path))
    else:
        wb = openpyxl.Workbook()
    ws = wb.active
    if ws.title == 'Sheet':
        ws.title = 'FICHE'

    def write(cell, value, number_format=None):
        try:
            ws[cell] = value
            if number_format:
                ws[cell].number_format = number_format
        except Exception:
            pass

    def round_currency(val):
        return float(Decimal(str(val or 0)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))

    # Section 1: en-tête / ancienneté
    write('L1', 'nb jour/mois')
    write('U1', '=365/12')
    write('M2', 'Clôture')
    write('O2', 'Début')
    write('P2', period_start)
    write('S2', 'Ancienneté en jours')
    write('T2', 'Partie en années')
    write('U2', 'Partie en mois')
    write('V2', 'Partie en jours')
    write('W2', 'Ancienneté')
    write('M10', period_end)
    write('O10', 'Fin')
    write('P10', period_end)
    if employee.hire_date:
        anciennete_jours = (period_end - employee.hire_date).days + 1
    else:
        anciennete_jours = 0
    years = int(anciennete_jours / 365) if anciennete_jours else 0
    months = int((anciennete_jours - (years * 365)) / 30.4167) if anciennete_jours else 0
    days_rest = int(anciennete_jours - (years * 365) - (months * 30.4167)) if anciennete_jours else 0
    write('S10', anciennete_jours)
    write('T10', years)
    write('U10', months)
    write('V10', days_rest)
    write('W10', f"{years} an(s) {months} mois {days_rest} jour(s)")

    # Section 2: informations employé
    full_name = f"{employee.last_name} {employee.first_name}".strip()
    write('C12', 'FICHE DE PAIE')
    write('D14', 'Nom et Prénoms :')
    write('F14', full_name)
    write('D15', 'Matricule :')
    write('F15', employee.matricule or '')
    write('H15', 'Classification :')
    write('J15', str(employee.category) if employee.category else '')
    write('D16', 'Fonction :')
    write('F16', employee.function or '')
    write('H16', 'Salaire de base :')
    write('J16', round_currency(base_salary), '#,##0.00')
    write('D17', 'N° CNaPS :')
    write('F17', employee.cnaps_number or '')
    write('H17', 'Taux journaliers :')
    write('J17', round_currency(daily_rate_contract), '#,##0.00')
    write('D18', "Date d'embauche :")
    write('F18', employee.hire_date or '')
    write('H18', 'Taux horaires :')
    write('J18', round_currency(hourly_rate), '#,##0.00')
    write('D19', 'Ancienneté :')
    write('F19', seniority_text)
    write('H19', 'Indice :')
    write('J19', indice)

    # Section 3: base congés
    write('M16', round_currency(base_salary), '#,##0.00')
    write('M17', round_currency(base_salary / 30 if base_salary else 0))
    write('M18', round_currency(base_salary / 173.33 if base_salary else 0))
    write('M19', round_currency((base_salary / 30 if base_salary else 0) / 1.334 if base_salary else 0))

    # Section 4: tableau des gains
    period_text = f"Salaire du {period_start.strftime('%d/%m/%y')} au {period_end.strftime('%d/%m/%y')}"
    write('D23', period_text)
    effective_days = round(days_worked, 2) if isinstance(days_worked, (int, float)) else days_worked
    salaire_principal = round_currency(gross)
    salaire_daily_rate = round_currency((salaire_principal / days_worked) if days_worked else daily_rate_contract)
    write('H23', effective_days)
    write('I23', salaire_daily_rate, '#,##0.00')
    write('J23', salaire_principal, '#,##0.00')
    write('D24', 'Absences déductibles')
    absence_rate = round_currency(absence_deduction / absence_days) if absence_days else salaire_daily_rate
    write('H24', absence_days)
    write('I24', absence_rate, '#,##0.00')
    write('J24', -round_currency(absence_deduction), '#,##0.00')

    write('D27', 'Heures supplémentaires majorées de 30%')
    overtime_rate = round_currency(overtime_amount / overtime_hours) if overtime_hours else round_currency(hourly_rate * 1.3)
    write('H27', overtime_hours)
    write('I27', overtime_rate, '#,##0.00')
    write('J27', round_currency(overtime_amount), '#,##0.00')

    write('D28', 'Heures supplémentaires majorées de 40%')
    write('H28', 0)
    write('I28', round_currency(hourly_rate * 1.4), '#,##0.00')
    write('J28', 0)

    write('D29', 'Heures supplémentaires majorées de 50%')
    write('H29', 0)
    write('I29', round_currency(hourly_rate * 1.5), '#,##0.00')
    write('J29', 0)

    write('D30', 'Heures supplémentaires majorées de 100%')
    combined_hours = sunday_hours + holiday_hours
    combined_amount = round_currency(sunday_amount + holiday_amount)
    combined_rate = round_currency(combined_amount / combined_hours) if combined_hours else round_currency(hourly_rate * 2)
    write('H30', round(combined_hours, 2))
    write('I30', combined_rate, '#,##0.00')
    write('J30', combined_amount, '#,##0.00')

    write('D31', 'Majoration pour heures de nuit')
    night_rate = (
        round_currency(night_amount / night_hours)
        if night_hours
        else round_currency(hourly_rate * 2)
    )
    write('H31', night_hours)
    write('I31', night_rate, '#,##0.00')
    write('J31', round_currency(night_amount), '#,##0.00')

    write('D34', 'Droits de congés')
    write('H34', 0)
    write('I34', salaire_daily_rate, '#,##0.00')
    write('J34', 0)

    write('I38', 'Salaire brut')
    write('J38', round_currency(gross), '#,##0.00')

    # Section 5: retenues
    write('G40', 'Retenue CNaPS 1%')
    write('J40', round_currency(cnaps_employee), '#,##0.00')
    write('M40', round_currency(PayrollCalculator.PLAFOND_CNAPS * PayrollCalculator.TAUX_CNAPS_SALARIE), '#,##0.00')
    write('N40', 'Plafond')
    write('G41', 'Retenue sanitaire')
    write('J41', round_currency(sanitaire), '#,##0.00')
    irsa_labels = [
        'Tranche IRSA INF 350 000',
        'Tranche IRSA DE 350 001 à 400 000',
        'Tranche IRSA DE 400 001 à 500 000',
        'Tranche IRSA DE 500 001 à 600 000',
        'Tranche IRSA DE 600 001 à 4 000 000',
        'Tranche IRSA PLUS DE 4 000 000',
    ]
    for idx, label in enumerate(irsa_labels):
        row = 43 + idx
        write(f'G{row}', label)
        write(f'J{row}', irsa_values[idx], '#,##0')
    write('G49', 'TOTAL IRSA')
    write('J49', irsa_total, '#,##0')
    write('I51', 'Total des retenues')
    write('J51', round_currency(total_retenues), '#,##0.00')

    # Section 6: net à payer
    # Respect template layout: labels in column I, amounts in column J
    net_template_amount = round_currency(gross - total_retenues + autres_indemnites)
    write('I52', 'Autres indemnités')
    write('J52', round_currency(autres_indemnites), '#,##0.00')
    write('I53', '')
    write('J53', '')
    write('I54', 'Net à payer')
    write('J54', net_template_amount, '#,##0.00')

    # Section 7: informations complémentaires
    write('C57', 'Avantages en nature :')
    write('C58', 'Déductions IRSA :')
    write('C59', 'Montant imposable :')
    write('F59', round_currency(montant_imposable), '#,##0.00')
    write('D62', 'Mode de paiement :')
    write('F62', payroll.notes or '')
    write('D64', "L'employeur")
    write('H64', "L'employé(e)")

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    xlsx_bytes = bio.read()

    # If caller explicitly wants the XLSX, return it immediately
    if req_format == 'xlsx':
        filename = f"fiche_{employee.matricule or employee.pk}_{payroll.year}_{payroll.month}.xlsx"
        resp = HttpResponse(xlsx_bytes, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f'attachment; filename="{filename}"'
        return resp

    # Try XLSX -> PDF via LibreOffice first
    with tempfile.TemporaryDirectory() as td:
        xlsx_path = Path(td) / f"fiche_{employee.matricule or employee.pk}_{payroll.year}_{payroll.month}.xlsx"
        pdf_path = xlsx_path.with_suffix('.pdf')
        with open(xlsx_path, 'wb') as f:
            f.write(xlsx_bytes)

        soffice = shutil.which('soffice') or shutil.which('libreoffice')
        pdf_converted = False
        if soffice:
            cmd = [soffice, '--headless', '--convert-to', 'pdf', '--outdir', str(td), str(xlsx_path)]
            try:
                subprocess.run(cmd, check=True, timeout=60, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                pdf_converted = pdf_path.exists()
            except Exception as e:
                logging.exception('LibreOffice conversion failed: %s', e)
                pdf_converted = False

        # If user asked for both, return a ZIP with XLSX and PDF (PDF only if conversion succeeded)
        if req_format == 'both':
            from zipfile import ZipFile
            zip_path = Path(td) / f"fiche_{employee.matricule or employee.pk}_{payroll.year}_{payroll.month}.zip"
            with ZipFile(zip_path, 'w') as zf:
                zf.writestr(xlsx_path.name, xlsx_bytes)
                if pdf_converted and pdf_path.exists():
                    with open(pdf_path, 'rb') as f:
                        zf.writestr(pdf_path.name, f.read())
            with open(zip_path, 'rb') as f:
                zip_bytes = f.read()
            resp = HttpResponse(zip_bytes, content_type='application/zip')
            resp['Content-Disposition'] = f'attachment; filename="{zip_path.name}"'
            return resp

        # If conversion to PDF worked, return the PDF
        if pdf_converted and pdf_path.exists():
            with open(pdf_path, 'rb') as f:
                pdf_bytes = f.read()
            resp = HttpResponse(pdf_bytes, content_type='application/pdf')
            resp['Content-Disposition'] = f'attachment; filename="fiche_{employee.matricule or employee.pk}_{payroll.year}_{payroll.month}.pdf"'
            return resp

        # --- Fallback: try HTML -> PDF via WeasyPrint ---
        try:
            salaire_brut = float(payroll.gross_salary or (employee.salary_base or 0))
            calculs = PayrollCalculator.generer_fiche_paie_complete(salaire_brut)
            diag = {
                'salary_base': salaire_brut,
                'gross': calculs.get('salaire_brut', salaire_brut),
                'cnaps_employee': calculs.get('cnaps_salarie', 0),
                'ostie': calculs.get('sanitaire_salarie', 0),
                'irsa': calculs.get('irsa', 0),
                'total_retenues': calculs.get('total_retenues', 0),
                'net': calculs.get('salaire_net', calculs.get('salaire_brut', salaire_brut) - calculs.get('total_retenues', 0)),
                'hourly_rate': calculs.get('taux_horaire', 0),
                'daily_rate': calculs.get('taux_journalier', 0),
            }
            context = {
                'employee': employee,
                'payroll': payroll,
                'diag_payroll_breakdown': diag,
                'diag_month': payroll.month,
                'diag_year': payroll.year,
            }
            html = render_to_string('core/fiche_pdf.html', context=context, request=request)
            try:
                from weasyprint import HTML
                pdf_bytes = HTML(string=html, base_url=request.build_absolute_uri('/')).write_pdf()
                resp = HttpResponse(pdf_bytes, content_type='application/pdf')
                resp['Content-Disposition'] = f'attachment; filename="fiche_{employee.matricule or employee.pk}_{payroll.year}_{payroll.month}.pdf"'
                return resp
            except Exception as e:
                logging.exception('WeasyPrint conversion failed: %s', e)
        except Exception:
            logging.exception('Failed to render HTML fiche, will return XLSX as last resort')

        # If we reach here, both conversions failed (or weren't available). Return XLSX as last resort.
        filename = f"fiche_{employee.matricule or employee.pk}_{payroll.year}_{payroll.month}.xlsx"
        resp = HttpResponse(xlsx_bytes, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f'attachment; filename="{filename}"'
        return resp


def export_contract_pdf(request, pk):
    """Render a Contract as PDF. Uses WeasyPrint if available, otherwise returns the rendered HTML.

    This view is intentionally lightweight: it looks for a template
    `core/contract_pdf.html` and renders the contract context. If `weasyprint`
    is installed it will convert the HTML to PDF and return it as an attachment.
    If not available we simply return the HTML so the user can print/save from
    the browser.
    """
    try:
        contract = Contract.objects.get(pk=pk)
    except Contract.DoesNotExist:
        return HttpResponse('Contrat introuvable', status=404)

    context = {'contract': contract}
    html = render_to_string('core/contract_pdf.html', context)

    # Try WeasyPrint conversion when available
    try:
        from weasyprint import HTML
        pdf_bytes = HTML(string=html, base_url=request.build_absolute_uri('/')).write_pdf()
        resp = HttpResponse(pdf_bytes, content_type='application/pdf')
        resp['Content-Disposition'] = f'attachment; filename="contract_{contract.pk}.pdf"'
        return resp
    except Exception:
        # If WeasyPrint isn't installed or conversion fails, return HTML fallback
        return HttpResponse(html)

def import_payroll_upload(request):
    """Allow a user to upload an ETAT DE PAIE Excel file and run the import_payroll command.
    Supports dry-run (default) and apply (real import). Shows the command output as a report.
    """
    base = Path(settings.BASE_DIR)
    uploads = base / 'uploads'
    uploads.mkdir(exist_ok=True)

    report = None
    saved_name = None
    if request.method == 'POST':
        action = request.POST.get('action', 'dry-run')
        f = request.FILES.get('file')
        if not f:
            return render(request, 'core/import_payroll.html', {'error': 'Aucun fichier envoyé'})

        # save uploaded file
        filename = f"{uuid.uuid4().hex}_{f.name}"
        dest = uploads / filename
        with open(dest, 'wb') as out:
            for chunk in f.chunks():
                out.write(chunk)
        saved_name = str(dest.name)

        # run management command with dry-run or real
        buf = StringIO()
        # call_command expects payroll_file relative to BASE_DIR as in the management command implementation
        relpath = os.path.join('uploads', dest.name)
        try:
            if action == 'apply':
                call_command('import_payroll', payroll_file=relpath, dry_run=False, stdout=buf)
            else:
                call_command('import_payroll', payroll_file=relpath, dry_run=True, stdout=buf)
        except Exception as e:
            buf.write(f"Erreur lors de l'import: {e}\n")
        report = buf.getvalue()

    return render(request, 'core/import_payroll.html', {'report': report, 'saved_name': saved_name})
    


# CRUD web views for Employee and Leave
class EmployeeCreateView(CreateView):
    model = Employee
    fields = ['matricule', 'email', 'first_name', 'last_name', 'category', 'function', 'hire_date', 'salary_base']
    template_name = 'core/employee_form.html'
    success_url = reverse_lazy('employee_list')
    def form_valid(self, form):
        # save then trigger the employees fiche update (signals also handle this but keep explicit)
        resp = super().form_valid(form)
        try:
            write_employees_to_fiche()
        except Exception:
            pass
        return resp


class EmployeeUpdateView(UpdateView):
    model = Employee
    fields = ['matricule', 'email', 'first_name', 'last_name', 'category', 'function', 'hire_date', 'salary_base']
    template_name = 'core/employee_form.html'
    success_url = reverse_lazy('employee_list')
    def form_valid(self, form):
        resp = super().form_valid(form)
        try:
            write_employees_to_fiche()
        except Exception:
            pass
        return resp


class EmployeeDeleteView(DeleteView):
    model = Employee
    template_name = 'core/employee_confirm_delete.html'
    success_url = reverse_lazy('employee_list')


class LeaveCreateView(CreateView):
    model = Leave
    fields = ['employee', 'start_date', 'end_date', 'leave_type', 'status', 'note']
    template_name = 'core/leave_form.html'
    success_url = reverse_lazy('leave_list')


class LeaveUpdateView(UpdateView):
    model = Leave
    fields = ['employee', 'start_date', 'end_date', 'leave_type', 'status', 'note']
    template_name = 'core/leave_form.html'
    success_url = reverse_lazy('leave_list')


class LeaveDeleteView(DeleteView):
    model = Leave
    template_name = 'core/leave_confirm_delete.html'
    success_url = reverse_lazy('leave_list')


def stats_workforce_by_gender(request):
    """Return headcount by gender (from view workforce_by_gender)"""
    with connection.cursor() as cur:
        cur.execute("SELECT gender_code, headcount FROM workforce_by_gender")
        rows = cur.fetchall()
    data = [{'gender': r[0], 'headcount': int(r[1])} for r in rows]
    return JsonResponse({'data': data})


def stats_avg_age(request):
    with connection.cursor() as cur:
        cur.execute("SELECT average_age FROM avg_age")
        row = cur.fetchone()
    return JsonResponse({'average_age': float(row[0]) if row and row[0] is not None else None})


def stats_avg_seniority(request):
    with connection.cursor() as cur:
        cur.execute("SELECT average_seniority FROM avg_seniority")
        row = cur.fetchone()
    return JsonResponse({'average_seniority': float(row[0]) if row and row[0] is not None else None})


def stats_turnover(request):
    with connection.cursor() as cur:
        cur.execute("SELECT terminations_12m, hires_12m, current_headcount FROM turnover_12m")
        row = cur.fetchone()
    if not row:
        return JsonResponse({'term': 0, 'hires': 0, 'headcount': 0})
    return JsonResponse({'terminations_12m': int(row[0]), 'hires_12m': int(row[1]), 'current_headcount': int(row[2])})


def stats_absenteeism_monthly(request):
    try:
        with connection.cursor() as cur:
            cur.execute("SELECT month, absence_days, headcount FROM absenteeism_monthly ORDER BY month DESC LIMIT 12")
            rows = cur.fetchall()
        data = []
        for r in rows:
            month_val = r[0]
            # if month is a date-like object, convert to isoformat, otherwise stringify
            if hasattr(month_val, 'isoformat'):
                month_str = month_val.isoformat()
            else:
                month_str = str(month_val)
            data.append({
                'month': month_str,
                'absence_days': int(r[1]) if r[1] is not None else 0,
                'headcount': int(r[2]) if r[2] is not None else 0,
            })
        return JsonResponse({'data': data})
    except Exception:
        # If the DB view/table isn't available or the query fails, return an empty dataset
        return JsonResponse({'data': []})


def stats_unused_leave_summary(request):
    """Return per-employee unused leave summary.
    This attempts to read from a DB view `unused_leave_summary`. If the view is not present
    the endpoint will return an empty list (defensive fallback to avoid import-time errors).
    Expected columns from the view: employee_id, matricule, first_name, last_name,
    total_allocated, used_days, unused_days
    """
    try:
        with connection.cursor() as cur:
            cur.execute(
                "SELECT employee_id, matricule, first_name, last_name, total_allocated, used_days, unused_days FROM unused_leave_summary"
            )
            rows = cur.fetchall()
        data = []
        for r in rows:
            data.append({
                'employee_id': int(r[0]) if r[0] is not None else None,
                'matricule': r[1] or '',
                'first_name': r[2] or '',
                'last_name': r[3] or '',
                'total_allocated': float(r[4]) if r[4] is not None else 0.0,
                'used_days': float(r[5]) if r[5] is not None else 0.0,
                'unused_days': float(r[6]) if r[6] is not None else 0.0,
            })
        return JsonResponse({'data': data})
    except Exception:
        return JsonResponse({'data': []})


def rh_dashboard_view(request):
    """Public HR dashboard with instant KPIs and daily status pie chart."""
    today = _local_today()
    month_start = today.replace(day=1)
    month_days_elapsed = (today - month_start).days + 1
    active_qs = Employee.objects.filter(is_active=True, archived=False)
    active_count = active_qs.count()

    gender_rows = {row['gender']: row['total'] for row in active_qs.values('gender').annotate(total=Count('id'))}
    male_pct = round((gender_rows.get('M', 0) / active_count) * 100, 1) if active_count else 0.0
    female_pct = round((gender_rows.get('F', 0) / active_count) * 100, 1) if active_count else 0.0

    hire_dates = list(active_qs.exclude(hire_date__isnull=True).values_list('hire_date', flat=True))
    avg_days = 0
    if hire_dates:
        total_days = sum(max((today - hd).days, 0) for hd in hire_dates if hd)
        avg_days = total_days / len(hire_dates)
    avg_years = avg_days // 365 if avg_days else 0
    avg_months = (avg_days % 365) // 30 if avg_days else 0
    seniority_display = f"{int(avg_years)} an{'s' if avg_years >= 2 else ''} {int(avg_months)} mois" if avg_days else '—'

    month_absences = Absence.objects.filter(date__gte=month_start, date__lte=today).count()
    potential_presence = max(1, active_count * max(1, month_days_elapsed))
    absence_rate = round((month_absences / potential_presence) * 100, 2) if potential_presence else 0.0

    year_start = date(today.year, 1, 1)
    headcount_start = Employee.objects.filter(hire_date__lt=year_start, archived=False).count()
    terminations = Contract.objects.filter(date_end__gte=year_start, date_end__lte=today).count()
    avg_headcount = max(1, (headcount_start + active_count) / 2)
    turnover_rate = round((terminations / avg_headcount) * 100, 2) if avg_headcount else 0.0

    limit_date = today + timedelta(days=30)
    expiring_qs = Contract.objects.filter(active=True, date_end__isnull=False, date_end__gte=today, date_end__lte=limit_date).select_related('employee').order_by('date_end')
    expiring_contracts = expiring_qs.count()
    expiring_contract_rows = [
        {
            'employee': contract.employee,
            'date_end': contract.date_end,
            'type': contract.get_type_display() if hasattr(contract, 'get_type_display') else contract.type,
            'days_left': (contract.date_end - today).days if contract.date_end else None,
        }
        for contract in expiring_qs[:6]
    ]

    balances = LeaveBalance.objects.filter(employee__in=active_qs, year__gte=today.year - 2)
    balance_rows = balances.values('employee').annotate(entitlement=Sum('entitlement_days'), used=Sum('used_days'))
    superstock_map = {}
    for row in balance_rows:
        ent = float(row['entitlement'] or 0)
        used = float(row['used'] or 0)
        available = ent - used
        if available > 60:
            superstock_map[row['employee']] = available
    leave_superstock_count = len(superstock_map)
    superstock_rows = []
    if superstock_map:
        employees_superstock = Employee.objects.filter(id__in=superstock_map.keys()).order_by('last_name')[:6]
        for emp in employees_superstock:
            superstock_rows.append({
                'employee': emp,
                'available_days': round(superstock_map.get(emp.id, 0), 1),
            })

    payroll_list = list(Payroll.objects.filter(year=today.year, month=today.month).select_related('employee'))
    gross_total = sum(float(p.gross_salary or 0) for p in payroll_list)
    net_total = sum(float(p.net_salary or 0) for p in payroll_list)
    cnaps_ostie_total = 0.0
    for payroll in payroll_list:
        breakdown = payroll.computed_breakdown or {}
        cnaps_val = breakdown.get('cnaps_employee')
        ostie_val = breakdown.get('ostie')
        if cnaps_val is None and payroll.gross_salary:
            cnaps_val = PayrollCalculator.calculer_cnaps_salarie(float(payroll.gross_salary))
        if ostie_val is None and payroll.gross_salary:
            ostie_val = PayrollCalculator.calculer_sanitaire_salarie(float(payroll.gross_salary))
        cnaps_ostie_total += float(cnaps_val or 0) + float(ostie_val or 0)

    statuses = _build_daily_status(today)
    status_summary = []
    status_urls = {}
    for state_key, data in statuses.items():
        url = reverse('rh_dashboard_state', args=[state_key])
        status_summary.append({
            'state': state_key,
            'label': data['label'],
            'color': data['color'],
            'count': data['count'],
            'url': url,
        })
        status_urls[state_key] = url

    metrics = {
        'active_employees': active_count,
        'gender': {'male': male_pct, 'female': female_pct},
        'avg_seniority_display': seniority_display,
        'absence_rate': absence_rate,
        'turnover_rate': turnover_rate,
        'contracts_expiring': expiring_contracts,
        'leave_superstock_count': leave_superstock_count,
        'gross_total': gross_total,
        'net_total': net_total,
        'cnaps_ostie_total': cnaps_ostie_total,
    }

    kpi_cards = [
        {'label': 'Employés actifs', 'value': metrics['active_employees'], 'subtext': 'Salariés en poste'},
        {'label': '% Hommes – Femmes', 'value': f"{male_pct:.1f}% · {female_pct:.1f}%", 'subtext': 'Répartition du personnel'},
        {'label': 'Ancienneté moyenne', 'value': seniority_display, 'subtext': f"{int(avg_years)} ans en moyenne" if avg_days else 'Aucune donnée'},
        {'label': 'Taux d’absentéisme', 'value': _format_percentage(absence_rate, digits=2), 'subtext': f"{month_absences} jour(s) dans le mois"},
        {'label': 'Turnover annuel', 'value': _format_percentage(turnover_rate, digits=2), 'subtext': f"{terminations} départ(s) en {today.year}"},
        {'label': 'Contrats < 30 jours', 'value': expiring_contracts, 'subtext': 'Actions à planifier'},
        {'label': 'Congés > 60 jours', 'value': leave_superstock_count, 'subtext': 'Employés à relancer'},
        {'label': 'Salaire brut du mois', 'value': f"{_format_currency(gross_total)} Ar"},
        {'label': 'Salaire net du mois', 'value': f"{_format_currency(net_total)} Ar"},
        {'label': 'CNAPS + OSTIE', 'value': f"{_format_currency(cnaps_ostie_total)} Ar", 'subtext': 'Cotisations salariales'},
    ]

    context = {
        'today': today,
        'kpi_cards': kpi_cards,
        'status_summary': status_summary,
        'pie_labels_json': json.dumps([item['label'] for item in status_summary]),
        'pie_values_json': json.dumps([item['count'] for item in status_summary]),
        'pie_colors_json': json.dumps([item['color'] for item in status_summary]),
        'status_state_order_json': json.dumps([item['state'] for item in status_summary]),
        'status_urls_json': json.dumps(status_urls),
    'expiring_contract_rows': expiring_contract_rows,
        'superstock_employees': superstock_rows,
        'metrics': metrics,
        'page_title': 'Tableau de bord RH',
        'month_label': MONTH_LABELS[today.month],
    }
    return render(request, 'core/rh_dashboard.html', context)


def rh_dashboard_state_view(request, state):
    today = _local_today()
    statuses = _build_daily_status(today)
    if state not in statuses:
        raise Http404('État inconnu')

    status_summary = []
    status_urls = {}
    for state_key, data in statuses.items():
        url = reverse('rh_dashboard_state', args=[state_key])
        status_summary.append({
            'state': state_key,
            'label': data['label'],
            'color': data['color'],
            'count': data['count'],
            'url': url,
        })
        status_urls[state_key] = url

    context = {
        'today': today,
        'state_key': state,
        'state_data': statuses[state],
        'status_summary': status_summary,
        'status_urls': status_urls,
    }
    return render(request, 'core/rh_dashboard_status_list.html', context)


def home(request):
    """Simple dashboard / front page showing key HR indicators and quick links."""
    today = _local_today()
    try:
        selected_year = int(request.GET.get('year', today.year))
    except (TypeError, ValueError):
        selected_year = today.year
    try:
        selected_month = int(request.GET.get('month', today.month))
    except (TypeError, ValueError):
        selected_month = today.month
    if selected_month not in range(1, 13):
        selected_month = today.month

    year_choices = list(range(today.year - 4, today.year + 1))
    if selected_year not in year_choices:
        year_choices.append(selected_year)
    year_choices = sorted(set(y for y in year_choices if y > 2000))
    month_choices = [{'value': i, 'label': MONTH_LABELS[i]} for i in range(1, 13)]

    # Employee selector for drill-down view
    employee_choices = []
    for emp in Employee.objects.order_by('last_name', 'first_name'):
        display = f"{emp.matricule or ''} - {emp.last_name} {emp.first_name}".strip('- ')
        employee_choices.append({'id': emp.id, 'label': display})

    selected_employee_id = request.GET.get('employee')
    employee_selected = None
    try:
        if selected_employee_id:
            employee_selected = Employee.objects.get(pk=int(selected_employee_id))
            selected_employee_id = employee_selected.id
    except (Employee.DoesNotExist, ValueError, TypeError):
        selected_employee_id = None
        employee_selected = None

    try:
        employee_year = int(request.GET.get('employee_year', selected_year))
    except (TypeError, ValueError):
        employee_year = selected_year
    try:
        employee_month = int(request.GET.get('employee_month', selected_month))
    except (TypeError, ValueError):
        employee_month = selected_month
    if employee_month not in range(1, 13):
        employee_month = selected_month
    if employee_year < 2000:
        employee_year = selected_year

    employees_count = Employee.objects.count()
    contracts_active = Contract.objects.filter(active=True).count()
    pending_leaves = Leave.objects.filter(status='PENDING').count()
    alerts_open = Alerte.objects.filter(statut='OPEN').count()

    # Try to get last payroll month/year
    last_payroll = Payroll.objects.order_by('-year', '-month').first()
    last_payroll_label = f"{last_payroll.month}/{last_payroll.year}" if last_payroll else '—'

    # workforce by gender (if view exists) - safe fallback
    gender_data = []
    try:
        with connection.cursor() as cur:
            cur.execute("SELECT gender_code, headcount FROM workforce_by_gender")
            rows = cur.fetchall()
        gender_data = [{'gender': r[0], 'headcount': int(r[1])} for r in rows]
    except Exception:
        gender_data = []

    # Monthly focus metrics (absences / leaves / late)
    absence_qs = Absence.objects.filter(date__year=selected_year, date__month=selected_month)
    monthly_absences = absence_qs.count()
    monthly_absences_unjustified = absence_qs.filter(justified=False).count()
    monthly_absences_justified = monthly_absences - monthly_absences_unjustified

    leaves_qs = Leave.objects.filter(start_date__year=selected_year, start_date__month=selected_month)
    monthly_leaves_total = leaves_qs.count()
    monthly_leaves_pending = leaves_qs.filter(status='PENDING').count()
    monthly_leaves_approved = leaves_qs.filter(status='APPROVED').count()

    late_threshold = getattr(settings, 'HR_PAYROLL', {}).get('LATE_THRESHOLD', 0)
    presence_filter = {
        'date__year': selected_year,
        'date__month': selected_month,
    }
    presence_qs = Presence.objects.filter(**presence_filter)
    if late_threshold:
        late_qs = presence_qs.filter(minutes_late__gte=late_threshold)
    else:
        late_qs = presence_qs.filter(minutes_late__gt=0)
    monthly_lates = late_qs.count()
    monthly_late_minutes = late_qs.aggregate(total=Sum('minutes_late')).get('total') or 0

    # Yearly overview for charts/list
    yearly_overview = []
    for month_idx in range(1, 13):
        abs_count = Absence.objects.filter(date__year=selected_year, date__month=month_idx).count()
        leave_count = Leave.objects.filter(start_date__year=selected_year, start_date__month=month_idx).count()
        if late_threshold:
            late_count = Presence.objects.filter(date__year=selected_year, date__month=month_idx, minutes_late__gte=late_threshold).count()
        else:
            late_count = Presence.objects.filter(date__year=selected_year, date__month=month_idx, minutes_late__gt=0).count()
        yearly_overview.append({
            'label': MONTH_LABELS[month_idx],
            'short': MONTH_LABELS[month_idx][:3],
            'absences': abs_count,
            'leaves': leave_count,
            'late': late_count,
        })

    employee_absences_detail = []
    employee_leaves_detail = []
    employee_lates_detail = []
    employee_payroll_history = []
    employee_month_summary = None

    if employee_selected:
        abs_qs = Absence.objects.filter(
            employee=employee_selected,
            date__year=employee_year,
            date__month=employee_month,
        ).order_by('date')
        employee_absences_detail = [
            {
                'date': a.date,
                'reason': a.reason or '',
                'justified': a.justified,
            }
            for a in abs_qs
        ]

        leaves_detail_qs = Leave.objects.filter(
            employee=employee_selected,
            start_date__year=employee_year,
            start_date__month=employee_month,
        ).order_by('start_date')
        employee_leaves_detail = [
            {
                'type': l.leave_type,
                'status': l.status,
                'start_date': l.start_date,
                'end_date': l.end_date,
                'days': l.days,
            }
            for l in leaves_detail_qs
        ]

        lates_detail_qs = Presence.objects.filter(
            employee=employee_selected,
            date__year=employee_year,
            date__month=employee_month,
            minutes_late__gt=0,
        ).order_by('date')
        employee_lates_detail = [
            {
                'date': p.date,
                'minutes_late': p.minutes_late,
                'worked_minutes': p.worked_minutes,
            }
            for p in lates_detail_qs
        ]

        total_worked_minutes = sum(item['worked_minutes'] or 0 for item in employee_lates_detail)
        total_absences = len(employee_absences_detail)
        total_leaves = len(employee_leaves_detail)
        summary_lates = sum(item['minutes_late'] or 0 for item in employee_lates_detail)
        employee_month_summary = {
            'absences': total_absences,
            'leaves': total_leaves,
            'late_occurrences': len(employee_lates_detail),
            'late_minutes': summary_lates,
            'worked_hours': round(total_worked_minutes / 60, 2) if total_worked_minutes else 0,
        }

        def _month_label(value):
            try:
                if 1 <= value < len(MONTH_LABELS):
                    return MONTH_LABELS[value]
            except Exception:
                pass
            return str(value)

        employee_payroll_history = [
            {
                'year': p.year,
                'month': p.month,
                'month_label': _month_label(p.month),
                'gross': float(p.gross_salary or 0),
                'net': float(p.net_salary or 0),
                'status': 'Payé' if float(p.net_salary or 0) > 0 else 'À régulariser',
                'etat': p.etat_paie,
                'created_at': p.created_at,
            }
            for p in Payroll.objects.filter(employee=employee_selected).order_by('year', 'month')
        ]

    context = {
        'employees_count': employees_count,
        'contracts_active': contracts_active,
        'pending_leaves': pending_leaves,
        'alerts_open': alerts_open,
        'last_payroll': last_payroll_label,
        'gender_data': gender_data,
        'selected_year': selected_year,
        'selected_month': selected_month,
        'selected_month_label': MONTH_LABELS[selected_month],
        'year_choices': year_choices,
        'month_choices': month_choices,
        'monthly_absences': monthly_absences,
        'monthly_absences_unjustified': monthly_absences_unjustified,
        'monthly_absences_justified': monthly_absences_justified,
        'monthly_leaves_total': monthly_leaves_total,
        'monthly_leaves_pending': monthly_leaves_pending,
        'monthly_leaves_approved': monthly_leaves_approved,
        'monthly_lates': monthly_lates,
        'monthly_late_minutes': int(monthly_late_minutes),
        'yearly_overview': yearly_overview,
        'employee_choices': employee_choices,
        'selected_employee_id': selected_employee_id,
        'employee_selected': employee_selected,
        'employee_month': employee_month,
        'employee_year': employee_year,
        'employee_month_label': MONTH_LABELS[employee_month],
        'employee_absences_detail': employee_absences_detail,
        'employee_leaves_detail': employee_leaves_detail,
        'employee_lates_detail': employee_lates_detail,
        'employee_month_summary': employee_month_summary,
        'employee_payroll_history': employee_payroll_history,
    }
    return render(request, 'core/home.html', context)


def user_is_manager(user):
    """Return True if user is in Manager group or is_staff."""
    if not user or not user.is_authenticated:
        return False
    if user.is_staff:
        return True
    try:
        return user.groups.filter(name='Manager').exists()
    except Exception:
        return False


@login_required
def leave_approval_page(request, pk):
    """Show a simple approval page for a manager to approve/reject a leave."""
    try:
        leave = Leave.objects.select_related('employee').get(pk=pk)
    except Leave.DoesNotExist:
        return HttpResponse('Demande introuvable', status=404)

    if not user_is_manager(request.user):
        return HttpResponse('Accès refusé', status=403)

    return render(request, 'core/leave_approval.html', {'leave': leave})


@login_required
@require_POST
def approve_leave(request, pk):
    """Handle approve/reject POST from manager. Expects 'action' in POST: 'approve' or 'reject'."""
    try:
        leave = Leave.objects.select_related('employee').get(pk=pk)
    except Leave.DoesNotExist:
        return HttpResponse('Demande introuvable', status=404)

    if not user_is_manager(request.user):
        return HttpResponse('Accès refusé', status=403)

    action = request.POST.get('action')
    note = request.POST.get('note', '')
    if action == 'approve':
        leave.status = 'APPROVED'
        leave.save()
        LeaveHistory.objects.create(employee=leave.employee, leave=leave, action='APPROVED', details=f'Approved by {request.user} {note}')
        # optional email notification
        try:
            send_mail(
                subject=f'Demande de congé approuvée - {leave.employee}',
                message=f'Votre demande de congé du {leave.start_date} au {leave.end_date} a été approuvée.',
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[getattr(leave.employee, 'email', '')] if getattr(leave.employee, 'email', None) else [],
                fail_silently=True,
            )
        except Exception:
            pass
    elif action == 'reject':
        leave.status = 'REJECTED'
        leave.save()
        LeaveHistory.objects.create(employee=leave.employee, leave=leave, action='REJECTED', details=f'Rejected by {request.user} {note}')
        try:
            send_mail(
                subject=f'Demande de congé refusée - {leave.employee}',
                message=f'Votre demande de congé du {leave.start_date} au {leave.end_date} a été refusée.',
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[getattr(leave.employee, 'email', '')] if getattr(leave.employee, 'email', None) else [],
                fail_silently=True,
            )
        except Exception:
            pass
    else:
        return HttpResponse('Action invalide', status=400)

    # If AJAX request, return JSON so frontend can update without reload
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest'
    if is_ajax:
        return JsonResponse({'result': 'ok', 'leave_id': leave.pk, 'status': leave.status})
    return redirect('leave_list')



@login_required
@require_POST
def approve_suggestion(request, pk):
    """Allow an HR/manager user to approve a SuggestedReplacement.

    POST -> approves the suggestion by calling suggestion.approve(request.user).
    Requires HR or Manager privileges (is_staff or member of 'HR'/'Manager' groups).
    """
    try:
        suggestion = SuggestedReplacement.objects.get(pk=pk)
    except SuggestedReplacement.DoesNotExist:
        return HttpResponse('Suggestion introuvable', status=404)

    user = request.user
    # simple RH check: staff or in HR/Manager groups
    try:
        allowed = user.is_staff or user.groups.filter(name__in=['HR', 'Manager']).exists()
    except Exception:
        allowed = bool(user.is_staff)

    if not allowed:
        return HttpResponse('Accès refusé', status=403)

    try:
        suggestion.approve(user)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

    # return the updated suggestion
    try:
        serializer = SuggestedReplacementSerializer(suggestion)
        return JsonResponse({'result': 'ok', 'suggestion': serializer.data}, status=200)
    except Exception:
        return JsonResponse({'result': 'ok'}, status=200)

