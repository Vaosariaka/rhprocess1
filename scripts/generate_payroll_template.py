#!/usr/bin/env python3
"""
Generate a payroll template XLSX with formulas based on the provided spec.
Creates: exports/FICHE_DE_PAIE.xlsx
"""
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers
import json
import os

OUT = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'exports', 'FICHE_DE_PAIE.xlsx')

wb = Workbook()
# Main sheet
ws = wb.active
ws.title = 'FICHE'

# Ensure columns A..W exist and set widths
for i in range(1, 24):
    ws.column_dimensions[get_column_letter(i)].width = 14

# --- Lignes 1-2 : métadonnées et calculs de base ---
ws['A1'] = 'FICHE DE PAIE'
ws['A2'] = 'nb jours/mois'
ws['B2'] = '=365/12'

# --- Lignes 10 : ancienneté area (use exact cells referenced in spec) ---
ws['A10'] = 'Ancienneté (jours)'
ws['B10'] = '=IF(AND(ISNUMBER(N10),ISNUMBER(B20)),N10-B20+1,"")'
ws['C10'] = 'Années'
ws['D10'] = '=IF(B10="","",INT(B10/365))'
ws['E10'] = 'Mois'
ws['F10'] = 'Date début période'
ws['G10'] = 'Mois (calc)'
ws['H10'] = '=IF(B10="","",INT(MOD(B10,365)/30))'
ws['I10'] = 'Jours restants'
ws['J10'] = '=IF(B10="","",MOD(B10,30))'

# --- Lignes 16-22 : informations employé ---
ws['A16'] = 'Nom et Prénoms'
ws['B16'] = 'NOM PRENOM'
ws['A17'] = 'Matricule'
ws['B17'] = 'MATRICULE'
ws['A18'] = 'Fonction'
ws['B18'] = 'FONCTION'
ws['A19'] = 'N° CNaPS'
ws['B19'] = 0
ws['A20'] = 'Date embauche'
ws['B20'] = '2011-03-25'
ws['A21'] = 'Classification'
ws['B21'] = 'CL'
ws['A22'] = 'Mode paiement'
ws['B22'] = 'Virement'

# --- Lignes 23-37 : Désignations (calculs salariaux) ---
# Place salary base in J16 per spec
ws['J16'] = 7800000
ws['I23'] = 'Designation'
ws['J23'] = 'Valeur'
ws['A23'] = 'Salaire du [date début] au [date fin]'
ws['H23'] = '1 mois'  # indicator
ws['J23'] = '=IF(H23="1 mois",J16,I23*H23)'

# Taux journalier / horaire on lines per spec
ws['A24'] = 'Taux journalier'
ws['J24'] = '=ROUND(J16/30,0)'
ws['A25'] = 'Taux horaire'
ws['J25'] = '=ROUND(J16/173.33,0)'

# Example heures sup lines (30%,50%,100%)
ws['A26'] = 'Heures sup 30%'
ws['H26'] = 0
ws['J26'] = '=ROUND(J25*1.3*H26,0)'
ws['A27'] = 'Heures sup 50%'
ws['H27'] = 0
ws['J27'] = '=ROUND(J25*1.5*H27,0)'
ws['A28'] = 'Heures sup 100%'
ws['H28'] = 0
ws['J28'] = '=ROUND(J25*2.0*H28,0)'

# Majoration nuit (22h-5h) as example
ws['A29'] = 'Majoration nuit 30%'
ws['H29'] = 0
ws['J29'] = '=ROUND(J25*0.3*H29,0)'

# Droits congés
ws['A34'] = 'Droits congés (jours)'
ws['H34'] = 0
ws['J34'] = '=H34*J24'

# Salaire brut (sum area J23:J37)
ws['A38'] = 'Salaire brut'
ws['J38'] = '=SUM(J23:J37)'

# --- Lignes 39-51 : Retenues ---
ws['A40'] = 'CNaPS 1% (employé)'
# M40 is the CNAPS cap (plafond)
ws['M40'] = 2800000
ws['J40'] = '=ROUND(MIN(J38*0.01,M40),0)'

ws['A41'] = 'Sanitaire 1%'
ws['J41'] = '=ROUND(J38*0.01,0)'

ws['A49'] = 'IRSA'
# IRSA progressive single-line formula
ws['J49'] = '=ROUND((MAX(0,MIN(J38-350000,50000))*0.05)+(MAX(0,MIN(J38-400000,100000))*0.10)+(MAX(0,MIN(J38-500000,100000))*0.15)+(MAX(0,MIN(J38-600000,3400000))*0.20)+(MAX(0,J38-4000000)*0.25),0)'

ws['A51'] = 'Total retenues'
ws['J51'] = '=SUM(J40,J41,J49)'

# Indemnities placeholder
ws['A52'] = 'Indemnites / Avantages'
ws['J52'] = 0

# Net a payer
ws['A53'] = 'Net a payer'
ws['J53'] = '=J38-J51+J52'

# Signatures / payment info
ws['A55'] = 'Mode paiement'
ws['B55'] = ws['B22'].value
ws['A56'] = 'Signature Employeur'
ws['A57'] = 'Signature Salarié'

# Apply number formats
num_cells = ['J16','J23','J24','J25','J26','J27','J28','J29','J34','J38','J40','J41','J49','J51','J52','J53']
for cell in num_cells:
    try:
        ws[cell].number_format = numbers.FORMAT_NUMBER
    except Exception:
        pass

# Create _META sheet
meta = wb.create_sheet('_META')
vars_example = {
    'NOM': ws['B16'].value,
    'MATRICULE': ws['B17'].value,
    'FONCTION': ws['B18'].value,
    'CNAPS_NUM': ws['B19'].value,
    'DATE_EMBAUCHE': ws['B20'].value,
    'DATE_DEBUT': ws['F10'].value,
    'DATE_FIN': ws['N10'].value if 'N10' in ws else '',
    'SALAIRE_BASE': ws['J16'].value,
}
meta['A1'] = 'variables_json'
meta['A2'] = json.dumps(vars_example)
meta['A4'] = 'generation_date'
import datetime
meta['B4'] = datetime.datetime.utcnow().isoformat() + 'Z'
meta['A5'] = 'generator_version'
meta['B5'] = 'script-openpyxl-v2'

# Save file
os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
print('Wrote', OUT)
